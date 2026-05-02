import streamlit as st
import PyPDF2
import pdfplumber
import re
import pandas as pd
from collections import Counter, defaultdict
import io
import plotly.express as px
import plotly.graph_objects as go
import calendar
from datetime import datetime
import hashlib
import json
import os
import time
import gc
import traceback
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas as rl_canvas
# Importar openpyxl explícitamente para exportación Excel
import openpyxl
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import urllib.request

# === SECCIÓN 1: IMPORTS ===
# (se mantienen explícitos para compatibilidad HF Spaces y evitar regresiones visuales)

# === SECCIÓN 2: CONFIGURACIÓN Y CONSTANTES ===
# --- 1. CONFIGURACIÓN DE LA PÁGINA ---
st.set_page_config(page_title="Analizador de Programaciones", page_icon="✈️", layout="wide")

# --- ORDEN DE BASES OFICIAL ---
BASES_ORDENADAS = [
    ('BCN', 'Barcelona (BCN)'),
    ('ALC', 'Alicante (ALC)'),
    ('BIO', 'Bilbao (BIO)'),
    ('LPA', 'Gran Canaria (LPA)'),
    ('IBZ', 'Ibiza (IBZ)'),
    ('AGP', 'Málaga (AGP)'),
    ('PMI', 'Palma de Mallorca (PMI)'),
    ('SCQ', 'Santiago de Compostela (SCQ)'),
    ('SVQ', 'Sevilla (SVQ)'),
    ('TFN', 'Tenerife Norte (TFN)'),
    ('VLC', 'Valencia (VLC)'),
]

# === SECCIÓN 3: AUTENTICACIÓN ===
# --- 1b. SISTEMA DE AUTENTICACIÓN ---
USERS = {
    'admin': hashlib.sha256('Master123'.encode()).hexdigest(),
    'laura.parra': hashlib.sha256('Elecciones2025'.encode()).hexdigest(),
    'david.llopis': hashlib.sha256('Elecciones2025'.encode()).hexdigest(),
    'neus.bofill': hashlib.sha256('Elecciones2025'.encode()).hexdigest(),
    'marian.calvo': hashlib.sha256('Elecciones2025'.encode()).hexdigest(),
    'marta.lestayo': hashlib.sha256('Elecciones2025'.encode()).hexdigest(),
    # Nuevos usuarios añadidos
    'veronica.huerta': hashlib.sha256('23041976'.encode()).hexdigest(),
    'carmen.pricop': hashlib.sha256('13091980'.encode()).hexdigest(),
    'ana.raya': hashlib.sha256('22101982'.encode()).hexdigest(),
    'iker.rodriguez': hashlib.sha256('15051983'.encode()).hexdigest(),
    'jose.villalta': hashlib.sha256('21121975'.encode()).hexdigest(),
    'rocio.muñoz': hashlib.sha256('27051983'.encode()).hexdigest(),
    'alberto.roldan': hashlib.sha256('02121984'.encode()).hexdigest(),
    'alina.simona': hashlib.sha256('17121983'.encode()).hexdigest(),
    'lara.aguado': hashlib.sha256('16121992'.encode()).hexdigest(),
    'nerea.millan': hashlib.sha256('18041987'.encode()).hexdigest(),
    'roberto.diaz': hashlib.sha256('18021981'.encode()).hexdigest(),
    'andrea.hernandez': hashlib.sha256('19031991'.encode()).hexdigest(),
    'carlos.arenas': hashlib.sha256('09121983'.encode()).hexdigest(),
    'fatima.penin': hashlib.sha256('13051982'.encode()).hexdigest(),
    'jaime.monguio': hashlib.sha256('13051993'.encode()).hexdigest(),
    'jose.gil': hashlib.sha256('23111990'.encode()).hexdigest(),
    'jose.roldan': hashlib.sha256('08031991'.encode()).hexdigest(),
    'lorena.tainta': hashlib.sha256('12041983'.encode()).hexdigest(),
    'marina.montemayor': hashlib.sha256('01071991'.encode()).hexdigest(),
    'marta.soto': hashlib.sha256('07071988'.encode()).hexdigest(),
    'monica.iglesias': hashlib.sha256('29071977'.encode()).hexdigest(),
    'natalia.buiucli': hashlib.sha256('05031986'.encode()).hexdigest(),
    'pedro.zarate': hashlib.sha256('01071989'.encode()).hexdigest(),
    'tabita.ortega': hashlib.sha256('04121993'.encode()).hexdigest(),
    'valentin.velasco': hashlib.sha256('27081976'.encode()).hexdigest(),
    'claudia.ventura': hashlib.sha256('20081986'.encode()).hexdigest(),
    'emilio.espinosa': hashlib.sha256('28051975'.encode()).hexdigest(),
}

LOGS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'access_logs.json')
INFORMES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'informes')
ADMIN_EMAIL = 'a.n.s15@hotmail.com'

def get_client_ip():
    """Intenta obtener la IP del cliente de forma segura para Hugging Face Spaces"""
    try:
        # Intento 1: Variables de entorno (común en Hugging Face Spaces y proxies)
        forwarded_for = os.environ.get('HTTP_X_FORWARDED_FOR')
        if forwarded_for:
            return forwarded_for.split(',')[0].strip()
        
        real_ip = os.environ.get('HTTP_X_REAL_IP')
        if real_ip:
            return real_ip.strip()
        
        # Intento 2: Streamlit context (solo si está disponible y activo)
        try:
            if hasattr(st, 'context') and st.context is not None:
                headers = getattr(st.context, 'headers', None)
                if headers:
                    ip = headers.get('X-Forwarded-For', headers.get('X-Real-IP', ''))
                    if ip:
                        return ip.split(',')[0].strip()
        except Exception:
            pass  # Silenciosamente ignorar si no hay contexto activo
        
        return 'No disponible'
    except Exception:
        return 'No disponible'

def get_location_from_ip(ip):
    """Intenta obtener ubicación aproximada desde IP de forma segura"""
    try:
        if not ip or ip == 'No disponible':
            return 'No disponible'
        if ip.startswith('127.') or ip.startswith('192.168.') or ip.startswith('10.') or ip.startswith('172.'):
            return 'Local/Privada'
        url = f"http://ip-api.com/json/{ip}?fields=city,country"
        with urllib.request.urlopen(url, timeout=1) as response:
            data = json.loads(response.read().decode())
            city = data.get('city', '')
            country = data.get('country', '')
            if city or country:
                return f"{city}, {country}".strip(', ')
            return 'No disponible'
    except Exception:
        return 'No disponible'

def send_login_notification(username, ip, location, login_time):
    """Envía email de notificación de conexión (silencioso si falla)"""
    try:
        subject = f"🔐 Nueva conexión - Analizador de Programaciones"
        body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; padding: 20px;">
            <div style="background: linear-gradient(135deg, #0D5F5D 0%, #117F7C 100%); padding: 20px; border-radius: 10px; color: white;">
                <h2 style="margin: 0;">✈️ Analizador de Programaciones</h2>
                <p style="margin: 5px 0;">Nueva conexión detectada</p>
            </div>
            <div style="padding: 20px; background: #f8f9fa; border-radius: 10px; margin-top: 15px;">
                <h3 style="color: #0D5F5D;">Detalles de la conexión:</h3>
                <table style="width: 100%; border-collapse: collapse;">
                    <tr><td style="padding: 8px; border-bottom: 1px solid #ddd;"><strong>👤 Usuario:</strong></td><td style="padding: 8px; border-bottom: 1px solid #ddd;">{username}</td></tr>
                    <tr><td style="padding: 8px; border-bottom: 1px solid #ddd;"><strong>📅 Fecha/Hora:</strong></td><td style="padding: 8px; border-bottom: 1px solid #ddd;">{login_time}</td></tr>
                    <tr><td style="padding: 8px; border-bottom: 1px solid #ddd;"><strong>🌐 Dirección IP:</strong></td><td style="padding: 8px; border-bottom: 1px solid #ddd;">{ip}</td></tr>
                    <tr><td style="padding: 8px;"><strong>📍 Ubicación:</strong></td><td style="padding: 8px;">{location}</td></tr>
                </table>
            </div>
            <p style="color: #666; font-size: 12px; margin-top: 20px;">Este es un mensaje automático del sistema de Analizador de Programaciones.</p>
        </body>
        </html>
        """
        # Guardar notificación en archivo para registro
        try:
            notif_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'email_notifications.json')
            notif = {
                'to': ADMIN_EMAIL,
                'subject': subject,
                'username': username,
                'ip': ip,
                'location': location,
                'time': login_time,
                'sent_at': datetime.now().isoformat()
            }
            notifications = []
            if os.path.exists(notif_file):
                try:
                    with open(notif_file, 'r') as f:
                        notifications = json.load(f)
                except Exception:
                    notifications = []
            notifications.append(notif)
            with open(notif_file, 'w') as f:
                json.dump(notifications, f, indent=2)
        except Exception:
            pass  # Silenciosamente ignorar errores de escritura
        return True
    except Exception:
        return False

def save_access_log(username):
    """Guarda log de acceso con IP, ubicación y envía email de notificación (nunca falla)"""
    try:
        ip = get_client_ip()
        location = get_location_from_ip(ip)
        login_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        log_entry = {
            'usuario': username,
            'fecha': datetime.now().strftime('%Y-%m-%d'),
            'hora': datetime.now().strftime('%H:%M:%S'),
            'timestamp': datetime.now().isoformat(),
            'ip': ip,
            'ubicacion': location,
            'hora_desconexion': None  # Se actualizará al cerrar sesión
        }
        
        logs = []
        if os.path.exists(LOGS_FILE):
            try:
                with open(LOGS_FILE, 'r') as f:
                    logs = json.load(f)
            except Exception:
                logs = []
        
        logs.append(log_entry)
        
        try:
            with open(LOGS_FILE, 'w') as f:
                json.dump(logs, f, indent=2)
        except Exception:
            pass  # Silenciosamente ignorar errores de escritura
        
        # Enviar notificación por email (silencioso si falla)
        try:
            send_login_notification(username, ip, location, login_time)
        except Exception:
            pass
    except Exception:
        pass  # La función NUNCA debe romper el flujo de login

def generate_report(username, base, tripulantes):
    """Genera informe de acceso/procesamiento"""
    report = {
        'usuario': username,
        'fecha': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'base': base,
        'tripulantes_analizados': tripulantes,
    }
    os.makedirs(INFORMES_DIR, exist_ok=True)
    report_file = os.path.join(INFORMES_DIR, f'informe_{username}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json')
    with open(report_file, 'w') as f:
        json.dump(report, f, indent=2)
    return report_file

def login_page():
    """Página de login"""
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
    * { font-family: 'Inter', sans-serif; }
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    </style>
    """, unsafe_allow_html=True)
    
    st.markdown("""
    <style>
    .stApp { background: linear-gradient(160deg, #f0fafa 0%, #e8f6f6 30%, #f5fffe 60%, #eaf9f9 100%) !important; }
    </style>
    """, unsafe_allow_html=True)
    
    st.markdown("""
    <div style="background: linear-gradient(135deg, #0A5C5A 0%, #0D6F6D 15%, #117F7C 35%, #1A9E9B 60%, #44BABC 85%, #61C1C3 100%); 
                padding: 45px; 
                border-radius: 28px; 
                text-align: center;
                margin: 40px auto;
                max-width: 550px;
                position: relative;
                overflow: hidden;
                box-shadow: 0 20px 60px rgba(17,127,124,0.35), 0 8px 25px rgba(11,132,127,0.2);">
        <div style="position:absolute;top:-50%;left:-50%;width:200%;height:200%;background:radial-gradient(circle at 30% 40%,rgba(97,193,195,0.15) 0%,transparent 50%),radial-gradient(circle at 70% 60%,rgba(68,186,188,0.12) 0%,transparent 50%);"></div>
        <div style="position:absolute;bottom:0;left:0;right:0;height:1px;background:linear-gradient(90deg,transparent,rgba(255,255,255,0.3),transparent);"></div>
        <div style="position:relative;z-index:1;">
            <div style="font-size:3rem;margin-bottom:8px;filter:drop-shadow(0 4px 8px rgba(0,0,0,0.15));">✈️</div>
            <h1 style="color: white; margin-bottom: 5px; font-size: 2.2rem; font-weight: 900; text-shadow: 0 2px 10px rgba(0,0,0,0.1);">ACCESO PRIVADO</h1>
            <div style="width: 80px; height: 4px; background: linear-gradient(90deg, #61C1C3, rgba(255,255,255,0.6)); border-radius: 2px; margin: 14px auto;"></div>
            <p style="color: rgba(255,255,255,0.9); font-size: 0.95rem; font-weight: 500;">Gestión y Análisis de Programación para Delegados</p>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("""
        <div style="background: rgba(255,255,255,0.85); backdrop-filter: blur(20px); border-radius: 20px; padding: 30px; margin-top: 20px;
                    box-shadow: 0 8px 32px rgba(17,127,124,0.08), 0 2px 8px rgba(0,0,0,0.04); border-top: 4px solid #117F7C;
                    border: 1px solid rgba(17,127,124,0.08);">
            <h3 style="color: #0D5F5D; text-align: center; margin-bottom: 20px;">🔐 Iniciar Sesión</h3>
        </div>
        """, unsafe_allow_html=True)
        username = st.text_input("👤 Usuario", key="login_user", placeholder="Introduce tu usuario")
        password = st.text_input("🔒 Contraseña", type="password", key="login_pass", placeholder="Introduce tu contraseña")
        
        # Detectar ENTER en contraseña o clic en botón
        login_clicked = st.button("🚀 Acceder", use_container_width=True, type="primary")
        
        # Ejecutar login con botón o si hay contraseña (permite ENTER)
        if login_clicked or (password and username):
            username_lower = username.lower().strip()
            if username_lower in USERS:
                hashed_password = hashlib.sha256(password.encode()).hexdigest()
                if USERS[username_lower] == hashed_password:
                    st.session_state['authenticated'] = True
                    st.session_state['username'] = username_lower.replace('.', ' ').title()
                    st.session_state['is_admin'] = (username_lower == 'admin')
                    save_access_log(username)
                    st.rerun()
                else:
                    st.error("❌ Contraseña incorrecta")
            else:
                st.error("❌ Usuario no encontrado")

# Control de autenticación
if 'authenticated' not in st.session_state:
    st.session_state['authenticated'] = False

if not st.session_state['authenticated']:
    login_page()
    st.stop()

# --- INICIALIZACIÓN DE SECCIÓN ---
if 'app_section' not in st.session_state:
    st.session_state['app_section'] = 'seleccion'
if 'analysis_complete' not in st.session_state:
    st.session_state['analysis_complete'] = False
if 'processed_data' not in st.session_state:
    st.session_state['processed_data'] = None

# --- SCROLL AL TOP AL CAMBIAR DE SECCIÓN ---
_prev_section = st.session_state.get('_prev_app_section', None)
_curr_section = st.session_state.get('app_section', 'seleccion')
if _prev_section is not None and _prev_section != _curr_section:
    # Inyectar JavaScript para scroll al top
    st.markdown("""
    <script>
    window.parent.document.querySelector('section.main').scrollTo({top: 0, behavior: 'instant'});
    window.scrollTo({top: 0, behavior: 'instant'});
    </script>
    """, unsafe_allow_html=True)
st.session_state['_prev_app_section'] = _curr_section

def update_logout_time():
    """Actualiza la hora de desconexión del último log del usuario"""
    try:
        if os.path.exists(LOGS_FILE):
            with open(LOGS_FILE, 'r') as f:
                logs = json.load(f)
            if logs:
                # Buscar el último log del usuario actual
                username = st.session_state.get('username', '')
                for i in range(len(logs) - 1, -1, -1):
                    if logs[i].get('usuario') == username and logs[i].get('hora_desconexion') is None:
                        logs[i]['hora_desconexion'] = datetime.now().strftime('%H:%M:%S')
                        break
                with open(LOGS_FILE, 'w') as f:
                    json.dump(logs, f, indent=2)
    except:
        pass

def render_admin_sidebar():
    """Renderiza el Panel de Administración en el sidebar (solo para admins)"""
    if not st.session_state.get('is_admin'):
        return
    
    with st.sidebar:
        st.markdown("""
        <div style="background: linear-gradient(135deg, #117F7C 0%, #0B847F 100%); 
                    padding: 15px; border-radius: 12px; text-align: center; margin-bottom: 15px;">
            <h3 style="color: white; margin: 0; font-weight: 700;">👑 Panel Admin</h3>
        </div>
        """, unsafe_allow_html=True)
        
        # Mostrar últimas conexiones
        st.markdown("##### 📊 Últimas Conexiones")
        try:
            if os.path.exists(LOGS_FILE):
                with open(LOGS_FILE, 'r') as f:
                    logs = json.load(f)
                if logs:
                    # Mostrar últimas 5 conexiones
                    for log in reversed(logs[-5:]):
                        estado = "🟢" if log.get('hora_desconexion') is None else "⚪"
                        st.markdown(f"""
                        <div style="background: #f8f9fa; padding: 8px 12px; border-radius: 8px; margin-bottom: 6px; font-size: 0.8rem;">
                            {estado} <strong>{log.get('usuario', 'N/A')}</strong><br>
                            <span style="color: #666;">📅 {log.get('fecha', 'N/A')} {log.get('hora_conexion', '')}</span>
                        </div>
                        """, unsafe_allow_html=True)
                else:
                    st.info("Sin registros de conexión")
            else:
                st.info("Sin registros de conexión")
        except Exception as e:
            st.warning(f"Error leyendo logs: {str(e)[:50]}")
        
        # Mostrar notificaciones de email pendientes
        st.markdown("##### 📧 Notificaciones Email")
        try:
            notif_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'email_notifications.json')
            if os.path.exists(notif_file):
                with open(notif_file, 'r') as f:
                    notifications = json.load(f)
                if notifications:
                    st.markdown(f"""
                    <div style="background: #FEF3C7; padding: 10px; border-radius: 8px; border-left: 3px solid #F59E0B;">
                        <strong>📬 {len(notifications)}</strong> notificaciones registradas
                    </div>
                    """, unsafe_allow_html=True)
                    with st.expander("Ver detalles"):
                        for notif in reversed(notifications[-5:]):
                            st.markdown(f"- **{notif.get('username', 'N/A')}** ({notif.get('time', 'N/A')})")
            else:
                st.markdown("""
                <div style="background: #F0F7F7; padding: 10px; border-radius: 8px;">
                    ✅ Sin notificaciones pendientes
                </div>
                """, unsafe_allow_html=True)
        except:
            pass
        
        st.markdown("---")
        
        # Link al panel completo
        if st.button("🔐 Panel Completo", key="sidebar_admin_full", use_container_width=True):
            st.session_state['app_section'] = 'admin_panel'
            st.rerun()

def selection_screen():
    """Pantalla de selección tras el login - Diseño Premium"""
    # Renderizar sidebar de admin si corresponde
    render_admin_sidebar()
    
    # ==================== CSS PREMIUM ====================
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
    * { font-family: 'Inter', sans-serif; }
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    
    /* ===== ANIMATED BACKGROUND ===== */
    .stApp {
        background: linear-gradient(160deg, #f0fafa 0%, #e8f6f6 30%, #f5fffe 60%, #eaf9f9 100%) !important;
    }
    
    /* ===== TOP BAR ===== */
    .top-bar {
        display: flex;
        justify-content: space-between;
        align-items: center;
        padding: 12px 20px;
        background: rgba(255,255,255,0.7);
        backdrop-filter: blur(20px);
        -webkit-backdrop-filter: blur(20px);
        border-radius: 16px;
        border: 1px solid rgba(17,127,124,0.08);
        box-shadow: 0 2px 16px rgba(17,127,124,0.06);
        margin-bottom: 24px;
    }
    .top-bar .user-info {
        display: flex;
        align-items: center;
        gap: 10px;
        font-size: 0.9rem;
        color: #0D5F5D;
        font-weight: 600;
    }
    .top-bar .user-avatar {
        width: 36px; height: 36px;
        background: linear-gradient(135deg, #117F7C, #44BABC);
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        color: white;
        font-size: 1rem;
        font-weight: 800;
        box-shadow: 0 3px 10px rgba(17,127,124,0.3);
    }
    .top-bar .btn-group {
        display: flex;
        gap: 8px;
    }
    
    /* ===== HERO BANNER ===== */
    .hero-banner {
        background: linear-gradient(135deg, #0A5C5A 0%, #0D6F6D 15%, #117F7C 35%, #1A9E9B 60%, #44BABC 85%, #61C1C3 100%);
        padding: 55px 50px;
        border-radius: 28px;
        text-align: center;
        margin: 0 auto 45px auto;
        max-width: 100%;
        position: relative;
        overflow: hidden;
        box-shadow: 0 20px 60px rgba(17,127,124,0.35), 0 8px 25px rgba(11,132,127,0.2);
    }
    .hero-banner::before {
        content: '';
        position: absolute;
        top: -50%; left: -50%;
        width: 200%; height: 200%;
        background: radial-gradient(circle at 30% 40%, rgba(97,193,195,0.15) 0%, transparent 50%),
                    radial-gradient(circle at 70% 60%, rgba(68,186,188,0.12) 0%, transparent 50%),
                    radial-gradient(circle at 50% 20%, rgba(255,255,255,0.05) 0%, transparent 40%);
        animation: shimmer 8s ease-in-out infinite;
    }
    @keyframes shimmer {
        0%, 100% { transform: translate(0, 0) rotate(0deg); }
        25% { transform: translate(-2%, 1%) rotate(0.5deg); }
        50% { transform: translate(1%, -1%) rotate(-0.5deg); }
        75% { transform: translate(-1%, 2%) rotate(0.3deg); }
    }
    .hero-banner::after {
        content: '';
        position: absolute;
        bottom: 0; left: 0; right: 0;
        height: 1px;
        background: linear-gradient(90deg, transparent, rgba(255,255,255,0.3), transparent);
    }
    .hero-banner .hero-icon {
        font-size: 3.5rem;
        margin-bottom: 8px;
        filter: drop-shadow(0 4px 8px rgba(0,0,0,0.15));
        position: relative;
        z-index: 1;
    }
    .hero-banner h1 {
        font-size: 2.6rem;
        font-weight: 900;
        color: white;
        margin: 0 0 6px 0;
        letter-spacing: -0.8px;
        position: relative;
        z-index: 1;
        text-shadow: 0 2px 10px rgba(0,0,0,0.1);
    }
    .hero-banner .divider {
        width: 90px;
        height: 4px;
        background: linear-gradient(90deg, #61C1C3, rgba(255,255,255,0.6));
        border-radius: 2px;
        margin: 16px auto;
        position: relative;
        z-index: 1;
    }
    .hero-banner p {
        color: rgba(255,255,255,0.9);
        font-size: 1.1rem;
        font-weight: 500;
        margin: 0;
        position: relative;
        z-index: 1;
    }
    
    /* ===== TOOL CARDS - GLASSMORPHISM ===== */
    .tool-card {
        background: rgba(255,255,255,0.65);
        backdrop-filter: blur(20px);
        -webkit-backdrop-filter: blur(20px);
        border-radius: 24px;
        padding: 40px 28px 32px 28px;
        text-align: center;
        border: 1px solid rgba(17,127,124,0.1);
        box-shadow: 0 8px 32px rgba(17,127,124,0.08), 0 2px 8px rgba(0,0,0,0.04);
        min-height: 320px;
        display: flex;
        flex-direction: column;
        justify-content: center;
        align-items: center;
        transition: all 0.4s cubic-bezier(0.175, 0.885, 0.32, 1.275);
        position: relative;
        overflow: hidden;
        cursor: default;
        pointer-events: none;
    }
    .tool-card::before {
        content: '';
        position: absolute;
        top: 0; left: 0; right: 0;
        height: 5px;
        border-radius: 24px 24px 0 0;
    }
    .tool-card.card-analizador::before { background: linear-gradient(90deg, #117F7C, #44BABC); }
    .tool-card.card-generador::before { background: linear-gradient(90deg, #0B847F, #61C1C3); }
    .tool-card.card-recortador::before { background: linear-gradient(90deg, #489999, #44BABC); }
    .tool-card.card-tablas::before { background: linear-gradient(90deg, #0D5F5D, #489999); }
    .tool-card.card-calculadora::before { background: linear-gradient(90deg, #117F7C, #61C1C3); }
    
    .tool-card .card-icon-wrap {
        width: 80px; height: 80px;
        border-radius: 22px;
        display: flex;
        align-items: center;
        justify-content: center;
        margin-bottom: 20px;
        font-size: 2.8rem;
        box-shadow: 0 8px 25px rgba(0,0,0,0.1);
        transition: all 0.4s ease;
    }
    .tool-card.card-analizador .card-icon-wrap { background: linear-gradient(145deg, #117F7C, #44BABC); }
    .tool-card.card-generador .card-icon-wrap { background: linear-gradient(145deg, #0B847F, #61C1C3); }
    .tool-card.card-recortador .card-icon-wrap { background: linear-gradient(145deg, #489999, #44BABC); }
    .tool-card.card-tablas .card-icon-wrap { background: linear-gradient(145deg, #0D5F5D, #489999); }
    .tool-card.card-calculadora .card-icon-wrap { background: linear-gradient(145deg, #117F7C, #61C1C3); }
    
    .tool-card h2 {
        color: #0D5F5D;
        font-size: 1.35rem;
        font-weight: 800;
        margin: 0 0 10px 0;
        letter-spacing: -0.3px;
    }
    .tool-card p {
        color: #64748B;
        font-size: 0.88rem;
        line-height: 1.55;
        margin: 0;
    }
    
    /* ===== BOTONES PRIMARY - Gradiente vibrante ===== */
    div[data-testid="stButton"] > button[kind="primary"],
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #0D5F5D 0%, #117F7C 40%, #44BABC 100%) !important;
        border: none !important;
        padding: 14px 28px !important;
        font-size: 1.05rem !important;
        font-weight: 700 !important;
        border-radius: 14px !important;
        margin-top: 14px !important;
        box-shadow: 0 6px 24px rgba(17,127,124,0.35) !important;
        pointer-events: auto !important;
        cursor: pointer !important;
        transition: all 0.3s cubic-bezier(0.175, 0.885, 0.32, 1.275) !important;
        color: white !important;
        letter-spacing: 0.2px !important;
    }
    div[data-testid="stButton"] > button[kind="primary"]:hover,
    .stButton > button[kind="primary"]:hover {
        background: linear-gradient(135deg, #0A5C5A 0%, #0D6F6D 40%, #117F7C 100%) !important;
        box-shadow: 0 10px 35px rgba(17,127,124,0.5) !important;
        transform: translateY(-3px) !important;
    }
    
    /* ===== BOTONES SECONDARY (Admin, Cerrar Sesión) ===== */
    div[data-testid="stButton"] > button[kind="secondary"],
    .stButton > button[kind="secondary"],
    button[kind="secondary"],
    button[data-testid="stBaseButton-secondary"] {
        background: rgba(255,255,255,0.8) !important;
        backdrop-filter: blur(10px) !important;
        color: #0D5F5D !important;
        border: 1.5px solid rgba(17,127,124,0.2) !important;
        font-weight: 600 !important;
        padding: 10px 20px !important;
        border-radius: 12px !important;
        box-shadow: 0 2px 10px rgba(17,127,124,0.08) !important;
        transition: all 0.3s ease !important;
        font-size: 0.88rem !important;
    }
    div[data-testid="stButton"] > button[kind="secondary"]:hover,
    .stButton > button[kind="secondary"]:hover,
    button[kind="secondary"]:hover,
    button[data-testid="stBaseButton-secondary"]:hover {
        transform: translateY(-2px) !important;
        background: rgba(17,127,124,0.08) !important;
        border-color: #117F7C !important;
        box-shadow: 0 6px 20px rgba(17,127,124,0.15) !important;
        color: #117F7C !important;
    }
    
    /* ===== GENERAL BUTTONS (default) ===== */
    .stButton > button,
    div[data-testid="stButton"] > button {
        background: linear-gradient(135deg, #0D5F5D 0%, #117F7C 100%) !important;
        color: white !important;
        font-weight: 700 !important;
        border: none !important;
        padding: 12px 24px !important;
        border-radius: 12px !important;
        box-shadow: 0 4px 14px rgba(17,127,124,0.3) !important;
        transition: all 0.3s ease !important;
    }
    .stButton > button:hover,
    div[data-testid="stButton"] > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 24px rgba(17,127,124,0.45) !important;
    }
    
    /* ===== FOOTER ===== */
    .home-footer {
        text-align: center;
        padding: 20px 0 8px 0;
        margin-top: 30px;
    }
    .home-footer .footer-line {
        width: 100px;
        height: 2px;
        background: linear-gradient(90deg, transparent, #44BABC, transparent);
        margin: 0 auto 12px auto;
        border-radius: 1px;
    }
    .home-footer p {
        color: #94A3B8;
        font-size: 0.78rem;
        font-weight: 500;
        margin: 0;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # ==================== TOP BAR ====================
    _username = st.session_state.get('username', '?')
    _initial = _username[0].upper() if _username else '?'
    _admin_badge = "<span style='background:linear-gradient(135deg,#FFD700,#FFA500);color:#fff;padding:2px 10px;border-radius:20px;font-size:0.72rem;font-weight:700;margin-left:4px;'>👑 Admin</span>" if st.session_state.get('is_admin') else ""
    st.html(f'<div class="top-bar" style="display:flex;justify-content:space-between;align-items:center;padding:12px 20px;background:rgba(255,255,255,0.7);backdrop-filter:blur(20px);-webkit-backdrop-filter:blur(20px);border-radius:16px;border:1px solid rgba(17,127,124,0.08);box-shadow:0 2px 16px rgba(17,127,124,0.06);margin-bottom:24px;"><div class="user-info" style="display:flex;align-items:center;gap:10px;font-size:0.9rem;color:#0D5F5D;font-weight:600;"><div class="user-avatar" style="width:36px;height:36px;background:linear-gradient(135deg,#117F7C,#44BABC);border-radius:50%;display:flex;align-items:center;justify-content:center;color:white;font-size:1rem;font-weight:800;box-shadow:0 3px 10px rgba(17,127,124,0.3);">{_initial}</div><span>{_username}</span>{_admin_badge}</div></div>')
    
    # Botones de Admin y Cerrar Sesión - dentro del recuadro top-bar
    if st.session_state.get('is_admin'):
        _usr_col1, _usr_col2, _usr_col3 = st.columns([5, 1.2, 1])
    else:
        _usr_col1, _usr_col2, _usr_col3 = st.columns([7, 0.01, 1])
    
    # Inject CSS to pull buttons up into the top-bar area
    st.markdown("""
    <style>
    /* Pull the button row up into the top-bar visually */
    div[data-testid="stColumns"]:has(button[data-testid="stBaseButton-secondary"]) {
        margin-top: -58px !important;
        margin-bottom: 14px !important;
        padding-right: 20px !important;
    }
    </style>
    """, unsafe_allow_html=True)
        
    with _usr_col1:
        st.markdown("")  # spacer
    
    if st.session_state.get('is_admin'):
        with _usr_col2:
            if st.button("🔐 Panel Admin", key="admin_panel_btn", type="secondary"):
                st.session_state['app_section'] = 'admin_panel'
                st.rerun()
    
    with _usr_col3:
        if st.button("🚪 Salir", key="logout_sel", type="secondary"):
            update_logout_time()
            for key in ['authenticated', 'username', 'is_admin', 'app_section']:
                st.session_state[key] = False if key == 'authenticated' else None
            st.session_state['app_section'] = 'seleccion'
            st.rerun()
    
    # ==================== HERO BANNER ====================
    st.markdown("""
    <div class="hero-banner">
        <div class="hero-icon">✈️</div>
        <h1>ÁREA DE TRABAJO</h1>
        <div class="divider"></div>
        <p>Selecciona la herramienta que deseas utilizar</p>
    </div>
    """, unsafe_allow_html=True)
    
    # ==================== TOOL CARDS ====================
    # Primera fila: Analizador y Generador de PDFs
    col1, col2 = st.columns(2, gap="large")
    
    with col1:
        st.markdown("""
        <div class="tool-card card-analizador">
            <div class="card-icon-wrap">📊</div>
            <h2>Analizador de Programaciones</h2>
            <p>Analiza programaciones de tripulación, visualiza estadísticas y genera informes detallados.</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("📊 Abrir Analizador", key="btn_analizador", use_container_width=True, type="primary"):
            st.session_state['app_section'] = 'analizador'
            st.rerun()
    
    with col2:
        st.markdown("""
        <div class="tool-card card-generador">
            <div class="card-icon-wrap">📄</div>
            <h2>Generador de PDFs</h2>
            <p>Genera PDFs de programaciones finales desde archivos de roster (.txt). Réplica fiel del diseño oficial.</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("📄 Abrir Generador", key="btn_pdf", use_container_width=True, type="primary"):
            st.session_state['app_section'] = 'pdf_generator'
            st.rerun()
    
    # Segunda fila: Recortador + Tablas Salariales
    st.markdown("<div style='margin-top: 20px;'></div>", unsafe_allow_html=True)
    col3, col4 = st.columns(2, gap="large")
    
    with col3:
        st.markdown("""
        <div class="tool-card card-recortador">
            <div class="card-icon-wrap">✂️</div>
            <h2>Recortar Programaciones</h2>
            <p>Extrae y genera PDFs de programaciones iniciales filtradas por base seleccionada.</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("✂️ Abrir Recortador", key="btn_recortador", use_container_width=True, type="primary"):
            st.session_state['app_section'] = 'recortador'
            st.rerun()
    
    with col4:
        st.markdown("""
        <div class="tool-card card-tablas">
            <div class="card-icon-wrap">💶</div>
            <h2>Tablas Salariales 2026</h2>
            <p>Consulta las tablas salariales vigentes con todos los niveles. Descarga en Excel o PDF.</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("💶 Abrir Tablas Salariales", key="btn_tablas", use_container_width=True, type="primary"):
            st.session_state['app_section'] = 'tablas_salariales'
            st.rerun()
    
    # Tercera fila: Calculadora de Nóminas (centrada)
    st.markdown("<div style='margin-top: 20px;'></div>", unsafe_allow_html=True)
    _spacer_l, col5, _spacer_r = st.columns([1, 2, 1])
    
    with col5:
        st.markdown("""
        <div class="tool-card card-calculadora">
            <div class="card-icon-wrap"><svg viewBox="0 0 24 24" style="width:2.4rem;height:2.4rem;stroke:white;fill:none;stroke-width:1.8;stroke-linecap:round;stroke-linejoin:round;"><rect width="16" height="20" x="4" y="2" rx="2"/><line x1="8" x2="16" y1="6" y2="6"/><line x1="16" x2="16" y1="14" y2="18"/><path d="M16 10h.01"/><path d="M12 10h.01"/><path d="M8 10h.01"/><path d="M12 14h.01"/><path d="M8 14h.01"/><path d="M12 18h.01"/><path d="M8 18h.01"/></svg></div>
            <h2>Calculadora de Nóminas</h2>
            <p>Calcula tu nómina estimada según tu nivel, tipo (TCP/Sobrecargo) y conceptos variables.</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("🔢 Abrir Calculadora", key="btn_calculadora", use_container_width=True, type="primary"):
            st.session_state['app_section'] = 'calculadora_nominas'
            st.rerun()
    
    # ==================== FOOTER ====================
    st.markdown("""
    <div class="home-footer">
        <div class="footer-line"></div>
        <p>✈️ Herramientas Sindicales © 2026</p>
    </div>
    """, unsafe_allow_html=True)


# --- ROUTING ---
if st.session_state['app_section'] == 'seleccion':
    selection_screen()
    st.stop()

# Paleta de colores del sindicato: tonos turquesa
# #117F7C (principal), #0B847F (secundario), #489999 (acento1), #44BABC (acento2), #61C1C3 (acento3)
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
    
    * { font-family: 'Inter', sans-serif; }
    
    /* ========== MAIN HEADER (Analizador) ========== */
    .main-header {
        background: linear-gradient(135deg, #0A5C5A 0%, #0D6F6D 15%, #117F7C 35%, #1A9E9B 60%, #44BABC 85%, #61C1C3 100%);
        color: white;
        padding: 36px 44px;
        border-radius: 20px;
        margin-bottom: 28px;
        position: relative;
        overflow: hidden;
        box-shadow: 0 12px 40px rgba(17,127,124,0.35), 0 4px 15px rgba(11,132,127,0.15);
    }
    .main-header::before {
        content: '';
        position: absolute;
        top: -50%; right: -20%;
        width: 400px; height: 400px;
        background: radial-gradient(circle, rgba(97,193,195,0.15) 0%, transparent 70%);
        border-radius: 50%;
    }
    .main-header::after {
        content: '';
        position: absolute;
        bottom: 0; left: 0; right: 0;
        height: 1px;
        background: linear-gradient(90deg, transparent, rgba(255,255,255,0.2), transparent);
    }
    .main-header h1 {
        font-size: 2.2rem;
        font-weight: 900;
        margin: 0 0 5px 0;
        letter-spacing: -0.5px;
        position: relative;
        z-index: 1;
        text-shadow: 0 2px 6px rgba(0,0,0,0.08);
    }
    .main-header p {
        font-size: 1rem;
        opacity: 0.9;
        margin: 0;
        font-weight: 500;
        position: relative;
        z-index: 1;
    }
    .main-header .accent-line {
        width: 80px;
        height: 4px;
        background: linear-gradient(90deg, #61C1C3, rgba(255,255,255,0.5));
        border-radius: 2px;
        margin: 12px 0;
        position: relative;
        z-index: 1;
    }
    
    /* ========== SECTION TITLE ========== */
    .section-title {
        font-size: 1.35rem;
        color: #0D5F5D;
        font-weight: 800;
        margin: 1.5rem 0 1rem 0;
        padding-bottom: 0.5rem;
        border-bottom: 3px solid #117F7C;
        letter-spacing: -0.3px;
    }
    
    /* ========== BASE BANNER ========== */
    .base-banner {
        background: linear-gradient(135deg, #0D5F5D 0%, #117F7C 60%, #489999 100%);
        color: white;
        padding: 14px 28px;
        border-radius: 12px;
        margin-bottom: 20px;
        display: flex;
        align-items: center;
        gap: 24px;
        font-size: 0.95rem;
        box-shadow: 0 6px 20px rgba(17,127,124,0.35);
        border-left: 5px solid #44BABC;
    }
    .base-banner .base-name {
        font-weight: 900;
        font-size: 1.15rem;
        background: rgba(68,186,188,0.3);
        padding: 5px 16px;
        border-radius: 8px;
    }
    .base-banner .stat {
        display: flex;
        align-items: center;
        gap: 6px;
        font-weight: 500;
    }
    .base-banner .stat-num {
        font-weight: 800;
        font-size: 1.1rem;
    }
    
    /* ========== KPI BOXES ========== */
    .kpi-box { 
        background: rgba(255,255,255,0.85);
        backdrop-filter: blur(12px);
        -webkit-backdrop-filter: blur(12px);
        border-radius: 18px; 
        padding: 24px 20px; 
        text-align: center; 
        border-top: 4px solid #117F7C; 
        box-shadow: 0 4px 20px rgba(17,127,124,0.08), 0 1px 4px rgba(0,0,0,0.04); 
        margin-bottom: 15px;
        transition: all 0.35s cubic-bezier(0.175, 0.885, 0.32, 1.275);
        border: 1px solid rgba(17,127,124,0.06);
    }
    .kpi-box:hover { transform: translateY(-5px); box-shadow: 0 14px 40px rgba(17,127,124,0.18); }
    .kpi-box h1 { font-size: 2.5rem; color: #0D5F5D; margin: 4px 0; font-weight: 900; letter-spacing: -1px; }
    .kpi-box p { font-size: 0.78rem; color: #64748B; font-weight: 700; margin: 0; text-transform: uppercase; letter-spacing: 0.8px; }
    .kpi-box.accent { border-top-color: #44BABC; }
    .kpi-box.accent h1 { color: #117F7C; }
    
    /* ========== CONCEPTO CARDS ========== */
    .concepto-card {
        background: rgba(255,255,255,0.8);
        backdrop-filter: blur(10px);
        border-radius: 14px;
        padding: 14px 20px;
        margin: 6px 0;
        box-shadow: 0 3px 16px rgba(17,127,124,0.06), 0 1px 3px rgba(0,0,0,0.03);
        display: flex;
        justify-content: space-between;
        align-items: center;
        border-left: 5px solid #117F7C;
        transition: all 0.3s cubic-bezier(0.175, 0.885, 0.32, 1.275);
        border: 1px solid rgba(17,127,124,0.06);
        border-left: 5px solid #117F7C;
    }
    .concepto-card:hover {
        transform: translateX(6px);
        box-shadow: 0 8px 28px rgba(17,127,124,0.15);
        border-left-color: #44BABC;
    }
    .concepto-card .concepto-name {
        font-weight: 700;
        font-size: 1rem;
        color: #0D5F5D;
    }
    .concepto-card .concepto-stats {
        display: flex;
        gap: 12px;
        align-items: center;
    }
    .concepto-card .concepto-pct {
        font-weight: 900;
        font-size: 1.5rem;
        color: #117F7C;
    }
    .concepto-card .concepto-detail {
        font-size: 0.78rem;
        color: #64748B;
        font-weight: 500;
    }
    
    /* ========== MINI KPI CARDS ========== */
    .mini-kpi {
        background: white;
        border-radius: 12px;
        padding: 16px;
        text-align: center;
        border: 1px solid #E0F2F1;
        box-shadow: 0 3px 12px rgba(0,0,0,0.06);
        border-top: 3px solid #117F7C;
        transition: transform 0.25s ease, box-shadow 0.25s ease;
    }
    .mini-kpi:hover { transform: translateY(-3px); box-shadow: 0 8px 22px rgba(17,127,124,0.14); }
    .mini-kpi h2 { margin: 0; font-size: 1.9rem; color: #0D5F5D; font-weight: 900; letter-spacing: -0.5px; }
    .mini-kpi p { margin: 5px 0 0 0; font-size: 0.78rem; color: #64748B; font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px; }
    
    /* ========== CATEGORY CARDS - PROFESIONAL & COMPACTO ========== */
    .category-card {
        border-radius: 14px;
        padding: 18px 14px;
        margin: 0 0 18px 0;
        box-shadow: 0 4px 14px rgba(0,0,0,0.14);
        transition: all 0.3s ease;
        position: relative;
        overflow: hidden;
        min-height: 200px;
        max-height: 220px;
        text-align: center;
        display: flex;
        flex-direction: column;
        justify-content: center;
        align-items: center;
    }
    .category-card::after {
        content: '';
        position: absolute;
        top: 0; right: 0;
        width: 50%; height: 100%;
        background: linear-gradient(135deg, transparent 0%, rgba(255,255,255,0.08) 100%);
        pointer-events: none;
    }
    .category-card:hover {
        transform: translateY(-4px);
        box-shadow: 0 10px 28px rgba(0,0,0,0.20);
    }
    .category-card h2 { font-size: 1.05rem; margin: 0 0 6px 0; font-weight: 700; letter-spacing: 0.2px; white-space: normal; line-height: 1.25; overflow: hidden; text-overflow: ellipsis; width: 100%; }
    .category-card .cat-number { font-size: 3.2rem; font-weight: 900; letter-spacing: -1px; line-height: 1; }
    .category-card .cat-trip-row { display: flex; align-items: baseline; justify-content: center; margin: 8px 0; }
    .category-card .cat-trip-label { font-size: 1.1rem; font-weight: 700; color: white; margin-left: 8px; }
    .category-card .cat-info { font-size: 0.82rem; margin: 3px 0; opacity: 0.93; line-height: 1.3; font-weight: 500; }
    .category-card .cat-codes { font-size: 0.68rem; margin-top: 6px; opacity: 0.85; line-height: 1.25; border-top: 1px solid rgba(255,255,255,0.25); padding-top: 5px; font-weight: 500; word-break: break-word; }
    
    /* Category gradient colors */
    .cat-vuelo { background: linear-gradient(135deg, #0D5F5D 0%, #117F7C 100%); color: white; }
    .cat-guardias { background: linear-gradient(135deg, #357070 0%, #489999 100%); color: white; }
    .cat-libres { background: linear-gradient(135deg, #2F9E9F 0%, #44BABC 100%); color: white; }
    .cat-formacion { background: linear-gradient(135deg, #4FAAAC 0%, #61C1C3 100%); color: white; }
    .cat-oficina { background: linear-gradient(135deg, #096B67 0%, #0B847F 100%); color: white; }
    .cat-medico { background: linear-gradient(135deg, #0D5F5D 0%, #117F7C 100%); color: white; }
    .cat-permisos { background: linear-gradient(135deg, #2D7A7A 0%, #489999 100%); color: white; }
    .cat-vacaciones { background: linear-gradient(135deg, #2F9E9F 0%, #44BABC 100%); color: white; }
    .cat-parttime { background: linear-gradient(135deg, #489999 0%, #44BABC 100%); color: white; }
    .cat-incidencias { background: linear-gradient(135deg, #3D9698 0%, #61C1C3 100%); color: white; }
    .cat-otros { background: linear-gradient(135deg, #357070 0%, #489999 100%); color: white; }
    
    /* ========== DAY CARDS (individual calendar) ========== */
    .day-card { 
        border-radius: 10px; 
        padding: 8px 5px; 
        text-align: center; 
        box-shadow: 0 2px 10px rgba(0,0,0,0.12); 
        margin-bottom: 6px; 
        min-height: 90px; 
        display: flex; 
        flex-direction: column; 
        justify-content: center;
        align-items: center;
        transition: all 0.25s ease;
    }
    .day-card:hover { transform: scale(1.05); box-shadow: 0 6px 18px rgba(0,0,0,0.2); }
    .day-weekday { font-size: 0.58rem; font-weight: 700; opacity: 0.7; text-transform: uppercase; letter-spacing: 0.5px; }
    .day-number { font-size: 0.7rem; font-weight: 800; opacity: 0.9; }
    .day-icon { font-size: 1.15rem; margin: 2px 0; }
    .day-code { font-size: 0.95rem; font-weight: 900; letter-spacing: 0.5px; }
    .day-desc { font-size: 0.48rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.3px; margin-top: 1px; }
    
    /* Day-card color themes */
    .dc-vuelo { background: linear-gradient(135deg, #B71C1C 0%, #D32F2F 100%); color: white; }
    .dc-verde { background: linear-gradient(135deg, #2F9E9F 0%, #44BABC 100%); color: white; }
    .dc-franco { background: linear-gradient(135deg, #1565C0 0%, #2196F3 100%); color: white; }
    .dc-formacion { background: linear-gradient(135deg, #F9A825 0%, #FBC02D 100%); color: #333; }
    .dc-sindical { background: linear-gradient(135deg, #5D4037 0%, #795548 100%); color: white; }
    .dc-sick { background: linear-gradient(135deg, #C0392B 0%, #E74C3C 100%); color: white; }
    .dc-ftg { background: linear-gradient(135deg, #CA6F1E 0%, #E67E22 100%); color: white; }
    .dc-permiso { background: linear-gradient(135deg, #357070 0%, #489999 100%); color: white; }
    .dc-otro { background: linear-gradient(135deg, #096B67 0%, #0B847F 100%); color: white; }
    .dc-vac { background: linear-gradient(135deg, #2F9E9F 0%, #44BABC 100%); color: white; }
    .dc-guardia { background: linear-gradient(135deg, #0D47A1 0%, #1565C0 100%); color: white; }
    .dc-pt { background: linear-gradient(135deg, #0B847F 0%, #117F7C 100%); color: white; }
    
    /* ========== BUTTONS ========== */
    .stButton > button {
        background: linear-gradient(135deg, #0D5F5D 0%, #117F7C 100%);
        color: white;
        font-weight: 700;
        border: none;
        padding: 12px 24px;
        border-radius: 10px;
        transition: all 0.3s ease;
        box-shadow: 0 4px 14px rgba(17,127,124,0.3);
        letter-spacing: 0.3px;
    }
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 24px rgba(17,127,124,0.45);
        background: linear-gradient(135deg, #117F7C 0%, #0D5F5D 100%);
    }
    
    /* Botón Secondary - estilo consistente */
    div[data-testid="stButton"] > button[kind="secondary"],
    button[data-testid="stBaseButton-secondary"] {
        background: linear-gradient(135deg, #64748B 0%, #94A3B8 100%) !important;
        color: white !important;
        border: none !important;
        font-weight: 700 !important;
        padding: 12px 24px !important;
        border-radius: 10px !important;
        box-shadow: 0 4px 14px rgba(100, 116, 139, 0.3) !important;
    }
    div[data-testid="stButton"] > button[kind="secondary"]:hover,
    button[data-testid="stBaseButton-secondary"]:hover {
        transform: translateY(-2px) !important;
        background: linear-gradient(135deg, #475569 0%, #64748B 100%) !important;
        box-shadow: 0 6px 18px rgba(100, 116, 139, 0.4) !important;
    }
    
    /* ========== RESULT CARD ========== */
    .result-card {
        background: white;
        border-radius: 14px;
        padding: 24px;
        margin: 15px 0;
        box-shadow: 0 6px 24px rgba(0,0,0,0.09);
        border-left: 5px solid #117F7C;
    }
    
    /* ========== BADGES ========== */
    .badge {
        display: inline-block;
        padding: 6px 14px;
        border-radius: 20px;
        font-size: 0.85rem;
        font-weight: 700;
        margin: 4px;
    }
    .badge-jc { background: #0D5F5D; color: white; }
    .badge-tc { background: #44BABC; color: white; }
    .badge-block { background: #F0F7F7; color: #0D5F5D; border: 2px solid #B2DFDB; }
    .badge-duty { background: #F0F7F7; color: #0D5F5D; border: 2px solid #B2DFDB; }
    .badge-estab { background: #F0F7F7; color: #0D5F5D; border: 2px solid #B2DFDB; }
    
    /* ========== CENSUS TABLE ========== */
    .censo-table-wrapper {
        overflow-x: auto;
        margin-top: 15px;
        border-radius: 12px;
        box-shadow: 0 4px 18px rgba(0,0,0,0.08);
    }
    .censo-table {
        width: 100%;
        border-collapse: collapse;
        border-radius: 12px;
        overflow: hidden;
        box-shadow: none;
        margin-top: 0;
    }
    .censo-table th {
        background: linear-gradient(135deg, #0D5F5D 0%, #117F7C 100%);
        color: white;
        padding: 13px 12px;
        text-align: center;
        font-weight: 700;
        font-size: 0.85rem;
        letter-spacing: 0.3px;
    }
    .censo-table td {
        padding: 11px 10px;
        border-bottom: 1px solid #E0F2F1;
        font-size: 0.85rem;
        text-align: center;
    }
    .censo-table td:nth-child(2) {
        white-space: nowrap;
        text-align: left;
        min-width: 280px;
        max-width: none;
        overflow: visible;
        text-overflow: clip;
        word-break: keep-all;
    }
    .censo-table tr:nth-child(even) { background: #F0F7F7; }
    .censo-table tr:hover { background: #E0F2F1; }
    
    /* ========== PROGRESS BAR ========== */
    .progress-bar {
        background: #E0F2F1;
        border-radius: 10px;
        height: 10px;
        overflow: hidden;
    }
    .progress-fill {
        height: 100%;
        border-radius: 10px;
        transition: width 0.5s ease;
    }
    .progress-green { background: linear-gradient(90deg, #27AE60, #1E8449); }
    .progress-yellow { background: linear-gradient(90deg, #F39C12, #D68910); }
    .progress-red { background: linear-gradient(90deg, #E74C3C, #C0392B); }
    
    /* ========== CODE SUMMARY CARD ========== */
    .code-summary-card {
        background: white;
        border-radius: 10px;
        padding: 12px 18px;
        margin: 5px 0;
        box-shadow: 0 2px 10px rgba(0,0,0,0.05);
        display: flex;
        justify-content: space-between;
        align-items: center;
        border-left: 4px solid #117F7C;
        transition: all 0.25s ease;
    }
    .code-summary-card:hover {
        transform: translateX(5px);
        box-shadow: 0 5px 16px rgba(17,127,124,0.14);
    }
    
    /* ========== CALENDAR DAILY ========== */
    .cal-day {
        background: white;
        border-radius: 10px;
        padding: 8px;
        margin: 4px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.07);
        min-height: 120px;
        font-size: 0.7rem;
        border-top: 3px solid #117F7C;
    }
    .cal-day .cal-day-num {
        font-weight: 900;
        font-size: 1.05rem;
        color: #0D5F5D;
        margin-bottom: 4px;
    }
    .cal-day .cal-item {
        display: flex;
        justify-content: space-between;
        padding: 1px 0;
        font-size: 0.63rem;
    }
    .cal-weekend { border-top-color: #117F7C; }
    .cal-weekend .cal-day-num { color: #117F7C; }
    
    /* ========== TABS ========== */
    .stTabs [data-baseweb="tab-list"] {
        gap: 4px;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px 8px 0 0;
        padding: 8px 16px;
        font-weight: 700;
    }
    .stTabs [aria-selected="true"] {
        background-color: #117F7C !important;
        color: white !important;
    }
    
    /* ========== FILE UPLOADER - CUSTOM DESIGN (sin botón +) ========== */
    /* Contenedor principal */
    [data-testid="stFileUploader"] {
        background: #f8fafa;
        border: 2px dashed #117F7C;
        border-radius: 12px;
        padding: 20px;
        transition: all 0.3s ease;
    }
    [data-testid="stFileUploader"]:hover {
        background: rgba(17, 127, 124, 0.06);
        border-color: #0B847F;
    }
    [data-testid="stFileUploader"] section {
        border: none !important;
    }
    [data-testid="stFileUploader"] section > div {
        text-align: center;
        color: #117F7C;
        font-size: 1rem;
    }
    
    /* Botón principal "Browse files" - estilo personalizado */
    [data-testid="stFileUploader"] [data-testid="stBaseButton-secondary"],
    [data-testid="stFileUploader"] button[data-testid="stFileUploaderButton"],
    [data-testid="stFileUploader"] section > button:first-of-type {
        background: linear-gradient(135deg, #117F7C 0%, #44BABC 100%) !important;
        color: white !important;
        border: none !important;
        padding: 12px 30px !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        font-size: 1rem !important;
        box-shadow: 0 4px 12px rgba(17, 127, 124, 0.25) !important;
        transition: all 0.3s ease !important;
        cursor: pointer !important;
        margin: 10px auto !important;
        display: block !important;
    }
    [data-testid="stFileUploader"] [data-testid="stBaseButton-secondary"]:hover,
    [data-testid="stFileUploader"] section > button:first-of-type:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 16px rgba(17, 127, 124, 0.35) !important;
    }
    
    /* ★★★ OCULTAR BOTÓN "+" — Cuando hay archivo cargado, ocultar dropzone ★★★ */
    /* Método 1: :has() — ocultar dropzone si ya hay un archivo subido */
    [data-testid="stFileUploader"] section:has([data-testid="stUploadedFileLayout"]) > [data-testid="stFileUploaderDropzone"],
    [data-testid="stFileUploader"] section:has([data-testid="stUploadedFile"]) > [data-testid="stFileUploaderDropzone"],
    [data-testid="stFileUploader"] section:has(.uploadedFile) > [data-testid="stFileUploaderDropzone"],
    [data-testid="stFileUploader"] section:has(li) > [data-testid="stFileUploaderDropzone"] {
        display: none !important;
        height: 0 !important;
        overflow: hidden !important;
        padding: 0 !important;
        margin: 0 !important;
        opacity: 0 !important;
        pointer-events: none !important;
    }
    /* Método 2: Ocultar botones pequeños/minimal que actúan como "+" */
    [data-testid="stFileUploader"] [data-testid="stBaseButton-minimal"],
    [data-testid="stFileUploader"] button[data-testid="baseButton-minimal"] {
        display: none !important;
        visibility: hidden !important;
        width: 0 !important;
        height: 0 !important;
        overflow: hidden !important;
        position: absolute !important;
        pointer-events: none !important;
    }
    /* Método 3: Ocultar cualquier botón secundario dentro del área de archivo subido */
    [data-testid="stFileUploader"] [data-testid="stUploadedFileLayout"] ~ [data-testid="stFileUploaderDropzone"],
    [data-testid="stFileUploader"] [data-testid="stUploadedFile"] ~ [data-testid="stFileUploaderDropzone"] {
        display: none !important;
    }
    /* Método 4: Si el dropzone contiene solo un icono pequeño (el "+"), ocultarlo */
    [data-testid="stFileUploaderDropzone"]:has(+ [data-testid="stUploadedFileLayout"]) {
        display: none !important;
    }
    /* Método 5: Forzar que solo se vea 1 botón: el Browse files O el archivo */
    [data-testid="stFileUploader"] section > div:first-child:not(:only-child) {
        display: none !important;
    }
    
    /* Botón X (eliminar archivo) - SIEMPRE VISIBLE */
    [data-testid="stFileUploader"] [data-testid="stFileUploaderDeleteBtn"],
    [data-testid="stFileUploader"] button[aria-label="Delete file"],
    [data-testid="stFileUploader"] .stFileUploaderFile button,
    [data-testid="stFileUploader"] button[kind="icon"] {
        display: inline-flex !important;
        visibility: visible !important;
        position: relative !important;
        clip: auto !important;
        width: 30px !important;
        height: 30px !important;
        padding: 2px !important;
        margin: 0 4px !important;
        background: #ef4444 !important;
        color: white !important;
        border-radius: 50% !important;
        border: none !important;
        align-items: center !important;
        justify-content: center !important;
        cursor: pointer !important;
        box-shadow: 0 2px 6px rgba(239, 68, 68, 0.3) !important;
        font-size: 0.85rem !important;
        overflow: visible !important;
        opacity: 1 !important;
    }
    [data-testid="stFileUploader"] [data-testid="stFileUploaderDeleteBtn"]:hover,
    [data-testid="stFileUploader"] button[aria-label="Delete file"]:hover {
        background: #dc2626 !important;
        transform: scale(1.1) !important;
    }
    
    /* Texto informativo del uploader */
    [data-testid="stFileUploader"] small {
        color: #666;
        font-size: 0.85rem;
    }
    
    /* Contenedores personalizados para los uploaders */
    .upload-container {
        background: white;
        border-radius: 14px;
        padding: 20px;
        box-shadow: 0 2px 12px rgba(0,0,0,0.08);
        border: 1px solid rgba(17, 127, 124, 0.15);
    }
    .upload-header {
        display: flex;
        align-items: center;
        gap: 8px;
        margin-bottom: 12px;
        font-weight: 700;
        font-size: 1rem;
        color: #0D5F5D;
    }
    .upload-header-icon {
        font-size: 1.2rem;
    }
    
    /* ========== USER BAR ========== */
    .user-bar {
        background: linear-gradient(135deg, #0D5F5D 0%, #117F7C 100%);
        color: white;
        padding: 8px 20px;
        border-radius: 10px;
        display: flex;
        justify-content: space-between;
        align-items: center;
        margin-bottom: 15px;
        font-size: 0.85rem;
        box-shadow: 0 3px 12px rgba(17,127,124,0.25);
    }
    
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    </style>
""", unsafe_allow_html=True)

# --- 2. DICCIONARIO COMPLETO DE CÓDIGOS ---
CODIGO_DESCRIPCIONES = {
    '2HBY': 'Hotel standby 2 días',
    '2SBY': 'Bloque standby 2 días',
    '3HBY': 'Hotel standby 3 días',
    '3SBY': 'Bloque standby 3 días',
    'AA': 'Actividad aeroportuaria',
    'AHUE': 'Adherido a la Huelga',
    'AOFI': 'Oficina en Aeropuerto mañana',
    'AOFT': 'Oficina en Aeropuerto tarde',
    'ART': 'Formación recurrente anual',
    'ASBY': 'Standby en aeropuerto',
    'ASEP': 'Reconocimiento médico Asepeyo',
    'BOD': 'Permiso Boda',
    'CDOF': 'Libre recuperado mes siguiente',
    'CISC': 'Curso estandarización CRMI',
    'CUR': 'Curso',
    'DANG': 'Mercancías peligrosas',
    'DEOF': 'Libre recuperado mismo mes',
    'DOFF': 'Día OFF en destacamento',
    'DS': 'Día sin cambios',
    'DSD': 'Día sin cambios daytoday',
    'E': 'Actividad corporativa',
    'EF': 'Franco ex ruta',
    'ELAC': 'Días extras de lactancia',
    'ELAT': 'Días de Lactancia (bloque)',
    'EMB': 'Embarazo',
    'EOFF': 'Día libre extra (+18 años)',
    'ESBY': 'Standby vuelos extendidos',
    'F': 'Franco',
    'F2': 'Franco futuro',
    'FALT': 'Falta (sanción)',
    'FDHC': 'Deadhead (posicionamiento)',
    'FDUT': 'Actividad de vuelo',
    'FOFF': 'Día libre en festivo',
    'FR': 'Servicio de Franco',
    'FTG': 'Fatiga',
    'FTRN': 'Entrenamiento de vuelo',
    'FW': 'Flexiworking',
    'FZ': 'Cambio forzoso',
    'HFR': 'Franco en hotel',
    'HIGH': 'Horas altas - no usar',
    'HSBY': 'Standby en hotel',
    'IET': 'Instructor en tierra',
    'INC': 'Incidencia',
    'ITCP': 'Instrucción TCP',
    'J2SB': 'JC bloque 2 SBY',
    'J3SB': 'TCP realizando JC 3SBY',
    'JHSB': 'TCP realizando JC HSBY',
    'JRSB': 'TCP realizando JC RSBY',
    'JTSB': 'CC Flexible Home SBY',
    'LAC': 'Lactancia',
    'LFR': 'Franco regulable en línea',
    'LMAE': 'Pase aeroportuario',
    'LMCC': 'Código de conducta',
    'LMCS': 'eLearning Ciberseguridad',
    'LMCX': 'eLearning Experiencia Cliente',
    'LMGG': 'eLearning Gate Gourmet',
    'LMIA': 'eLearning IAG training shield',
    'LMPE': 'eLearning Directiva AESA',
    'LMS9': 'eLearning cabin familiarizator',
    'LMSB': 'Formación manipulación alimentos',
    'LMSC': 'eLearning CBT',
    'LMSE': 'eLearning Seguridad',
    'LMSG': 'eLearning Mercancías peligrosas',
    'LMSW': 'eLearning WIFI',
    'LMSZ': 'eLearning mes en curso',
    'LPRL': 'eLearning prevención riesgos laborales',
    'LSCK': 'Enfermedad larga',
    'LSIC': 'Enfermedad en línea',
    'MAT': 'Permiso de maternidad',
    'MED': 'Revisión médica',
    'MGS': 'Reunión sindical',
    'MGSF': 'Reunión sindical (FR)',
    'MTG': 'Reunión',
    'MUD': 'Permiso por mudanza',
    'N/S': 'No Show',
    'NANQ': 'No cualificado para volar',
    'NAVL': 'No disponible',
    'NJSK': 'Baja no justificada',
    'NOFF': 'Día libre no compensable',
    'NQ': 'No cualificado',
    'NROF': 'ROFF no ganado en subasta',
    'NVAC': 'Vacaciones compensables',
    'OABY': 'Fuera de standby aeropuerto',
    'OCC': 'Curso conversión operador',
    'OCCL': 'OCC curso largo',
    'OCCR': 'OCC reincorporados',
    'OFF': 'Día libre',
    'OFIC': 'Oficina',
    'OFIT': 'Oficina departamento formación',
    'OSB2': 'Fuera de standby bloque 2',
    'OSB3': 'Fuera de standby bloque 3',
    'OSBY': 'Fuera de standby',
    'PAT': 'Permiso de paternidad',
    'PER': 'Permiso no retribuido',
    'PPA': 'Permiso parental no retribuido',
    'PPAI': 'PPA invadido',
    'PQP': 'Plan de cualificación',
    'PT': 'Part time (días libres)',
    'PTD': 'PT devuelto',
    'PTF': 'PT flexible',
    'PTI': 'Part time invadido',
    'PUR': 'Curso de sobrecargo',
    'PVAC': 'Vacaciones pendientes',
    'REC': 'Examen de recuperación',
    'REST': 'Descanso compensable',
    'REU': 'Reunión retribuida',
    'REUT': 'Reunión retribuida (formación)',
    'RF': 'Franco restringido',
    'RFSM': 'Refresco obligatorio',
    'RFT': 'Reservado para formación',
    'ROFF': 'Día libre solicitado',
    'RPER': 'Permiso retribuido',
    'RPUR': 'Permiso urgente retribuido',
    'RRES': 'Descanso recurrente',
    'RSBY': 'Standby ex ruta',
    'RSF': 'Standby ex ruta + vuelo asignado',
    'RSV': 'Período pre-roster',
    'RVAC': 'Vacaciones voluntarias',
    'RVCC': 'Revacunación',
    'SBY': 'Standby desde casa',
    'SBYA': 'SBY vacuna fiebre amarilla',
    'SEC': 'Seguridad',
    'SICD': 'Enfermo desde servicio',
    'SICK': 'Enfermedad',
    'SIND': 'Horas sindicales',
    'SINF': 'Horas sindicales (FR)',
    'SOFF': 'ROFF personal oficina',
    'SROF': 'ROFF ganado en subasta',
    'SSMM': 'Servicios mínimos',
    'TAX': 'Taxi',
    'TRT': 'Formación recurrente trianual',
    'TSBY': 'Standby tripulante junior',
    'UNFD': 'Fatiga durante servicio',
    'UNFT': 'No apto para volar',
    'VAC': 'Vacaciones forzosas',
    'VICC': 'Curso de instructor',
    'XOF1': 'Libre navidad prioridad 1',
    'XOF2': 'Libre navidad prioridad 2',
    'XOF3': 'Libre navidad prioridad 3',
    'XSOF': 'Libre navidad sin prioridad',
}

# --- 3. SISTEMA DE CATEGORÍAS ---
CATEGORIAS_PRINCIPALES = {
    'vuelo': {
        'nombre': 'Actividad de Vuelo',
        'icono': '✈️',
        'color': 'cat-vuelo',
        'codigos': ['FDUT', 'FDHC', 'RSF', 'FR', 'HFR', 'SSMM'],
        'incluye_vuelos': True
    },
    'guardias': {
        'nombre': 'Guardias / Standby',
        'icono': '⏰',
        'color': 'cat-guardias',
        'codigos': ['SBY', 'ASBY', 'HSBY', 'RSBY', 'ESBY', 'TSBY', '2SBY', '3SBY', '2HBY', '3HBY',
                    'J2SB', 'J3SB', 'JHSB', 'JRSB', 'JTSB', 'OSB2', 'OSB3', 'OSBY', 'OABY', 'SBYA']
    },
    'libres': {
        'nombre': 'Días Libres / Descanso',
        'icono': '🟢',
        'color': 'cat-libres',
        'codigos': ['OFF', 'ROFF', 'SROF', 'NROF', 'NOFF', 'EOFF', 'FOFF', 'DOFF',
                    'REST', 'RRES', 'LFR', 'XOF1', 'XOF2', 'XOF3', 'XSOF', 'CDOF', 'DEOF', 'SOFF']
    },
    'disponibilidad': {
        'nombre': 'Disponibilidad Empresa',
        'icono': '🔵',
        'color': 'cat-guardias',
        'codigos': ['F', 'F2', 'RF', 'EF']
    },
    'formacion': {
        'nombre': 'Formación / Cursos',
        'icono': '📚',
        'color': 'cat-formacion',
        'codigos': ['FTRN', 'ART', 'TRT', 'RFSM', 'RFT', 'CUR', 'PUR', 'CISC', 'OCC', 'OCCL', 'OCCR', 'VICC',
                    'ITCP', 'IET', 'PQP', 'REC', 'DANG', 'LMSG', 'SEC', 'LMSE', 'LMSB', 'REUT',
                    'LMAE', 'LMCC', 'LMCS', 'LMCX', 'LMGG', 'LMIA', 'LMPE', 'LMS9', 'LMSC', 'LMSW', 'LMSZ', 'LPRL']
    },
    'oficina': {
        'nombre': 'Oficina / Reuniones',
        'icono': '🏢',
        'color': 'cat-oficina',
        'codigos': ['OFIC', 'AOFI', 'AOFT', 'OFIT', 'MTG', 'REU', 'MGS', 'MGSF', 'SIND', 'SINF', 'AHUE', 'AA', 'E']
    },
    'medico': {
        'nombre': 'Médico / Salud',
        'icono': '🏥',
        'color': 'cat-medico',
        'codigos': ['MED', 'ASEP', 'RVCC', 'SICK', 'SICD', 'LSIC', 'LSCK', 'NJSK', 'UNFT', 'FTG', 'UNFD']
    },
    'permisos': {
        'nombre': 'Permisos Personales',
        'icono': '👤',
        'color': 'cat-permisos',
        # Solo estos 5 tipos específicos de permisos en calendario diario
        'codigos': ['PER', 'RPER', 'BOD', 'MUD', 'RPUR']
    },
    'permisos_otros': {
        'nombre': 'Otros Permisos',
        'icono': '👶',
        'color': 'cat-permisos',
        'codigos': ['PAT', 'MAT', 'EMB', 'LAC', 'ELAC', 'ELAT', 'PPA', 'PPAI']
    },
    'vacaciones': {
        'nombre': 'Vacaciones',
        'icono': '🌴',
        'color': 'cat-vacaciones',
        'codigos': ['VAC', 'RVAC', 'NVAC', 'PVAC']
    },
    'parttime': {
        'nombre': 'Reducciones / Part-Time',
        'icono': '⏸️',
        'color': 'cat-parttime',
        'codigos': ['PT', 'PTF', 'PTD', 'PTI']
    },
    'incidencias': {
        'nombre': 'Incidencias',
        'icono': '⚠️',
        'color': 'cat-incidencias',
        'codigos': ['N/S', 'FALT', 'INC', 'FZ', 'DS', 'DSD', 'HIGH', 'RSV', 'FW']
    }
}

# Conceptos para porcentajes globales
CONCEPTOS_GLOBALES = {
    'Reducciones (Part Time)': {
        'codigos': ['PT', 'PTD', 'PTF', 'PTI'],
        'color': '#117F7C',
        'icono': '⏸️'
    },
    'Vacaciones': {
        'codigos': ['VAC', 'RVAC', 'NVAC', 'PVAC'],
        'color': '#44BABC',
        'icono': '🌴'
    },
    'Días Libres': {
        'codigos': ['OFF', 'ROFF', 'SROF', 'NOFF', 'EOFF', 'FOFF', 'REST', 'DOFF',
                    'NROF', 'LFR', 'RRES', 'XOF1', 'XOF2', 'XOF3', 'XSOF', 'CDOF', 'DEOF', 'SOFF'],
        'color': '#2196F3',
        'icono': '🟢'
    },
    'Disponibilidad Empresa': {
        'codigos': ['F', 'F2', 'RF', 'EF'],
        'color': '#1565C0',
        'icono': '🔵'
    },
    'Formación': {
        'codigos': ['FTRN', 'ART', 'TRT', 'RFSM', 'CUR', 'PUR', 'CISC', 'OCC', 'OCCL', 'OCCR', 'VICC',
                    'PQP', 'DANG', 'LMSG', 'SEC', 'LMSE', 'LMSB', 'ITCP', 'IET', 'RFT', 'REC',
                    'LMAE', 'LMCC', 'LMCS', 'LMCX', 'LMGG', 'LMIA', 'LMPE', 'LMS9', 'LMSC', 'LMSW', 'LMSZ', 'LPRL', 'REUT'],
        'color': '#61C1C3',
        'icono': '📚'
    },
    'Guardias (Standby)': {
        'codigos': ['SBY', 'ASBY', 'HSBY', 'RSBY', 'ESBY', 'TSBY', '2SBY', '3SBY', 'J2SB', 'J3SB',
                    '2HBY', '3HBY', 'OSB2', 'OSB3', 'OSBY', 'OABY', 'JHSB', 'JRSB', 'JTSB', 'SBYA'],
        'color': '#489999',
        'icono': '⏰'
    },
    'Bajas Médicas': {
        'codigos': ['MED', 'ASEP', 'SICK', 'SICD', 'LSIC', 'LSCK', 'NJSK', 'UNFT', 'FTG', 'UNFD'],
        'color': '#E74C3C',
        'icono': '🏥'
    },
    'Permisos Personales': {
        'codigos': ['BOD', 'PAT', 'MAT', 'LAC', 'ELAC', 'ELAT', 'MUD', 'RPER', 'PER', 'RPUR', 'PPA', 'PPAI', 'EMB'],
        'color': '#0B847F',
        'icono': '👤'
    }
}

def get_categoria_principal(code):
    c = str(code).strip().upper()
    if c.isdigit() and len(c) == 4:
        return 'vuelo'
    for cat_id, cat_info in CATEGORIAS_PRINCIPALES.items():
        if c in cat_info['codigos']:
            return cat_id
    return 'otros'

def get_codigo_descripcion(code):
    c = str(code).strip().upper()
    if c == '--' or c == '-' or c == '':
        return "Sin asignación"
    if c in CODIGO_DESCRIPCIONES:
        return CODIGO_DESCRIPCIONES[c]
    elif c.isdigit() and len(c) == 4:
        return f"Vuelo VY{c}"
    else:
        return c

def get_visual_style(code):
    """Returns (css_class, icon, label) for a given code.
    Full color scheme:
    - Vuelos (4 digits): Rojo
    - F/F2: Azul (franco)
    - PT/PTF/PTD/PTI: Rosa (reducción)
    - SIND/SINF/MGS/MGSF/OFIC/AOFI/AOFT/IET/ITCP: Marrón (sindical/oficina)
    - AA/SBY/ASBY/HSBY/RSBY/ESBY/TSBY/2SBY/3SBY/2HBY/3HBY/J2SB/J3SB/JHSB/JRSB/JTSB/OSB2/OSB3/OSBY/OABY/SBYA: Azul oscuro (guardia)
    - Formaciones (ART/TRT/CUR/FTRN/PUR/CISC/OCC/LM*/etc): Amarillo
    """
    c = str(code).strip().upper()
    
    # Vuelos (4 dígitos) -> Rojo
    if c.isdigit() and len(c) == 4:
        return "dc-vuelo", "✈️", "VUELO"
    
    # Disponibilidad Empresa codes -> Azul
    if c in ['F', 'F2', 'RF', 'EF']:
        return "dc-franco", "🔵", "DISPONIBILIDAD"
    
    # Part-time / reductions -> Rosa
    if c in ['PT', 'PTF', 'PTD', 'PTI']:
        return "dc-pt", "⏸️", "REDUCCIÓN"
    
    # Sindicales y oficina/instrucción en tierra -> Marrón
    if c in ['SIND', 'SINF', 'MGS', 'MGSF', 'OFIC', 'AOFI', 'AOFT', 'IET', 'ITCP']:
        return "dc-sindical", "🏢", "OFICINA/SIND."
    
    # Actividad aeroportuaria e imaginarias -> Azul oscuro
    if c in ['AA', 'SBY', 'ASBY', 'HSBY', 'RSBY', 'ESBY', 'TSBY', '2SBY', '3SBY',
             '2HBY', '3HBY', 'J2SB', 'J3SB', 'JHSB', 'JRSB', 'JTSB', 'OSB2', 'OSB3',
             'OSBY', 'OABY', 'SBYA']:
        return "dc-guardia", "⏰", "GUARDIA"
    
    # Formaciones -> Amarillo
    if c in ['ART', 'TRT', 'CUR', 'FTRN', 'PUR', 'CISC', 'OCC', 'OCCL', 'OCCR', 'VICC',
             'PQP', 'REC', 'RFSM', 'RFT', 'DANG', 'SEC', 'REUT',
             'LMAE', 'LMCC', 'LMCS', 'LMCX', 'LMGG', 'LMIA', 'LMPE', 'LMS9',
             'LMSB', 'LMSC', 'LMSE', 'LMSG', 'LMSW', 'LMSZ', 'LPRL']:
        return "dc-formacion", "📚", "FORMACIÓN"
    
    # Fallback por categoría
    cat = get_categoria_principal(c)
    estilos = {
        'vuelo': ("dc-vuelo", "✈️", "VUELO"),
        'guardias': ("dc-guardia", "⏰", "GUARDIA"),
        'libres': ("dc-verde", "🟢", "LIBRE"),
        'formacion': ("dc-formacion", "📚", "FORMACIÓN"),
        'oficina': ("dc-sindical", "🏢", "OFICINA"),
        'medico': ("dc-sick", "🏥", "MÉDICO"),
        'permisos': ("dc-permiso", "👤", "PERMISO"),
        'vacaciones': ("dc-vac", "🌴", "VACACIONES"),
        'parttime': ("dc-pt", "⏸️", "REDUCCIÓN"),
        'incidencias': ("dc-ftg", "⚠️", c),
        'otros': ("dc-otro", "📋", c)
    }
    return estilos.get(cat, ("dc-otro", "📋", c))

def get_fechas_texto(indices):
    if not indices:
        return ""
    indices = sorted(list(set(indices)))
    ranges = []
    start = indices[0]
    for i in range(1, len(indices)):
        if indices[i] != indices[i-1] + 1:
            ranges.append((start, indices[i-1]))
            start = indices[i]
    ranges.append((start, indices[-1]))
    res = []
    for r in ranges:
        if r[0] == r[1]:
            res.append(f"{r[0]:02d}")
        else:
            res.append(f"{r[0]:02d}-{r[1]:02d}")
    return ", ".join(res)

def get_color_categoria(cat):
    colores = {
        'vuelo': '#117F7C',
        'guardias': '#489999',
        'libres': '#44BABC',
        'formacion': '#61C1C3',
        'oficina': '#0B847F',
        'medico': '#E74C3C',
        'permisos': '#489999',
        'vacaciones': '#44BABC',
        'parttime': '#117F7C',
        'incidencias': '#61C1C3',
        'otros': '#489999'
    }
    return colores.get(cat, '#117F7C')

def get_nombre_dia_semana(dia, mes, anio):
    try:
        dias_es = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
        from datetime import date
        d = date(anio, mes, dia)
        return dias_es[d.weekday()]
    except:
        return ""

def es_fin_de_semana(dia, mes, anio):
    try:
        from datetime import date
        d = date(anio, mes, dia)
        return d.weekday() >= 5
    except:
        return False

def parse_hours(h):
    if h and h != '--':
        try:
            parts = h.split(':')
            return int(parts[0]) + int(parts[1]) / 60
        except:
            return 0
    return 0


def format_duration(hours_decimal):
    """Formatea duración decimal en formato 'Xh YYmin'."""
    hours = int(hours_decimal)
    minutes = int(round((hours_decimal - hours) * 60))
    if minutes == 60:
        hours += 1
        minutes = 0
    return f"{hours}h {minutes:02d}min"


def format_diff_hm(hours_decimal):
    """Formatea diferencia de horas decimales como '+Xh YYmin' / '-Xh YYmin' / '0h 00min'."""
    if abs(hours_decimal) < 0.001:
        return "0h 00min"
    sign = "+" if hours_decimal > 0 else "-"
    abs_hours = abs(hours_decimal)
    h = int(abs_hours)
    m = int(round((abs_hours - h) * 60))
    if m == 60:
        h += 1
        m = 0
    return f"{sign}{h}h {m:02d}min"


def _safe_rewind_files(file_list):
    """Rebobina archivos tipo UploadedFile/BytesIO sin romper flujo."""
    for fo in file_list:
        if fo is not None and hasattr(fo, 'seek'):
            try:
                fo.seek(0)
            except Exception:
                pass


def _release_uploaded_files(file_list, clear_widget_keys=False):
    """Libera referencias de archivos subidos para reducir uso de memoria.

    - Cierra objetos archivo si exponen `.close()`.
    - Elimina claves de widgets uploader en session_state si se solicita.
    - Fuerza `gc.collect()` al final.
    """
    for fo in file_list or []:
        if fo is None:
            continue
        try:
            if hasattr(fo, 'close'):
                fo.close()
        except Exception:
            pass

    if clear_widget_keys:
        for key in ('ini_a', 'ini_b', 'fin_a', 'fin_b', 'inicial_pdf', 'final_pdf'):
            if key in st.session_state:
                del st.session_state[key]

    gc.collect()


# --- Regex precompiladas globales para reducir coste por llamada ---
REGEX_BASE_LINE = re.compile(r'^\s*([A-Z]{3})\s+(\S+)\s+(JC|TC)\s+(.*)')
REGEX_TIME_POS = re.compile(r'(\d{1,2}):(\d{2})')
REGEX_CODE_POS = re.compile(r'\*\s?[A-Z]{2,5}\d?|[A-Z]{2,5}\d?')
REGEX_AIRPORT = re.compile(r'^\*?\s?[A-Z]{3}$')
REGEX_DAY_HEADER = re.compile(r'(ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)(\d{2})')
REGEX_BLOCK_TOTAL = re.compile(r'(\d{2,6})\s+BLOCK\s*>\s*(\d+):(\d{2})')
REGEX_ID_LINE = re.compile(r'^\s*(\d{2,6})\s')
REGEX_SEPARATOR = re.compile(r'[_]{20,}')
REGEX_TIME_ONLY = re.compile(r'^\d{1,2}:\d{2}$')

# Umbral de cache para evitar duplicar en memoria PDFs muy grandes.
CACHE_BYPASS_BYTES = int(os.getenv("CACHE_BYPASS_MB", "120")) * 1024 * 1024


def _estimate_file_size(fo):
    """Estimación robusta del tamaño del archivo sin cargarlo completo en RAM."""
    if fo is None:
        return 0

    size_attr = getattr(fo, 'size', None)
    if isinstance(size_attr, int) and size_attr >= 0:
        return size_attr

    if hasattr(fo, 'getbuffer'):
        try:
            return len(fo.getbuffer())
        except Exception:
            pass

    if hasattr(fo, 'seek') and hasattr(fo, 'tell'):
        try:
            current = fo.tell()
            fo.seek(0, os.SEEK_END)
            size = fo.tell()
            fo.seek(current)
            return max(size, 0)
        except Exception:
            return 0

    return 0


def _estimate_total_size(file_list):
    return sum(_estimate_file_size(fo) for fo in (file_list or []) if fo is not None)


def _files_to_cache_payload(file_list):
    """Convierte UploadedFile/BytesIO a payload hashable para cache de Streamlit."""
    payload = []
    for fo in file_list or []:
        if fo is None:
            continue
        try:
            if hasattr(fo, 'seek'):
                fo.seek(0)
            content = fo.read() if hasattr(fo, 'read') else bytes(fo)
            if hasattr(fo, 'seek'):
                fo.seek(0)
            name = getattr(fo, 'name', 'uploaded.pdf')
            payload.append((name, content))
        except Exception:
            continue
    return tuple(payload)


def _payload_to_fileobjs(payload):
    return [io.BytesIO(content) for _, content in payload]


# === SECCIÓN 4: FUNCIONES CACHEADAS DE PROCESAMIENTO ===
# --- 3a. EXTRACCIÓN DE LONG DUTIES (>9h 40min) ---
# Umbral para Long Duties: 9h 40min (ya incluye 20min tras calzos)
UMBRAL_LONG_DUTY = 9 + (40 / 60)  # = 9.666... horas

def _extract_daily_duty_data_impl(file_list, progress_callback=None):
    """Extrae datos de duty diario con lógica basada en columnas del PDF.
    
    Para cada día de cada tripulante:
      1. Localizar TODOS los tokens (horas y códigos) en la columna de ese día
      2. Si el día contiene un código administrativo/tierra → EXCLUIR día
      3. Si el día contiene un código de imaginaria (SBY/SB2):
         → hora_inicio = primera hora DESPUÉS del código de imaginaria
         → hora_fin = última hora del día
      4. Si NO hay imaginaria:
         → hora_inicio = primera hora del día
         → hora_fin = última hora del día
      5. Duración = FIN - INICIO
      6. Si duración > 9h 40min → Long Duty
    
    Returns:
        dict: {crew_id: {
            'daily_duties': [(day_num, duty_hours_float), ...],
            'long_duty_days': [(day_num, duty_hours_float, ruta_str), ...],  # >9h40min only, con ruta
        }}
    """
    re_base_line = REGEX_BASE_LINE
    re_time_pos = REGEX_TIME_POS
    re_code_pos = REGEX_CODE_POS  # Códigos: OFIC, OSBY, SB2, *TFN, * BCN (con/sin espacio)
    re_airport = REGEX_AIRPORT  # Aeropuertos: AGP, MAD, BCN, *TFN, * BCN...
    re_day_header = REGEX_DAY_HEADER
    re_block_total = REGEX_BLOCK_TOTAL
    re_id_line = REGEX_ID_LINE
    re_separator = REGEX_SEPARATOR
    
    # Códigos administrativos/tierra/formación → EXCLUIR día completo (solo cuantificar vuelo)
    codigos_tierra = {
        # Oficina / Tierra / Reuniones
        'OFIC', 'OFI', 'AOFI', 'AOFT', 'MTG', 'SIND', 'SINF', 'OFF',
        'MGS', 'MGSF', 'REU', 'REUT',
        # Formación / Instrucción
        'ART', 'TRT', 'CUR', 'FTRN', 'IET', 'ITCP',
        'CISC', 'OCC', 'PQP', 'REC',
        # Formaciones adicionales
        'RFSM', 'RFT', 'PUR', 'OCCL', 'OCCR', 'VICC',
        'DANG', 'LMSG', 'SEC', 'LMSE', 'LMSB',
        # Last Minute / Cambios de última hora (todos son tierra/formación)
        'LMSZ', 'LMGG', 'LMSC', 'LMIA', 'LMGR',
        # Disponibilidad / Libranza / No operativos
        'RSV', 'F', 'FR', 'RF', 'NOFF', 'NOFR', 'ROFF', 'SROF',
        # Permisos / Vacaciones / Bajas
        'PT', 'PTF', 'VAC', 'RVAC', 'PVAC',
        'SICK', 'LSCK', 'MAT', 'EMB', 'LAC',
        'EOFF', 'DEOF', 'RPER', 'NQ', 'FW',
    }
    
    # Códigos de transporte terrestre → NO saltar el día, pero excluir tiempos de transporte
    codigos_transporte = {
        'TAXI', 'TREN', 'BUS', 'LMAP', 'CML',
    }
    
    # Patrones de imaginaria → ajustar hora de inicio
    # Lista explícita de códigos de imaginaria conocidos
    imaginaria_codes_explicit = {
        'SBY', 'ASBY', 'HSBY', 'RSBY', 'ESBY', 'TSBY',
        '2SBY', '3SBY', 'SB2', 'SB3',
        'OSBY', 'OSB2', 'OSB3',  # Variantes con prefijo O (Out of Standby)
    }
    
    def es_imaginaria(code):
        """Detecta si un código es de imaginaria (standby).
        Primero comprueba la lista explícita, luego busca patrones SBY/SB2/SB3."""
        if code in imaginaria_codes_explicit:
            return True
        if 'SBY' in code or 'SB2' in code or 'SB3' in code:
            return True
        return False
    
    def format_duration(hours_decimal):
        """Formatea duración decimal en formato 'Xh YYmin'."""
        hours = int(hours_decimal)
        minutes = int(round((hours_decimal - hours) * 60))
        return f"{hours}h {minutes:02d}min"
    
    all_crew_data = {}

    total_pages_est = 0
    for _fo in file_list:
        if _fo is None:
            continue
        try:
            _fo.seek(0)
            total_pages_est += len(PyPDF2.PdfReader(_fo).pages)
        except Exception:
            pass
        finally:
            try:
                _fo.seek(0)
            except Exception:
                pass

    processed_pages = 0

    for file_obj in file_list:
        if file_obj is None:
            continue
        try:
            file_obj.seek(0)
        except Exception:
            pass

        reader = PyPDF2.PdfReader(file_obj)

        for page_idx in range(len(reader.pages)):
            processed_pages += 1
            if progress_callback and (processed_pages % 20 == 0 or processed_pages == total_pages_est):
                progress_callback(processed_pages, max(total_pages_est, 1), "Extrayendo long duties (páginas PDF)")
            if processed_pages % 120 == 0:
                gc.collect()

            text = reader.pages[page_idx].extract_text()
            if not text:
                continue
            
            lines = text.split('\n')
            
            # --- Encontrar posiciones de columna de cada día desde la cabecera ---
            header_day_positions = {}  # {day_num: col_start}  (posición en la cabecera)
            col_width = 5  # ancho por defecto
            for line in lines[:10]:
                matches = list(re_day_header.finditer(line))
                if len(matches) >= 20:
                    for i, m_hdr in enumerate(matches):
                        day_num = int(m_hdr.group(2))
                        header_day_positions[day_num] = m_hdr.start()
                    # Calcular ancho medio de columna
                    if len(matches) >= 2:
                        col_width = (matches[-1].start() - matches[0].start()) / (len(matches) - 1)
                    break
            
            if not header_day_positions:
                continue
            
            header_day1_pos = header_day_positions.get(1, min(header_day_positions.values()))
            
            # --- Calcular offset entre cabecera y datos ---
            offset = 0
            for line in lines:
                sm = re_base_line.match(line)
                if sm:
                    data_start = sm.start(4)
                    offset = data_start - header_day1_pos
                    break
            
            # Construir columnas ajustadas: {day_num: center_position}
            day_centers = {}
            for day_num, hpos in header_day_positions.items():
                day_centers[day_num] = hpos + offset + col_width / 2
            
            # --- Encontrar bloques de tripulante separados por líneas ___ ---
            blocks = []
            current_block_start = 0
            for i, line in enumerate(lines):
                if re_separator.search(line) and len(line.strip()) > 20:
                    if i > current_block_start:
                        blocks.append((current_block_start, i))
                    current_block_start = i + 1
            if current_block_start < len(lines):
                blocks.append((current_block_start, len(lines)))
            
            for block_start, block_end in blocks:
                block_lines = lines[block_start:block_end]
                if len(block_lines) < 3:
                    continue
                
                # Buscar línea de schedule (BASE FLEET CAT codes...)
                schedule_idx = None
                for bi, bl in enumerate(block_lines):
                    if re_base_line.match(bl.strip()):
                        schedule_idx = bi
                        break
                
                if schedule_idx is None:
                    continue
                
                # Buscar ID del tripulante
                crew_id = None
                for bi in range(schedule_idx + 1, min(len(block_lines), schedule_idx + 8)):
                    bl = block_lines[bi]
                    bt = re_block_total.search(bl)
                    if bt:
                        crew_id = int(bt.group(1))
                        break
                    if crew_id is None:
                        id_m = re_id_line.match(bl.strip())
                        if id_m and 'BLOCK' not in bl.upper() and 'DUTY' not in bl.upper() and 'FLTS' not in bl.upper():
                            crew_id = int(id_m.group(1))
                
                if crew_id is None:
                    continue
                
                # --- Recoger TODOS los tokens (horas y códigos) por día ---
                # Cada token tiene: posición en línea, tipo ('time'/'code'), valor, y orden global
                day_tokens = {}  # {day_num: [{'type': 'time'|'code', 'value': ..., 'pos': int, 'order': int}, ...]}
                half_col = col_width / 2 + 1
                global_order = 0
                
                # PRIMERO: Extraer códigos de imaginaria de la LÍNEA DE SCHEDULE
                # Los códigos de imaginaria pueden aparecer concatenados con números de vuelo
                # Ej: "3905SBY", "2SBY", "ESBY", "OSBY", "OSB2"
                # Usamos regex que captura cualquier token que contenga SBY o SB2
                schedule_line = block_lines[schedule_idx]
                sm_sched = re_base_line.match(schedule_line.strip())
                # Extraer BASE del tripulante (primer grupo de re_base_line: 3 letras)
                base_tripulante = sm_sched.group(1) if sm_sched else None
                if sm_sched:
                    sched_data = sm_sched.group(4)  # Parte de datos después de BASE FLEET CAT
                    sched_data_start = schedule_line.find(sched_data)
                    
                    # Regex para encontrar imaginarias (incluso concatenadas con números)
                    re_imag_sched = re.compile(r'[A-Z0-9]*(?:SBY|SB[23])')
                    for m in re_imag_sched.finditer(sched_data):
                        code = m.group().upper()
                        if not es_imaginaria(code):
                            continue
                        code_center = sched_data_start + m.start() + len(m.group()) / 2
                        best_day = None
                        best_dist = float('inf')
                        for day_num, center in day_centers.items():
                            dist = abs(code_center - center)
                            if dist < best_dist:
                                best_dist = dist
                                best_day = day_num
                        if best_day is not None and best_dist <= half_col:
                            day_tokens.setdefault(best_day, []).append({
                                'type': 'code',
                                'value': code,
                                'pos': sched_data_start + m.start(),
                                'order': -1  # Antes de tokens de detalle
                            })
                    
                    # También buscar códigos de tierra en la línea de schedule
                    # Busca tanto códigos standalone como concatenados (ej: 3905OFIC, 1234ART)
                    # Patrón 1: códigos puros de 2-5 letras
                    re_sched_code = re.compile(r'[A-Z]{2,5}')
                    for m in re_sched_code.finditer(sched_data):
                        code = m.group().upper()
                        if code in codigos_tierra:
                            code_center = sched_data_start + m.start() + len(m.group()) / 2
                            best_day = None
                            best_dist = float('inf')
                            for day_num, center in day_centers.items():
                                dist = abs(code_center - center)
                                if dist < best_dist:
                                    best_dist = dist
                                    best_day = day_num
                            if best_day is not None and best_dist <= half_col:
                                day_tokens.setdefault(best_day, []).append({
                                    'type': 'code',
                                    'value': code,
                                    'pos': sched_data_start + m.start(),
                                    'order': -1
                                })
                    
                    # Patrón 2: códigos concatenados con números (ej: 3905OFIC, 1234SIND)
                    # Construir regex dinámico con los códigos tierra
                    tierra_pattern = '|'.join(re.escape(c) for c in codigos_tierra)
                    re_tierra_concat = re.compile(r'[A-Z0-9]*(?:' + tierra_pattern + r')')
                    for m in re_tierra_concat.finditer(sched_data):
                        # Verificar que el match realmente contiene un código tierra
                        matched = m.group().upper()
                        found_code = None
                        for ct in codigos_tierra:
                            if ct in matched:
                                found_code = ct
                                break
                        if found_code:
                            code_center = sched_data_start + m.start() + len(m.group()) / 2
                            best_day = None
                            best_dist = float('inf')
                            for day_num, center in day_centers.items():
                                dist = abs(code_center - center)
                                if dist < best_dist:
                                    best_dist = dist
                                    best_day = day_num
                            if best_day is not None and best_dist <= half_col:
                                # Solo añadir si no hay ya un código tierra para este día
                                existing_codes = [t['value'] for t in day_tokens.get(best_day, []) if t['type'] == 'code']
                                if found_code not in existing_codes:
                                    day_tokens.setdefault(best_day, []).append({
                                        'type': 'code',
                                        'value': found_code,
                                        'pos': sched_data_start + m.start(),
                                        'order': -1
                                    })
                
                for bi in range(schedule_idx + 1, len(block_lines)):
                    bl = block_lines[bi]
                    bl_stripped = bl.strip()
                    
                    # Parar al llegar a la línea BLOCK diaria
                    if bl_stripped.startswith('BLOCK') and '>' not in bl_stripped:
                        break
                    
                    # Líneas resumen con ">" (BLOCK >, DUTY >, FLTS >, etc.)
                    # Contienen tiempos TOTALES en la parte izquierda (resumen),
                    # PERO también contienen datos REALES por día en las columnas
                    # de cada día (ej: BCN de salida, hora de report, etc.).
                    # ═══════════════════════════════════════════════════════
                    # FIX 8 CRÍTICO: Antes se saltaban TODOS los tiempos de
                    # líneas resumen, perdiendo datos reales de vuelo.
                    # Ahora: extraer tiempos que caen en columnas de día válidas
                    # y EXCLUIR solo el tiempo del resumen (parte izquierda).
                    # El resumen total (ej: "48:32" en "DUTY > 48:32") está
                    # ANTES de la primera columna de día.
                    # ═══════════════════════════════════════════════════════
                    is_summary_line = '>' in bl
                    summary_end_pos = 0  # Posición hasta donde llega el resumen
                    if is_summary_line:
                        # Encontrar la posición del ">" y el tiempo/número que le sigue
                        gt_pos = bl.find('>')
                        if gt_pos >= 0:
                            # El resumen ocupa hasta ~10 chars después del ">"
                            # Ejemplo: "BLOCK > 22:32" → resumen termina ~pos 20
                            # Ejemplo: "DUTY  > 48:32" → resumen termina ~pos 20
                            # Ejemplo: "FLTS  >    12" → resumen termina ~pos 18
                            # Todo lo que está a la DERECHA son datos por día
                            summary_end_pos = gt_pos + 10  # Margen generoso
                    
                    # ═══════════════════════════════════════════════════════
                    # FIX 7: Detectar líneas con formato roto (broken alignment)
                    # PyPDF2 a veces extrae líneas sin whitespace, colapsando
                    # columnas. Estas líneas tienen lead<3 Y longitud mucho
                    # menor que las líneas normales (225 chars). Sus posiciones
                    # de columna son incorrectas → NO extraer tiempos de ellas.
                    # Los aeropuertos tampoco son fiables en estas líneas.
                    # ═══════════════════════════════════════════════════════
                    leading_ws = len(bl) - len(bl.lstrip())
                    is_broken_line = (leading_ws < 3 and len(bl) < 150
                                      and not bl_stripped.startswith('BLOCK')
                                      and not bl_stripped.startswith('CREDIT')
                                      and not bl_stripped.startswith('DHD'))
                    
                    # Recoger horas con posición
                    # FIX 8: También de líneas resumen, pero SOLO tiempos en columnas de día
                    # (excluyendo el tiempo total del resumen que está a la izquierda)
                    if not is_broken_line:
                        for m in re_time_pos.finditer(bl):
                            h_val = int(m.group(1))
                            m_val = int(m.group(2))
                            
                            if h_val > 23:
                                continue
                            
                            # En líneas resumen: SOLO extraer tiempos que están
                            # en la zona de columnas de día (después del resumen)
                            if is_summary_line and m.start() < summary_end_pos:
                                continue  # Este tiempo es parte del resumen total → skip
                            
                            time_center = m.start() + len(m.group()) / 2
                            
                            best_day = None
                            best_dist = float('inf')
                            for day_num, center in day_centers.items():
                                dist = abs(time_center - center)
                                if dist < best_dist:
                                    best_dist = dist
                                    best_day = day_num
                            
                            if best_day is not None and best_dist <= half_col:
                                day_tokens.setdefault(best_day, []).append({
                                    'type': 'time',
                                    'value': (h_val, m_val),
                                    'pos': m.start(),
                                    'order': global_order
                                })
                                global_order += 1
                    
                    # Recoger códigos alfabéticos con posición (de TODAS las líneas,
                    # EXCEPTO líneas con formato roto donde las posiciones son incorrectas)
                    if is_broken_line:
                        continue  # Skip toda la línea rota (Fix 7)
                    for m in re_code_pos.finditer(bl):
                        code = m.group()
                        # Normalizar: "* BCN" → "*BCN" (quitar espacio entre * y código)
                        code_upper = re.sub(r'\*\s+', '*', code).upper()
                        
                        # Capturar: tierra, transporte, imaginaria Y aeropuertos (3 letras)
                        is_tierra = code_upper in codigos_tierra
                        is_transporte = code_upper in codigos_transporte
                        is_imag = es_imaginaria(code_upper)
                        is_airport = bool(re_airport.match(code_upper))
                        
                        # En líneas resumen (">"), SOLO capturar aeropuertos.
                        # No capturar códigos tierra/imaginaria porque pueden
                        # pertenecer a otros días en la misma línea resumen.
                        if is_summary_line and not is_airport:
                            continue
                        
                        if not is_tierra and not is_imag and not is_airport and not is_transporte:
                            continue
                        
                        code_center = m.start() + len(m.group()) / 2
                        
                        best_day = None
                        best_dist = float('inf')
                        for day_num, center in day_centers.items():
                            dist = abs(code_center - center)
                            if dist < best_dist:
                                best_dist = dist
                                best_day = day_num
                        
                        if best_day is not None and best_dist <= half_col:
                            if is_transporte:
                                token_type = 'transport'
                            elif is_airport and not is_tierra and not is_imag:
                                token_type = 'airport'
                            else:
                                token_type = 'code'
                            day_tokens.setdefault(best_day, []).append({
                                'type': token_type,
                                'value': code_upper,
                                'pos': m.start(),
                                'order': global_order
                            })
                            global_order += 1
                
                # --- Calcular duty por día con filtros ---
                daily_duties = []
                long_duty_days = []  # Ahora: [(day_num, duty_hours, ruta_str), ...]
                
                # ═══════════════════════════════════════════════════════
                # PASO PREVIO: Detectar overflow de medianoche en TODOS
                # los días ANTES de calcular duties.
                # Cuando un vuelo cruza medianoche (ej: BCN→DSS, BCN→BJL),
                # la hora de llegada aparece en la columna del día siguiente.
                # Debemos:
                #   a) Eliminar esos tiempos del día siguiente
                #   b) Atribuirlos al día anterior para calcular duty correcto
                # ═══════════════════════════════════════════════════════
                overflow_from_previous = {}  # {day_num: {'max_time': (h,m), 'airports': [str], 'tokens': [...]}}
                
                sorted_days = sorted(day_tokens.keys())
                for day_num in sorted_days:
                    tokens_raw = sorted(day_tokens[day_num], key=lambda t: t['order'])
                    times_raw = [t for t in tokens_raw if t['type'] == 'time']
                    
                    # Comprobar si este día tiene un código de imaginaria (SBY/SB2/etc.)
                    # En ese caso, los tiempos tempranos son inicio de standby, NO overflow
                    has_imaginaria_day = any(
                        t['type'] == 'code' and es_imaginaria(t['value'])
                        for t in tokens_raw
                    )
                    
                    if len(times_raw) >= 2:
                        clock_sorted = sorted(times_raw,
                                              key=lambda t: t['value'][0] * 60 + t['value'][1])
                        overflow_boundary = -1
                        for gi in range(len(clock_sorted) - 1):
                            if all(t['value'][0] < 5 for t in clock_sorted[:gi + 1]):
                                curr_min = clock_sorted[gi]['value'][0] * 60 + clock_sorted[gi]['value'][1]
                                next_min = clock_sorted[gi+1]['value'][0] * 60 + clock_sorted[gi+1]['value'][1]
                                gap = next_min - curr_min
                                if gap > 300:
                                    overflow_boundary = gi
                                    break
                            else:
                                break
                        
                        if overflow_boundary >= 0:
                            overflow_times = clock_sorted[:overflow_boundary + 1]
                            overflow_max_order = max(t['order'] for t in overflow_times)
                            # Guardar overflow para atribuir al día anterior
                            max_overflow_time = max(overflow_times, key=lambda t: t['value'][0]*60+t['value'][1])
                            overflow_airports = [t['value'] for t in tokens_raw 
                                                 if t['type'] == 'airport' and t['order'] <= overflow_max_order]
                            
                            # NO guardar como overflow si el día tiene imaginaria
                            # (el tiempo temprano es inicio de standby, no llegada del día anterior)
                            if not has_imaginaria_day:
                                overflow_from_previous[day_num] = {
                                    'max_time': max_overflow_time['value'],
                                    'airports': overflow_airports,
                                    'tokens': [t for t in tokens_raw if t['order'] <= overflow_max_order],
                                }
                            # Limpiar overflow del día actual (siempre, independientemente de imaginaria)
                            day_tokens[day_num] = [t for t in tokens_raw if t['order'] > overflow_max_order]
                    elif len(times_raw) == 1:
                        # Solo 1 tiempo en el día y es < 5:00 AM → posible overflow puro
                        # (día que SOLO tiene la llegada de un vuelo nocturno, sin vuelos propios)
                        single_t = times_raw[0]
                        if single_t['value'][0] < 5 and not has_imaginaria_day:
                            # Verificar que hay al menos un aeropuerto (llegada)
                            airports_raw = [t for t in tokens_raw if t['type'] == 'airport']
                            if airports_raw:
                                overflow_from_previous[day_num] = {
                                    'max_time': single_t['value'],
                                    'airports': [t['value'] for t in airports_raw],
                                    'tokens': list(tokens_raw),
                                }
                                # Limpiar: este día solo era overflow
                                day_tokens[day_num] = [t for t in tokens_raw 
                                                       if t['type'] not in ('time',) or t['value'][0] >= 5]
                
                for day_num in sorted_days:
                    tokens = sorted(day_tokens[day_num], key=lambda t: t['order'])
                    
                    # Separar códigos y horas
                    codes_in_day = [t['value'] for t in tokens if t['type'] in ('code', 'airport')]
                    times_in_day = [t for t in tokens if t['type'] == 'time']
                    
                    # ═══════════════════════════════════════════════════════
                    # 0) VUELOS NOCTURNOS: Comprobar si el día SIGUIENTE
                    # tiene overflow que corresponde a ESTE día
                    # ═══════════════════════════════════════════════════════
                    next_day = day_num + 1
                    overnight_end_time = None  # (h, m) del aterrizaje post-medianoche
                    overnight_airports = []
                    if next_day in overflow_from_previous:
                        # Hay tiempos de madrugada en el día siguiente que son overflow.
                        # Atribuir al día actual si:
                        #   a) El día tiene tiempos >= 13:00 (vuelos de tarde/noche), O
                        #   b) El día tiene al menos 1 tiempo y aeropuertos (actividad real)
                        # El umbral bajo (13h) es necesario porque vuelos de larga
                        # distancia pueden salir a las 14:00 y llegar de madrugada.
                        has_activity = len(times_in_day) >= 1 and any(
                            t['type'] == 'airport' for t in tokens
                        )
                        has_afternoon_times = any(t['value'][0] >= 13 for t in times_in_day)
                        if has_activity and has_afternoon_times:
                            # Validar que el duty resultante es razonable
                            # Calcular duración estimada con overnight
                            earliest_time = min(times_in_day, key=lambda t: t['value'][0]*60+t['value'][1])
                            latest_time = max(times_in_day, key=lambda t: t['value'][0]*60+t['value'][1])
                            ov_time = overflow_from_previous[next_day]['max_time']
                            earliest_min = earliest_time['value'][0] * 60 + earliest_time['value'][1]
                            ov_min = ov_time[0] * 60 + ov_time[1]
                            estimated_overnight_dur = (1440 - earliest_min) + ov_min
                            
                            # Solo atribuir overnight si:
                            # 1. Duración <= 16h (razonable para duty), O
                            # 2. El día solo tiene 1-2 tiempos (vuelo largo único) Y último >= 19:00, O
                            # 3. El último tiempo del día es >= 20:00 (actividad hasta tarde)
                            latest_hour = latest_time['value'][0]
                            n_times = len(times_in_day)
                            is_reasonable = (
                                estimated_overnight_dur <= 960 or  # <= 16h
                                (n_times <= 2 and latest_hour >= 19) or  # Vuelo largo único
                                latest_hour >= 20  # Actividad hasta tarde
                            )
                            if is_reasonable:
                                overnight_end_time = ov_time
                                overnight_airports = overflow_from_previous[next_day].get('airports', [])
                    
                    # ═══════════════════════════════════════════════════════
                    # 1) VERIFICACIÓN: Códigos de tierra
                    # Si hay código tierra Y NO hay aeropuertos (vuelos) → skip
                    # Si hay código tierra PERO también hay aeropuertos → procesar
                    #   solo la parte de vuelo (ajustar tiempos)
                    # ═══════════════════════════════════════════════════════
                    has_tierra = any(c in codigos_tierra for c in codes_in_day)
                    # También verificar si algún código contiene un código tierra como substring
                    # IMPORTANTE: Solo buscar en tokens tipo 'code', NO en airports.
                    # El check anterior ya usa codes_in_day que incluye airports por nombre
                    # (ej: 'FUE'), y un substring match como 'F' in 'FUE' da falso positivo.
                    if not has_tierra:
                        # Solo buscar substrings en tokens que NO sean aeropuertos puros
                        non_airport_codes = [t['value'] for t in tokens 
                                             if t['type'] == 'code']
                        has_tierra = any(
                            ct in code_val 
                            for code_val in non_airport_codes 
                            for ct in codigos_tierra
                        )
                    
                    # Comprobar si hay vuelos reales en el día
                    # Vuelos reales = aeropuertos CON tiempos de vuelo (no solo aeropuertos de líneas resumen)
                    flight_airports_in_day = [t for t in tokens if t['type'] == 'airport']
                    flight_times_count = len(times_in_day)
                    
                    if has_tierra and (not flight_airports_in_day or flight_times_count < 2):
                        continue  # Solo tierra, sin vuelos reales → skip completo
                    
                    # Si TODOS los códigos tierra son del schedule (order==-1), verificar que hay
                    # vuelos reales en las líneas de detalle. Si solo hay aeropuertos aislados
                    # (de líneas resumen) sin estructura de vuelo real, skip el día.
                    if has_tierra:
                        tierra_from_schedule_only = all(
                            t['order'] == -1 
                            for t in tokens 
                            if t['type'] == 'code' and t['value'] in codigos_tierra
                        )
                        if tierra_from_schedule_only:
                            # Contar aeropuertos distintos en el día
                            distinct_airports = set(t['value'] for t in tokens if t['type'] == 'airport')
                            # Un día de vuelo real tiene al menos 2 aeropuertos DISTINTOS
                            # (origen + destino). Si solo hay 1 (ej: BCN de resumen), skip.
                            if len(distinct_airports) <= 1:
                                continue  # No hay ruta real (solo base de resumen) → skip
                    
                    # Si hay tierra + vuelos: ajustar para usar solo tiempos de vuelo
                    # Los códigos tierra del schedule line (order == -1) no afectan
                    tierra_in_detail = False
                    if has_tierra and flight_airports_in_day and flight_times_count >= 2:
                        tierra_tokens_detail = [t for t in tokens 
                                                if t['type'] == 'code' 
                                                and t['value'] in codigos_tierra
                                                and t['order'] != -1]
                        if tierra_tokens_detail:
                            tierra_in_detail = True
                            # Determinar si tierra está DESPUÉS de todos los vuelos
                            max_airport_order = max(t['order'] for t in flight_airports_in_day)
                            tierra_after_flights = all(
                                t['order'] > max_airport_order for t in tierra_tokens_detail
                            )
                            if tierra_after_flights:
                                # Tierra después de vuelos: filtrar tiempos post-tierra
                                min_tierra_order = min(t['order'] for t in tierra_tokens_detail)
                                tokens = [t for t in tokens if t['order'] < min_tierra_order or t['type'] == 'airport']
                                times_in_day = [t for t in tokens if t['type'] == 'time']
                            else:
                                # Tierra antes/durante vuelos: filtrar tiempos pre-vuelo
                                min_airport_order = min(t['order'] for t in flight_airports_in_day)
                                tokens = [t for t in tokens if t['order'] >= min_airport_order - 5 or t['type'] == 'airport']
                                times_in_day = [t for t in tokens if t['type'] == 'time']
                    
                    # ═══════════════════════════════════════════════════════
                    # 2) DETECCIÓN DE IMAGINARIA
                    # ═══════════════════════════════════════════════════════
                    has_imaginaria = False
                    imaginaria_idx = -1
                    for ti, t in enumerate(tokens):
                        if t['type'] == 'code' and es_imaginaria(t['value']):
                            has_imaginaria = True
                            imaginaria_idx = ti
                            break
                    
                    # VALIDACIÓN: Debe haber al menos 1 aeropuerto para ser vuelo
                    all_airports_day = [t for t in tokens if t['type'] == 'airport']
                    if not all_airports_day:
                        continue  # Sin aeropuertos → no hay vuelos → skip
                    
                    # Detectar códigos de transporte (TAXI, TREN, etc.) en el día
                    has_transport = any(t['type'] == 'transport' for t in tokens)
                    transport_tokens = [t for t in tokens if t['type'] == 'transport']
                    
                    if has_imaginaria:
                        # Guardia + vuelos: calcular duty SOLO de la parte de vuelo
                        # (excluir tiempo de espera en standby)
                        # Buscar primer aeropuerto DESPUÉS de la imaginaria
                        aeropuerto_order = -1
                        for ti in range(imaginaria_idx + 1, len(tokens)):
                            if tokens[ti]['type'] == 'airport':
                                aeropuerto_order = tokens[ti]['order']
                                break
                        
                        if aeropuerto_order == -1:
                            continue  # Sin aeropuertos tras imaginaria → skip
                        
                        # Tiempos de VUELO: todos los que vienen DESPUÉS del aeropuerto
                        # Usa min/max CRONOLÓGICO (el PDF no respeta orden temporal)
                        flight_times = [t for t in times_in_day
                                        if t['order'] > aeropuerto_order]
                        if len(flight_times) < 2:
                            continue
                        
                        ft_vals = [t['value'] for t in flight_times]
                        min_ft = min(ft_vals, key=lambda v: v[0]*60+v[1])
                        max_ft = max(ft_vals, key=lambda v: v[0]*60+v[1])
                        first_h, first_m = min_ft
                        last_h, last_m = max_ft
                    else:
                        # Sin imaginaria: primera y última hora del día
                        # Filtrar tiempos "huérfanos" (order >> último aeropuerto)
                        airport_orders = [t['order'] for t in tokens if t['type'] == 'airport']
                        valid_times = times_in_day
                        if airport_orders and len(times_in_day) >= 3:
                            max_airport_order = max(airport_orders)
                            near_times = [t for t in times_in_day
                                          if t['order'] - max_airport_order <= 40]
                            if len(near_times) >= 2:
                                valid_times = near_times
                        
                        if len(valid_times) < 2:
                            # ═══════════════════════════════════════════════
                            # CASO ESPECIAL: Vuelo nocturno con 1 solo tiempo
                            # En vuelos de larga distancia (BCN→DSS, BCN→BJL),
                            # a veces solo se captura la hora de salida en la
                            # columna del día. La hora de llegada está en el
                            # día siguiente (overflow). Si hay overnight_end_time,
                            # usamos el único tiempo como inicio del duty.
                            # ═══════════════════════════════════════════════
                            if len(valid_times) == 1 and overnight_end_time is not None:
                                first_h, first_m = valid_times[0]['value']
                                last_h, last_m = first_h, first_m  # Same as start (only 1 time)
                            else:
                                continue
                        else:
                            # Usar min/max CRONOLÓGICO (el PDF no siempre respeta
                            # el orden temporal en las columnas de datos)
                            tv = [t['value'] for t in valid_times]
                            min_tv = min(tv, key=lambda v: v[0]*60+v[1])
                            max_tv = max(tv, key=lambda v: v[0]*60+v[1])
                            first_h, first_m = min_tv
                            last_h, last_m = max_tv
                    
                    # ═══════════════════════════════════════════════════════
                    # 3) CÁLCULO DE DURACIÓN
                    # ═══════════════════════════════════════════════════════
                    is_overnight = False  # Flag para vuelo nocturno
                    
                    if overnight_end_time is not None:
                        # ═══════════════════════════════════════════════════
                        # VUELO NOCTURNO: El duty cruza medianoche.
                        # first_h/first_m = inicio del duty (ej: 14:50)
                        # overnight_end_time = llegada post-medianoche (ej: 2:05)
                        # Duración = (24:00 - inicio) + llegada_madrugada
                        # ═══════════════════════════════════════════════════
                        first_minutes = first_h * 60 + first_m
                        overnight_h, overnight_m = overnight_end_time
                        overnight_minutes = overnight_h * 60 + overnight_m
                        duration_minutes = (1440 - first_minutes) + overnight_minutes
                        duration_hours = duration_minutes / 60
                        is_overnight = True
                        # Usar last_h/last_m del mismo día para comparaciones internas
                        last_minutes = last_h * 60 + last_m
                    else:
                        first_minutes = first_h * 60 + first_m
                        last_minutes = last_h * 60 + last_m
                        span = last_minutes - first_minutes
                        
                        # Si span > 16h → cruce de medianoche invertido
                        # (min_time post-medianoche, max_time pre-medianoche)
                        if span > 960:  # > 16h
                            duration_minutes = (first_minutes + 1440) - last_minutes
                        else:
                            duration_minutes = span
                        
                        duration_hours = duration_minutes / 60
                    
                    # ═══════════════════════════════════════════════════════
                    # VALIDACIÓN FIX 5: Excluir tiempos L1 en casos específicos
                    #
                    # La primera línea de datos (L1) contiene la hora de
                    # presentación. En la MAYORÍA de casos, esta hora ES el
                    # inicio legítimo del duty (salida real desde base).
                    #
                    # Solo se excluye en DOS casos concretos:
                    #
                    # Método A – Deadhead a base:
                    #   Si el primer tramo de la ruta va de un aeropuerto
                    #   externo HACIA la base del tripulante (ej: AMS→BCN),
                    #   el tiempo L1 corresponde al deadhead, no al duty
                    #   operativo. Se recalcula sin él.
                    #   Ejemplo: KAIL 12:20 [AMS→BCN deadhead] → 14:35-22:35 = 8h
                    #
                    # Método B – Gap enorme + ruta simple:
                    #   Si hay un gap > 4h entre el 1er y 2do tiempo cronológico
                    #   Y la ruta solo tiene ≤ 2 aeropuertos distintos,
                    #   el 1er tiempo es anómalo (misalineación de columna o
                    #   report time sin vuelo operativo asociado).
                    #   Ejemplo: IZAGUIRRE 6:10 → [gap 7.8h] → 14:00-17:35 = 3.58h
                    # ═══════════════════════════════════════════════════════
                    if duration_hours > UMBRAL_LONG_DUTY and not has_imaginaria and not is_overnight:
                        # Obtener aeropuertos ordenados por aparición
                        airports_ordered = [t for t in tokens if t['type'] == 'airport' and t.get('order', -1) >= 0]
                        distinct_airports_fix5 = set(t['value'] for t in all_airports_day)
                        
                        active_times_fix5 = valid_times if not has_imaginaria else times_in_day
                        chrono_fix5 = sorted(active_times_fix5, key=lambda t: t['value'][0]*60 + t['value'][1])
                        
                        fix5_applied = False
                        
                        # Método A: Deadhead a base
                        # Condición: primer_aeropuerto ≠ base Y segundo_aeropuerto = base
                        # SAFEGUARD: Solo aplicar si el primer aeropuerto aparece UNA sola vez.
                        # Si aparece múltiples veces, es la base operacional del tripulante
                        # (ej: BIO para tripulantes desplazados con base oficial BCN),
                        # y el primer tramo NO es un deadhead sino un vuelo real.
                        if (not fix5_applied and base_tripulante and 
                            len(airports_ordered) >= 2):
                            first_ap = airports_ordered[0]['value']
                            second_ap = airports_ordered[1]['value']
                            # Contar cuántas veces aparece el primer aeropuerto
                            first_ap_count = sum(1 for t in airports_ordered if t['value'] == first_ap)
                            if (first_ap != base_tripulante and second_ap == base_tripulante
                                and first_ap_count == 1):
                                # El primer tramo es un deadhead a base (aeropuerto origen solo aparece 1 vez).
                                # Recalcular sin el tiempo L1 (order < primer aeropuerto)
                                first_airport_order = airports_ordered[0]['order']
                                if chrono_fix5 and chrono_fix5[0]['order'] < first_airport_order:
                                    post_dh_times = [t for t in chrono_fix5 if t['order'] >= first_airport_order]
                                    if len(post_dh_times) >= 2:
                                        new_min_v = post_dh_times[0]['value']
                                        new_max_v = post_dh_times[-1]['value']
                                        new_span = (new_max_v[0]*60 + new_max_v[1]) - (new_min_v[0]*60 + new_min_v[1])
                                        new_dur = new_span / 60
                                        if 0 <= new_dur <= UMBRAL_LONG_DUTY:
                                            first_h, first_m = new_min_v
                                            last_h, last_m = new_max_v
                                            first_minutes = first_h * 60 + first_m
                                            last_minutes = last_h * 60 + last_m
                                            duration_minutes = last_minutes - first_minutes
                                            duration_hours = duration_minutes / 60
                                            fix5_applied = True
                        
                        # Método B: Gap > 4h al inicio + ruta simple (≤ 2 aeropuertos)
                        if (not fix5_applied and duration_hours > UMBRAL_LONG_DUTY and 
                            len(distinct_airports_fix5) <= 2 and len(chrono_fix5) >= 3):
                            g0 = ((chrono_fix5[1]['value'][0]*60 + chrono_fix5[1]['value'][1]) -
                                  (chrono_fix5[0]['value'][0]*60 + chrono_fix5[0]['value'][1]))
                            if g0 > 240:  # > 4h gap al inicio
                                remaining_b = chrono_fix5[1:]
                                new_min_b = remaining_b[0]['value']
                                new_max_b = remaining_b[-1]['value']
                                new_span_b = (new_max_b[0]*60 + new_max_b[1]) - (new_min_b[0]*60 + new_min_b[1])
                                new_dur_b = new_span_b / 60
                                if 0 <= new_dur_b <= UMBRAL_LONG_DUTY:
                                    first_h, first_m = new_min_b
                                    last_h, last_m = new_max_b
                                    first_minutes = first_h * 60 + first_m
                                    last_minutes = last_h * 60 + last_m
                                    duration_minutes = last_minutes - first_minutes
                                    duration_hours = duration_minutes / 60
                                    fix5_applied = True
                    
                    # ═══════════════════════════════════════════════════════
                    # VALIDACIÓN FIX 6: Rutas con un solo aeropuerto
                    # Si la ruta solo tiene 1 aeropuerto distinto (ej: "BCN"),
                    # no es un vuelo real → no contar como long duty.
                    # Esto filtra actividades de tierra/briefing que aparecen
                    # con tiempos largos pero sin vuelos reales.
                    # ═══════════════════════════════════════════════════════
                    if duration_hours > UMBRAL_LONG_DUTY:
                        distinct_airports_day = set(t['value'] for t in all_airports_day)
                        # Para vuelos nocturnos, incluir aeropuertos del overflow
                        if is_overnight and overnight_airports:
                            distinct_airports_day.update(overnight_airports)
                        if len(distinct_airports_day) <= 1:
                            # Solo 1 aeropuerto → actividad de tierra, no long duty
                            if duration_hours > 0.5:
                                daily_duties.append((day_num, round(duration_hours, 2)))
                            continue  # No añadir a long_duty_days
                    
                    # ═══════════════════════════════════════════════════════
                    # VALIDACIÓN: Días con transporte terrestre (TAXI, TREN, etc.)
                    # Si el día tiene códigos de transporte y el duty calculado
                    # es desproporcionado respecto al número de vuelos, es falso positivo.
                    # Heurística: con transporte, si duty > 2.5 * (nº aeropuertos * 1.5h),
                    # reducir a la estimación de vuelo real.
                    # ═══════════════════════════════════════════════════════
                    if has_transport and duration_hours > UMBRAL_LONG_DUTY:
                        # Calcular estimación de bloque real basada en aeropuertos
                        # Usar ruta deduplicada para contar legs reales
                        ap_list = [t['value'] for t in all_airports_day]
                        dedup_route = [ap_list[0]] if ap_list else []
                        for ap in ap_list[1:]:
                            if ap != dedup_route[-1]:
                                dedup_route.append(ap)
                        n_real_legs = max(1, len(dedup_route) - 1)
                        estimated_block = n_real_legs * 1.75  # ~1.75h por leg promedio
                        
                        # Si duty > 2x el bloque estimado, es inflado por transporte
                        if duration_hours > estimated_block * 2.0 and estimated_block < 10:
                            # Día con transporte terrestre: NO contabilizar como long duty
                            # pero sí como duty normal (limitado)
                            if duration_hours > 0.5:
                                daily_duties.append((day_num, round(duration_hours, 2)))
                            continue  # No añadir a long_duty_days
                    
                    if duration_hours > 0.5:
                        daily_duties.append((day_num, round(duration_hours, 2)))
                    
                    if duration_hours > UMBRAL_LONG_DUTY:
                        # 3) EXTRACCIÓN DE RUTA: orden EXACTO de aeropuertos
                        #    Captura aeropuertos tal como aparecen en el PDF
                        #    SIN añadir BASE artificialmente entre destinos
                        #    Elimina duplicados consecutivos (escala en tierra)
                        #    Ejemplo MAITE: BCN,SVQ,SVQ,LPA,LPA,SVQ,SVQ,BCN
                        #    → BCN-SVQ-LPA-SVQ-BCN (dedup consecutivos)
                        # ═══════════════════════════════════════════════════════
                        try:
                            # Recoger TODOS los aeropuertos del día en orden exacto
                            all_airports = [t['value'] for t in tokens if t['type'] == 'airport']
                            
                            # Para vuelos nocturnos, añadir aeropuertos del overflow (llegada)
                            if is_overnight and overnight_airports:
                                all_airports.extend(overnight_airports)
                            
                            if all_airports:
                                # Eliminar duplicados CONSECUTIVOS solamente
                                ruta_limpia = [all_airports[0]]
                                for ap in all_airports[1:]:
                                    if ap != ruta_limpia[-1]:
                                        ruta_limpia.append(ap)
                                
                                if base_tripulante and len(ruta_limpia) == 0:
                                    ruta_limpia = [base_tripulante]
                                
                                ruta_str = '-'.join(ruta_limpia)
                                # Añadir indicador 🌙 para vuelos nocturnos
                                if is_overnight:
                                    ruta_str += ' 🌙'
                            else:
                                ruta_str = ''
                        except Exception as e:
                            ruta_str = ''
                        
                        if ruta_str:
                            long_duty_days.append((day_num, round(duration_hours, 2), ruta_str))
                
                # Guardar datos: MERGEAR en lugar de reemplazar
                # Si el crew ya existe, combinar long duties de ambas apariciones
                if crew_id not in all_crew_data:
                    all_crew_data[crew_id] = {
                        'daily_duties': daily_duties,
                        'long_duty_days': long_duty_days,
                    }
                else:
                    existing = all_crew_data[crew_id]
                    # Si nueva extracción tiene más días, usar sus daily_duties
                    if len(daily_duties) > len(existing.get('daily_duties', [])):
                        existing['daily_duties'] = daily_duties
                    # SIEMPRE mergear long_duty_days (unión por day_num)
                    existing_ld_days = set(ld[0] for ld in existing.get('long_duty_days', []))
                    for ld in long_duty_days:
                        if ld[0] not in existing_ld_days:
                            existing['long_duty_days'].append(ld)

        # Liberación de memoria por PDF procesado
        try:
            del reader
        except Exception:
            pass
        gc.collect()

    return all_crew_data

@st.cache_data(show_spinner=False, max_entries=8)
def _extract_daily_duty_data_cached(payload):
    files = _payload_to_fileobjs(payload)
    return _extract_daily_duty_data_impl(files, progress_callback=None)


def extract_daily_duty_data(file_list, progress_callback=None):
    """Wrapper cacheado; mantiene compatibilidad de firma original."""
    total_bytes = _estimate_total_size(file_list)
    if total_bytes <= 0:
        return {}

    # Evitar cachear PDFs masivos para no duplicar bytes en RAM.
    if total_bytes > CACHE_BYPASS_BYTES:
        _safe_rewind_files(file_list)
        return _extract_daily_duty_data_impl(file_list, progress_callback=progress_callback)

    payload = _files_to_cache_payload(file_list)
    if not payload:
        return {}

    # En modo cacheado no se usa progreso página a página para evitar romper hash de cache.
    if progress_callback:
        progress_callback(1, 1, "Long duties")
    return _extract_daily_duty_data_cached(payload)

# --- 3b. GENERADOR DE PDF ---
def extract_crew_pages_from_text(content):
    """Extract individual crew member pages from text content"""
    content = content.replace('\r\n', '\n').replace('\r', '\n')
    sections = re.split(r'^[_-]{50,}$', content, flags=re.MULTILINE)
    crew_pages = []
    for section in sections:
        section = section.strip()
        if len(section) > 200 and 'Vueling Airlines' in section:
            crew_pages.append(section)
    return crew_pages

def extract_base_from_crew_page(page_text):
    """Extract the base code (SVQ, PMI, ALC, etc.) from a crew page text.
    
    The base appears at the beginning of the line containing the roster data,
    typically in format: 'BASE  320  JC' or 'BASE  320  TC' followed by roster codes.
    Also handles non-numeric fleet codes like 'TRA' (training).
    
    Returns the base code or None if not found.
    """
    # List of known base codes
    known_bases = ['BCN', 'ALC', 'BIO', 'LPA', 'IBZ', 'AGP', 'PMI', 'SCQ', 'SVQ', 'TFN', 'VLC']
    
    lines = page_text.split('\n')
    for line in lines:
        # Skip header lines
        if 'Vueling Airlines' in line or 'CREW SCHEDULE' in line:
            continue
        if 'NAME' in line and 'ID' in line:
            continue
        if line.strip().startswith('ID '):
            continue
        
        # Look for base code at beginning of line (with possible leading whitespace)
        stripped = line.strip()
        for base in known_bases:
            # Check if line starts with base code followed by space and fleet type (numeric "320" or alpha "TRA")
            if re.match(rf'^{base}\s+\w{{2,4}}\s+(JC|TC)', stripped):
                return base
            # Also check for pattern where base appears after minimal whitespace
            if re.match(rf'^\s*{base}\s+\w{{2,4}}\s+(JC|TC)', line):
                return base
    
    # Fallback: analyze schedule content for base indicators
    # Look for airport codes that appear frequently as departure/arrival points
    base_mentions = {}
    for base in known_bases:
        # Count occurrences of base code in schedule lines (not in header)
        count = 0
        for line in lines:
            if 'Vueling Airlines' in line or 'CREW SCHEDULE' in line:
                continue
            # Count appearances as standalone airport code in schedule data
            matches = re.findall(rf'\b{base}\b', line)
            count += len(matches)
        if count > 3:  # Threshold: >3 mentions suggests this is the crew's base
            base_mentions[base] = count
    
    if base_mentions:
        # Return the base with most mentions
        return max(base_mentions, key=base_mentions.get)
    
    return None

def extract_operational_base_from_crew_page(page_text):
    """Determina la base OPERATIVA real de un tripulante analizando los aeropuertos
    de inicio/fin de sus jornadas en el mes.
    
    Lógica: cuenta cuántos días operativos comienza el tripulante desde cada aeropuerto.
    La base operativa es el aeropuerto desde el que más días sale.
    
    Returns: (base_operativa, dict_dias_por_base) o (None, {})
    """
    known_bases = ['BCN', 'ALC', 'BIO', 'LPA', 'IBZ', 'AGP', 'PMI', 'SCQ', 'SVQ', 'TFN', 'VLC']
    
    lines = page_text.split('\n')
    
    # Encontrar la línea principal del schedule (BASE FLEET CAT + códigos)
    schedule_line_idx = None
    official_base = None
    for idx, line in enumerate(lines):
        stripped = line.strip()
        m = re.match(r'^([A-Z]{3})\s+\w{2,4}\s+(JC|TC)', stripped)
        if m:
            schedule_line_idx = idx
            official_base = m.group(1)
            break
    
    if schedule_line_idx is None:
        return None, {}
    
    # Buscar las líneas con aeropuertos después de la línea del schedule
    # Estructura típica de un crew page:
    # Línea 0: schedule codes (3044 2227 3160 EOFF OFF ...)
    # Línea 1: nombre + tiempos
    # Línea 2: ID + BLOCK > + aeropuertos (primera salida de cada día)
    # Línea 3: DUTY > + aeropuertos (destino)
    # ... más líneas con aeropuertos alternados
    
    # Recoger todos los códigos de aeropuerto de 3 letras que aparecen en las líneas
    # posteriores al schedule, en las posiciones de las columnas de días
    airport_lines = []
    for idx in range(schedule_line_idx + 1, min(schedule_line_idx + 40, len(lines))):
        line = lines[idx]
        stripped = line.strip()
        # Parar si encontramos separador o nuevo tripulante
        if re.match(r'^[_-]{30,}$', stripped):
            break
        if re.match(r'^([A-Z]{3})\s+\w{2,4}\s+(JC|TC)', stripped):
            break
        # Buscar líneas que contengan aeropuertos conocidos
        airports_in_line = re.findall(r'\b(' + '|'.join(known_bases) + r')\b', line)
        if airports_in_line:
            airport_lines.append(airports_in_line)
    
    if not airport_lines:
        return None, {}
    
    # La PRIMERA línea con aeropuertos (generalmente después de BLOCK >) 
    # muestra el aeropuerto de SALIDA de cada día operativo.
    # Esa es la referencia para la base operativa.
    dias_por_base = {}
    
    # Analizar TODAS las líneas con aeropuertos, dando prioridad a la primera
    # (que suele ser la base de salida)
    # La primera aparición de un aeropuerto por cada "columna/día" es la base de salida
    first_airports = airport_lines[0] if airport_lines else []
    
    for ap in first_airports:
        if ap in known_bases:
            dias_por_base[ap] = dias_por_base.get(ap, 0) + 1
    
    # También contar en líneas pares (que suelen ser retornos a base)
    for i in range(2, len(airport_lines), 2):
        for ap in airport_lines[i]:
            if ap in known_bases:
                dias_por_base[ap] = dias_por_base.get(ap, 0) + 1
    
    if dias_por_base:
        base_operativa = max(dias_por_base, key=dias_por_base.get)
        return base_operativa, dias_por_base
    
    return None, {}


def extract_crew_pages_with_base(content):
    """Extract crew pages and their associated base codes from text content.
    
    MEJORA v3.0: Detecta tanto la base oficial (de la línea del schedule) como
    la base operativa real (analizando aeropuertos de salida). Si difieren,
    marca al tripulante como desplazado.
    
    Returns list of tuples: [(page_text, base_oficial, base_operativa, es_desplazado), ...]
    """
    pages = extract_crew_pages_from_text(content)
    pages_with_base = []
    for page in pages:
        base_oficial = extract_base_from_crew_page(page)
        base_operativa, dias_por_base = extract_operational_base_from_crew_page(page)
        
        # Si no se pudo determinar base operativa, usar la oficial
        if not base_operativa:
            base_operativa = base_oficial
        
        es_desplazado = (base_oficial != base_operativa) if (base_oficial and base_operativa) else False
        
        pages_with_base.append((page, base_oficial, base_operativa, es_desplazado))
    return pages_with_base

def generate_roster_pdf_bytes(pages):
    """Generate PDF bytes from crew roster pages matching reference layout"""
    page_width, page_height = landscape(A4)
    buffer = io.BytesIO()
    c = rl_canvas.Canvas(buffer, pagesize=landscape(A4))
    font_name = "Courier"
    available_width = page_width - 40
    chars_per_line = 225
    font_size = round((available_width / chars_per_line) / 0.6, 1)
    line_height = font_size + 1.5
    
    for page_text in pages:
        lines = page_text.split('\n')
        y = page_height - 20
        c.setFont(font_name, font_size)
        for line in lines:
            if y < 20:
                break
            c.drawString(20, y, line)
            y -= line_height
        c.showPage()
    
    c.save()
    buffer.seek(0)
    return buffer.getvalue(), len(pages)

# --- 4. MOTOR DE EXTRACCIÓN ---
def _extraer_dias_desde_cabecera(full_text):
    """Extrae el número de días del mes desde la cabecera del PDF.
    Busca 'CREW SCHEDULE FROM :DD/MM/YYYY TO DD/MM/YYYY' y calcula los días.
    Funciona correctamente para todos los meses: 28, 29 (bisiesto), 30 y 31 días."""
    header_match = re.search(
        r'CREW\s+SCHEDULE\s+FROM\s*:\s*(\d{2})/(\d{2})/(\d{4})\s+TO\s+(\d{2})/(\d{2})/(\d{4})',
        full_text
    )
    if header_match:
        dia_from = int(header_match.group(1))
        mes_from = int(header_match.group(2))
        anio_from = int(header_match.group(3))
        dia_to = int(header_match.group(4))
        mes_to = int(header_match.group(5))
        anio_to = int(header_match.group(6))
        
        # Calcular días reales del mes usando calendar (maneja bisiestos correctamente)
        import calendar
        dias_en_mes = calendar.monthrange(anio_from, mes_from)[1]
        
        # Verificar consistencia: si FROM es día 1 y TO es último día del mes
        if dia_from == 1 and dia_to == dias_en_mes and mes_from == mes_to:
            return dias_en_mes, mes_from, anio_from
        
        # Si el rango es del mismo mes, usar dia_to como cantidad de días
        if mes_from == mes_to and anio_from == anio_to:
            return dia_to, mes_from, anio_from
        
        # Fallback: usar los días del mes FROM
        return dias_en_mes, mes_from, anio_from
    return None, None, None


def split_concatenated_tokens(tokens):
    """Split tokens where flight numbers and activity codes got concatenated
    due to PDF text extraction (no space between columns).
    
    Handles:
    - '3976SBY' -> ['3976', 'SBY']   (flight + pure alpha code 2+ letters)
    - '3911OFF' -> ['3911', 'OFF']   (flight + OFF)
    - '3149ROFF' -> ['3149', 'ROFF'] (flight + ROFF)
    - '3942F2' -> ['3942', 'F2']     (flight + alphanumeric code like F2)
    - '3935F' -> ['3935', 'F']       (flight + single letter code like F)
    - 'SIND3940' -> ['SIND', '3940'] (alpha code + flight)
    
    Tokens that are pure codes (like 'ROFF', 'RVAC') or pure numbers ('3905') stay unchanged.
    """
    result = []
    for token in tokens:
        # Pattern 1: flight_number (4 digits) followed by activity_code (2+ uppercase letters)
        m = re.match(r'^(\d{3,4})([A-Z]{2,})$', token)
        if m:
            result.append(m.group(1))
            result.append(m.group(2))
            continue
        # Pattern 2: flight_number (4 digits) followed by short alphanumeric code (1-3 chars with letter+digit)
        # Handles cases like '3942F2' where F2 is a schedule code
        m = re.match(r'^(\d{4})([A-Z]\d{1,2})$', token)
        if m:
            result.append(m.group(1))
            result.append(m.group(2))
            continue
        # Pattern 3: flight_number (4 digits) followed by single uppercase letter
        # Handles cases like '3935F', '3648F', '3950F' where F is a standalone schedule code
        # CRITICAL FIX: Without this, tokens like '3935F' remain unsplit, causing token count != dias_mes,
        # which triggers column-based extraction that misaligns with fleet/cat prefixes (320, JC/TC)
        # and produces completely wrong schedules (shifted by 2-4 days).
        m = re.match(r'^(\d{4})([A-Z])$', token)
        if m:
            result.append(m.group(1))
            result.append(m.group(2))
            continue
        # Pattern 4: activity_code followed by flight_number (rare but handle)
        m = re.match(r'^([A-Z]{2,})(\d{3,4})$', token)
        if m:
            result.append(m.group(1))
            result.append(m.group(2))
            continue
        result.append(token)
    return result


def _extract_sectors_by_day_pdfplumber_impl(file_list, progress_callback=None):
    """Extrae TODOS los sectores (números de vuelo) por día para cada tripulante usando pdfplumber.
    
    Devuelve un dict: {crew_id: {day_num: set_of_sector_numbers}}
    
    Esto complementa extract_roster_stream() para detectar cambios de sectores
    cuando el código principal es el mismo pero los vuelos cambian.
    """
    all_sectors = {}  # {crew_id: {day: set(sectors)}}

    re_base_fleet = re.compile(r'^([A-Z]{3})\s+(\S+)\s+(JC|TC)')
    processed_pages = 0
    total_pages_global = 0

    for _fo in file_list:
        if _fo is None:
            continue
        try:
            if hasattr(_fo, 'seek'):
                _fo.seek(0)
            with pdfplumber.open(_fo) as _pdf_count:
                total_pages_global += len(_pdf_count.pages)
        except Exception:
            continue
        finally:
            try:
                if hasattr(_fo, 'seek'):
                    _fo.seek(0)
            except Exception:
                pass

    total_pages_global = max(total_pages_global, 1)

    for file_obj in file_list:
        if file_obj is None:
            continue
        
        # pdfplumber needs a file path or file-like object
        # Reset position if file-like
        if hasattr(file_obj, 'seek'):
            file_obj.seek(0)
        
        try:
            pdf = pdfplumber.open(file_obj)
        except Exception:
            continue
        
        try:
            for page in pdf.pages:
                processed_pages += 1
                if progress_callback and (
                    processed_pages == 1
                    or processed_pages == total_pages_global
                    or processed_pages % 5 == 0
                ):
                    progress_callback(processed_pages, total_pages_global, "Extrayendo sectores")

                words = page.extract_words()
                if not words:
                    continue
                
                # Find all crew header lines (TC/JC words)
                tc_jc_words = sorted(
                    [w for w in words if w['text'] in ('TC', 'JC')],
                    key=lambda x: x['top']
                )
                
                for tc_idx, tc_word in enumerate(tc_jc_words):
                    header_y = tc_word['top']
                    
                    # Verify this is actually a crew header by checking for base code before it
                    header_words_line = sorted(
                        [w for w in words if abs(w['top'] - header_y) < 3],
                        key=lambda x: x['x0']
                    )
                    
                    # Check if line starts with a base pattern
                    if len(header_words_line) < 3:
                        continue
                    line_text = ' '.join(w['text'] for w in header_words_line[:3])
                    if not re_base_fleet.match(line_text):
                        continue
                    
                    # Get schedule words (after TC/JC)
                    schedule_words = sorted(
                        [w for w in header_words_line if w['x0'] > tc_word['x0'] + 10],
                        key=lambda x: x['x0']
                    )
                    
                    if len(schedule_words) < 20:  # Too few for a full month
                        continue
                    
                    # Build day column positions from schedule words
                    day_cols = {i+1: w['x0'] for i, w in enumerate(schedule_words)}
                    
                    # Find crew ID: search words below header for a 4-5 digit number
                    crew_id = None
                    for w in words:
                        if (w['top'] > header_y + 3 and w['top'] < header_y + 30
                            and re.match(r'^\d{4,5}$', w['text'])):
                            crew_id = int(w['text'])
                            break
                    
                    if not crew_id:
                        continue
                    
                    # Find section boundaries
                    # Next crew header or BLOCK summary line
                    if tc_idx + 1 < len(tc_jc_words):
                        next_header_y = tc_jc_words[tc_idx + 1]['top']
                        section_end = next_header_y - 5
                    else:
                        section_end = page.height
                    
                    # Look for BLOCK summary line (last BLOCK in section)
                    block_words = [w for w in words 
                                  if w['text'] == 'BLOCK' 
                                  and w['top'] > header_y + 40 
                                  and w['top'] < section_end]
                    if block_words:
                        last_block = sorted(block_words, key=lambda x: x['top'])[-1]
                        section_end = min(section_end, last_block['top'] - 2)
                    
                    # Extract all 4-digit flight numbers in section
                    flight_words = [w for w in words 
                                  if w['top'] > header_y + 2 
                                  and w['top'] < section_end
                                  and re.match(r'^\d{4}$', w['text'])]
                    
                    # Assign to day columns
                    sectors = {d: set() for d in range(1, len(day_cols) + 1)}
                    for fw in flight_words:
                        best_day = min(day_cols, key=lambda d: abs(day_cols[d] - fw['x0']))
                        dist = abs(day_cols[best_day] - fw['x0'])
                        if dist < 12:  # tolerance in pixels
                            sectors[best_day].add(fw['text'])
                    
                    all_sectors[crew_id] = sectors

                if processed_pages % 80 == 0:
                    gc.collect()
        finally:
            pdf.close()
            del pdf
            gc.collect()
    
    return all_sectors

@st.cache_data(show_spinner=False, max_entries=8)
def _extract_sectors_by_day_cached(payload):
    files = _payload_to_fileobjs(payload)
    return _extract_sectors_by_day_pdfplumber_impl(files, progress_callback=None)


def extract_sectors_by_day_pdfplumber(file_list, progress_callback=None):
    total_bytes = _estimate_total_size(file_list)
    if total_bytes <= 0:
        return {}

    if total_bytes > CACHE_BYPASS_BYTES:
        _safe_rewind_files(file_list)
        return _extract_sectors_by_day_pdfplumber_impl(file_list, progress_callback=progress_callback)

    payload = _files_to_cache_payload(file_list)
    if not payload:
        return {}
    if progress_callback:
        progress_callback(1, 1, "Sectores")
    return _extract_sectors_by_day_cached(payload)


def count_real_changes(i_sched, f_sched, dias_mes, i_sectors=None, f_sectors=None):
    """Cuenta los cambios REALES entre dos programaciones.
    
    Un cambio real es:
    1. Código principal diferente (OFF→FLT, SBY→IMAG, etc.)
    2. Mismo código principal pero sectores/vuelos diferentes
    
    NO es cambio:
    - Mismo código y mismos sectores (aunque cambien horas)
    - Diferencias de formato
    
    Args:
        i_sched: lista de códigos principales del schedule inicial
        f_sched: lista de códigos principales del schedule final
        dias_mes: número de días del mes
        i_sectors: dict {day_num: set(sectors)} del inicial (opcional)
        f_sectors: dict {day_num: set(sectors)} del final (opcional)
    
    Returns:
        int: número de cambios reales
    """
    mods = 0
    for idx in range(dias_mes):
        i_code = i_sched[idx] if idx < len(i_sched) else '--'
        f_code = f_sched[idx] if idx < len(f_sched) else '--'
        
        if i_code != f_code:
            mods += 1
            continue
        
        # Same main code - check if it's a flight day (4-digit code)
        if re.match(r'^\d{4}$', i_code) and i_sectors and f_sectors:
            day_num = idx + 1
            i_sec = i_sectors.get(day_num, set())
            f_sec = f_sectors.get(day_num, set())
            if i_sec and f_sec and i_sec != f_sec:
                mods += 1
    
    return mods


def _extract_roster_stream_impl(file_list, base_seleccionada, progress_callback=None, progress_label="Extrayendo roster"):
    """Extrae tripulantes de los PDFs de programación.
    
    CAMBIO CRÍTICO v2.0: Extrae TODOS los tripulantes del PDF sin filtrar por base.
    Esto es necesario porque los PDFs de BCN contienen tripulantes de múltiples bases
    (AGP, BIO, SCQ, etc.) que operan desde BCN. Filtrar por base causaba pérdida
    de ~10% de tripulantes. El filtrado por base se realiza opcionalmente después.
    
    Sistema de validación:
    - Cuenta líneas de cabecera de tripulante (crew header lines)
    - Cuenta líneas BLOCK (cada tripulante tiene exactamente una)
    - Compara ambos conteos para detectar discrepancias
    - Registra tripulantes sin ID y tripulantes con schedules anómalos
    """
    rosters = {}
    _header_dias = None
    _header_mes = None
    _header_anio = None
    
    # === CONTADORES DE VALIDACIÓN ===
    validation = {
        'crew_header_lines': 0,        # Total líneas que matchean BASE FLEET CAT
        'crew_with_id': 0,             # Tripulantes con ID encontrado
        'crew_without_id': 0,          # Tripulantes sin ID (PÉRDIDA)
        'crew_overwritten': 0,         # IDs duplicados (sobreescritos)
        'crew_schedule_empty': 0,      # Tripulantes sin schedule (TRA fleet, etc.)
        'block_lines_found': 0,        # Líneas BLOCK > encontradas
        'bases_encontradas': {},       # Desglose por base: {base: count}
        'categorias_encontradas': {},  # Desglose por categoría: {cat: count}
        'fleet_codes': {},             # Desglose por flota: {fleet: count}
        'ids_sin_schedule': [],        # IDs de tripulantes sin schedule (para debug)
        'warnings': [],                # Advertencias durante extracción
    }
    
    # Pre-compilar regex para rendimiento
    # CAMBIO: Capturar TODAS las bases, no solo la seleccionada
    re_any_base = re.compile(r'^([A-Z]{3})\s+(\S+)\s+(JC|TC)')
    re_block = re.compile(r'(\d{2,6})\s+BLOCK\s*>\s*(\d{1,3}):(\d{2})')
    re_block_alt = re.compile(r'BLOCK\s*>\s*(\d{1,3}):(\d{2})')
    re_duty = re.compile(r'DUTY\s*>\s*(\d{1,3}):(\d{2})')
    re_id = re.compile(r'^\s*(\d{2,6})(\s|$)')
    re_block_ci = re.compile(r'[Bb][Ll][Oo][Cc][Kk]\s*>\s*(\d{1,3}):(\d{2})')
    re_name_parts = re.compile(r'[A-ZÁÉÍÓÚÀÈÌÒÙÂÊÎÔÛÄËÏÖÜÑÇŽŠŒ]{2,}')
    
    total_pages_global = 0
    for _fo in file_list:
        if _fo is None:
            continue
        try:
            _fo.seek(0)
            total_pages_global += len(PyPDF2.PdfReader(_fo).pages)
        except Exception:
            pass
        finally:
            try:
                _fo.seek(0)
            except Exception:
                pass

    processed_pages_global = 0

    for file_obj in file_list:
        if file_obj is None:
            continue

        try:
            file_obj.seek(0)
        except Exception:
            pass

        reader = PyPDF2.PdfReader(file_obj)
        total_pages = len(reader.pages)

        # Procesar página por página
        carry_over_lines = []
        # Regex para detectar posiciones de columnas de días en el header (feb01, mar01, etc.)
        _re_day_col = re.compile(r'(?:ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)\d{2}', re.IGNORECASE)
        # Posiciones de columna de cada día (se detectan del header de cada página)
        _day_col_positions = None  # [(day_num, start_pos), ...]
        
        for page_idx in range(total_pages):
            processed_pages_global += 1
            if progress_callback and (processed_pages_global % 15 == 0 or processed_pages_global == total_pages_global):
                progress_callback(processed_pages_global, max(total_pages_global, 1), progress_label)
            if processed_pages_global % 120 == 0:
                gc.collect()

            text = reader.pages[page_idx].extract_text()
            if not text:
                carry_over_lines = []
                continue
            
            # Extraer cabecera solo de la primera página con texto
            if not _header_dias:
                h_dias, h_mes, h_anio = _extraer_dias_desde_cabecera(text)
                if h_dias:
                    _header_dias = h_dias
                    _header_mes = h_mes
                    _header_anio = h_anio
            
            page_lines = text.split('\n')
            
            # ── Detectar posiciones de columnas desde la línea del header (ID feb01 feb02...) ──
            page_day_positions = None
            for pline in page_lines[:6]:  # El header está en las primeras líneas
                day_matches = list(_re_day_col.finditer(pline))
                if len(day_matches) >= 20:  # Al menos 20 días para ser un header válido
                    page_day_positions = []
                    for dm in day_matches:
                        day_str = dm.group()
                        day_num = int(day_str[-2:])
                        page_day_positions.append((day_num, dm.start()))
                    # Guardar como referencia global si es la primera vez
                    if _day_col_positions is None:
                        _day_col_positions = page_day_positions
                    else:
                        page_day_positions = _day_col_positions  # Usar global si ya detectamos
                    break
            
            if page_day_positions is None and _day_col_positions is not None:
                page_day_positions = _day_col_positions
            
            # Combinar carry-over para contexto en límites de página
            all_lines = carry_over_lines + page_lines
            offset = len(carry_over_lines)
            
            for i in range(offset, len(all_lines)):
                stripped_line = all_lines[i].strip()
                
                # CAMBIO: Matchear CUALQUIER base, no solo la seleccionada
                base_pattern = re_any_base.match(stripped_line)
                
                if not base_pattern:
                    continue
                
                crew_base = base_pattern.group(1)
                fleet_code = base_pattern.group(2)
                cat = base_pattern.group(3)
                
                validation['crew_header_lines'] += 1
                validation['bases_encontradas'][crew_base] = validation['bases_encontradas'].get(crew_base, 0) + 1
                validation['categorias_encontradas'][cat] = validation['categorias_encontradas'].get(cat, 0) + 1
                validation['fleet_codes'][fleet_code] = validation['fleet_codes'].get(fleet_code, 0) + 1
                
                # ── Extracción de tokens: preferir columnas si están disponibles ──
                raw_codes = stripped_line.split(cat, 1)[-1].strip()
                tokens_split = raw_codes.split()
                tokens_split = split_concatenated_tokens(tokens_split)
                
                # Usar la línea ORIGINAL (no stripped) para extracción por columnas
                original_line = all_lines[i]
                
                # Intentar extracción por columnas si hay posiciones detectadas
                # y los tokens por split no coinciden con el número de días esperado
                expected_days = _header_dias or 28
                tokens = tokens_split  # Default: split-based
                
                if page_day_positions and len(tokens_split) != expected_days:
                    # Extracción por columnas: más precisa para líneas con blancos o horarios
                    col_tokens = []
                    for idx_col, (day_num, start_pos) in enumerate(page_day_positions):
                        if idx_col + 1 < len(page_day_positions):
                            end_pos = page_day_positions[idx_col + 1][1]
                        else:
                            end_pos = len(original_line)
                        
                        if start_pos < len(original_line):
                            cell = original_line[start_pos:min(end_pos, len(original_line))].strip()
                        else:
                            cell = ''
                        col_tokens.append(cell)
                    
                    # Validar: la extracción por columnas debe tener exactamente expected_days tokens
                    if len(col_tokens) == expected_days:
                        # SAFETY CHECK: Detect if column extraction picked up fleet/category
                        # tokens (e.g., '320', 'JC', 'TC') at the start, which indicates the
                        # day column positions don't align with the data line (different PDF format).
                        # In that case, reject column-based extraction and keep split-based.
                        _fleet_cat_contaminated = False
                        if len(col_tokens) >= 3:
                            first_vals = [col_tokens[0].strip(), col_tokens[1].strip()]
                            if any(v in ('JC', 'TC', '320', '321', 'TRA') for v in first_vals):
                                _fleet_cat_contaminated = True
                        
                        if not _fleet_cat_contaminated:
                            # Filtrar tokens de horario HH:MM que no son códigos reales
                            _re_time_only = REGEX_TIME_ONLY
                            cleaned_col = []
                            for ct in col_tokens:
                                if _re_time_only.match(ct):
                                    cleaned_col.append('')  # Horario → tratar como blanco
                                else:
                                    # Aplicar split de tokens concatenados a cada celda
                                    parts = split_concatenated_tokens(ct.split())
                                    cleaned_col.append(parts[0] if parts else '')
                            tokens = cleaned_col
                
                if len(tokens) == 0:
                    validation['crew_schedule_empty'] += 1
                
                name, crew_id = "DESCONOCIDO", None
                block_hours, duty_hours = None, None
                
                context_lines = []
                for j in range(1, 12):
                    if i + j < len(all_lines):
                        context_lines.append(all_lines[i + j].strip())
                
                context_text = "\n".join(context_lines)
                
                # Buscar ID y BLOCK hours
                # IMPORTANTE: Limitar búsqueda a las primeras líneas de contexto
                # para no confundir datos del siguiente tripulante.
                # Formato PDF: Header -> Nombre -> ID -> BLOCK > HH:MM -> DUTY > HH:MM
                block_found_at = -1
                for j, ctx_line in enumerate(context_lines):
                    block_match = re_block.search(ctx_line)
                    if block_match:
                        crew_id = int(block_match.group(1))
                        block_hours = f"{block_match.group(2)}:{block_match.group(3)}"
                        validation['block_lines_found'] += 1
                        block_found_at = j
                        if j > 0:
                            name_line = context_lines[j - 1]
                            name_parts = re_name_parts.findall(name_line)
                            if name_parts:
                                name = name_parts[0]
                        break
                    
                    block_alt = re_block_alt.search(ctx_line)
                    if block_alt and not block_hours:
                        block_hours = f"{block_alt.group(1)}:{block_alt.group(2)}"
                        validation['block_lines_found'] += 1
                        block_found_at = j
                        # Buscar ID en la línea justo anterior al BLOCK
                        if j > 0:
                            id_before_block = re_id.match(context_lines[j - 1])
                            if id_before_block:
                                crew_id = int(id_before_block.group(1))
                                # Nombre está 2 líneas antes del BLOCK
                                if j > 1:
                                    name_parts = re_name_parts.findall(context_lines[j - 2])
                                    if name_parts:
                                        name = name_parts[0]
                        break
                
                # Buscar DUTY hours (solo en líneas cercanas al BLOCK, no más allá)
                search_limit = min(len(context_lines), (block_found_at + 4) if block_found_at >= 0 else 6)
                for j in range(search_limit):
                    duty_match = re_duty.search(context_lines[j])
                    if duty_match:
                        duty_hours = f"{duty_match.group(1)}:{duty_match.group(2)}"
                        break
                
                # Fallback: buscar ID sin BLOCK (solo en primeras líneas, antes de siguiente crew)
                if not crew_id:
                    # Limitar a las primeras 4 líneas o hasta donde se encontró BLOCK
                    id_search_limit = min(len(context_lines), max(4, block_found_at + 1) if block_found_at >= 0 else 4)
                    for j in range(id_search_limit):
                        ctx_line = context_lines[j]
                        id_match = re_id.match(ctx_line)
                        if id_match:
                            crew_id = int(id_match.group(1))
                            if j > 0:
                                name_parts = re_name_parts.findall(context_lines[j - 1])
                                if name_parts:
                                    name = name_parts[0]
                            break
                
                # Fallback: buscar BLOCK case-insensitive
                if not block_hours:
                    for j in range(min(len(context_lines), 6)):
                        ctx_line = context_lines[j]
                        m = re_block_ci.search(ctx_line)
                        if m:
                            block_hours = f"{m.group(1)}:{m.group(2)}"
                            break
                
                # === REGISTRAR TRIPULANTE ===
                if crew_id:
                    validation['crew_with_id'] += 1
                    
                    # Detectar sobreescritura de duplicados
                    if crew_id in rosters:
                        validation['crew_overwritten'] += 1
                        # Mantener el registro con más datos (más tokens)
                        existing_tokens = len(rosters[crew_id]['Schedule'])
                        if len(tokens) <= existing_tokens:
                            continue  # Mantener el existente si tiene más datos
                    
                    # NO filtrar por longitud de tokens (eliminado el filtro <= 40)
                    rosters[crew_id] = {
                        'Name': name,
                        'Categoria': cat,
                        'Base': crew_base,
                        'Fleet': fleet_code,
                        'Schedule': tokens,
                        'Block': block_hours or '--',
                        'Duty': duty_hours or '--'
                    }
                else:
                    validation['crew_without_id'] += 1
                    validation['warnings'].append(
                        f"Sin ID en página {page_idx}: {stripped_line[:80]}"
                    )
            
            carry_over_lines = page_lines[-12:] if len(page_lines) >= 12 else page_lines[:]

        # Liberación de memoria por PDF procesado
        try:
            del reader
        except Exception:
            pass
        gc.collect()
    
    # === METADATOS Y VALIDACIÓN ===
    validation['unique_ids_final'] = len([k for k in rosters if k != '__meta__'])
    
    if rosters:
        rosters['__meta__'] = {
            'header_dias': _header_dias,
            'header_mes': _header_mes,
            'header_anio': _header_anio,
            'validation': validation,
            # Backwards compatibility
            'tripulantes_encontrados': validation['crew_header_lines'],
            'tripulantes_otras_bases': 0,  # Ya no aplica: se extraen todas las bases
        }
    
    return rosters

@st.cache_data(show_spinner=False, max_entries=8)
def _extract_roster_stream_cached(payload, base_seleccionada):
    files = _payload_to_fileobjs(payload)
    return _extract_roster_stream_impl(files, base_seleccionada, progress_callback=None, progress_label="Extrayendo roster")


def extract_roster_stream(file_list, base_seleccionada, progress_callback=None, progress_label="Extrayendo roster"):
    total_bytes = _estimate_total_size(file_list)
    if total_bytes <= 0:
        return {}

    if total_bytes > CACHE_BYPASS_BYTES:
        _safe_rewind_files(file_list)
        return _extract_roster_stream_impl(file_list, base_seleccionada, progress_callback=progress_callback, progress_label=progress_label)

    payload = _files_to_cache_payload(file_list)
    if not payload:
        return {}
    if progress_callback:
        progress_callback(1, 1, progress_label)
    return _extract_roster_stream_cached(payload, base_seleccionada)


def detectar_dias_mes(rosters):
    """Detecta el número de días del mes.
    Prioriza la información de la cabecera del PDF sobre la longitud de los schedules."""
    if not rosters:
        return 31
    
    # Primero intentar usar la cabecera del PDF (más fiable)
    meta = rosters.get('__meta__', {})
    if meta.get('header_dias'):
        return meta['header_dias']
    
    # Fallback: usar la moda (valor más frecuente) de las longitudes de schedule
    dias = []
    for k, data in rosters.items():
        if k == '__meta__':
            continue
        dias.append(len(data['Schedule']))
    
    if dias:
        from collections import Counter
        conteo = Counter(dias)
        # Usar el valor más frecuente (moda) en vez del máximo
        moda = conteo.most_common(1)[0][0]
        return moda
    
    return 31


# === SECCIÓN 5: FUNCIONES DE VISUALIZACIÓN ===
# --- 5. INTERFAZ PRINCIPAL ---


# === SECCIÓN 7: INTERFAZ PRINCIPAL ===
# ===== SECCIÓN: RECORTADOR DE PROGRAMACIONES (standalone) =====
if st.session_state['app_section'] == 'recortador':
    # Renderizar sidebar de admin si corresponde
    render_admin_sidebar()
    
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
    * { font-family: 'Inter', sans-serif; }
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}

    /* Botones Secondary - glassmorphism elegante */
    div[data-testid="stButton"] > button[kind="secondary"],
    .stButton > button[kind="secondary"],
    button[kind="secondary"],
    button[data-testid="stBaseButton-secondary"] {
        background: rgba(255,255,255,0.8) !important;
        backdrop-filter: blur(10px) !important;
        color: #0D5F5D !important;
        border: 1.5px solid rgba(17,127,124,0.2) !important;
        font-weight: 600 !important;
        padding: 10px 20px !important;
        border-radius: 12px !important;
        box-shadow: 0 2px 10px rgba(17,127,124,0.08) !important;
        transition: all 0.3s ease !important;
    }
    div[data-testid="stButton"] > button[kind="secondary"]:hover,
    .stButton > button[kind="secondary"]:hover,
    button[kind="secondary"]:hover,
    button[data-testid="stBaseButton-secondary"]:hover {
        transform: translateY(-2px) !important;
        background: rgba(17,127,124,0.08) !important;
        border-color: #117F7C !important;
        box-shadow: 0 6px 20px rgba(17,127,124,0.15) !important;
        color: #117F7C !important;
    }
    </style>
    """, unsafe_allow_html=True)
    
    _usr_col1, _usr_col2, _usr_col3 = st.columns([1, 5, 1])
    with _usr_col1:
        if st.button("← Volver", key="back_recortador"):
            st.session_state['app_section'] = 'seleccion'
            st.rerun()
    with _usr_col2:
        st.markdown(f"👤 Conectado como: **{st.session_state.get('username', '')}**" + 
                    (" 👑" if st.session_state.get('is_admin') else ""))
    with _usr_col3:
        if st.button("🚪 Cerrar Sesión", key="logout_recortador", type="secondary"):
            update_logout_time()
            for key in ['authenticated', 'username', 'is_admin']:
                st.session_state[key] = False if key == 'authenticated' else None
            st.session_state['app_section'] = 'seleccion'
            st.rerun()
    
    st.markdown("""
    <div style="background: linear-gradient(135deg, #0A5C5A 0%, #0D6F6D 15%, #117F7C 35%, #1A9E9B 60%, #44BABC 85%, #61C1C3 100%); 
                padding: 35px 44px; 
                border-radius: 24px; 
                text-align: center;
                margin: 15px 0 30px 0;
                position: relative;
                overflow: hidden;
                box-shadow: 0 16px 50px rgba(17,127,124,0.3), 0 6px 20px rgba(11,132,127,0.15);">
        <div style="position:absolute;top:-50%;left:-20%;width:400px;height:400px;background:radial-gradient(circle,rgba(97,193,195,0.15) 0%,transparent 70%);border-radius:50%;"></div>
        <div style="position:absolute;bottom:0;left:0;right:0;height:1px;background:linear-gradient(90deg,transparent,rgba(255,255,255,0.2),transparent);"></div>
        <div style="position:relative;z-index:1;">
            <div style="font-size:2.8rem;margin-bottom:6px;filter:drop-shadow(0 3px 6px rgba(0,0,0,0.12));">✂️</div>
            <h1 style="color: white; margin: 0; font-weight: 900; font-size: 2rem; text-shadow: 0 2px 8px rgba(0,0,0,0.08);">Recortador de Programaciones</h1>
            <div style="width: 70px; height: 3px; background: linear-gradient(90deg, #61C1C3, rgba(255,255,255,0.5)); border-radius: 2px; margin: 14px auto;"></div>
            <p style="color: rgba(255,255,255,0.9); margin: 5px 0 0 0; font-size: 0.95rem; font-weight: 500;">Extrae programaciones iniciales filtradas por base</p>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Selector de base
    bases_recorte = [
        ('BCN', 'Barcelona (BCN) → Genera 2 PDFs: JC y TCP'),
        ('AGP', 'Málaga (AGP)'),
        ('ALC', 'Alicante (ALC)'),
        ('BIO', 'Bilbao (BIO)'),
        ('LPA', 'Gran Canaria (LPA)'),
        ('PMI', 'Palma de Mallorca (PMI)'),
        ('SCQ', 'Santiago de Compostela (SCQ)'),
        ('SVQ', 'Sevilla (SVQ)'),
        ('TFN', 'Tenerife Norte (TFN)'),
        ('VLC', 'Valencia (VLC)'),
        ('IBZ', 'Ibiza (IBZ)'),
    ]
    
    st.markdown("### 📍 Selecciona la Base")
    base_seleccionada = st.selectbox(
        "Base para filtrar:",
        options=[b[0] for b in bases_recorte],
        format_func=lambda x: dict(bases_recorte).get(x, x),
        key="recorte_base"
    )
    
    if base_seleccionada == 'BCN':
        st.info("ℹ️ Para **BCN** se generarán **dos PDFs separados**: uno con los **JC** (Jefes de Cabina) y otro con los **TCP** (Tripulantes de Cabina).")
    
    st.markdown("### 📁 Sube el PDF de Programaciones Iniciales")
    pdf_file = st.file_uploader(
        "Arrastra o selecciona el archivo PDF",
        type=['pdf'],
        key="recorte_pdf"
    )
    
    if pdf_file is not None:
        st.success(f"✅ Archivo cargado: {pdf_file.name}")
        
        if st.button("✂️ Recortar Programaciones", type="primary", use_container_width=True):
            try:
                import PyPDF2
                from PyPDF2 import PdfReader, PdfWriter
                import re as re_mod
                
                # Leer el PDF
                pdf_reader = PdfReader(pdf_file)
                total_pages = len(pdf_reader.pages)
                
                # Animación de progreso
                st.markdown("""
                <style>
                @keyframes flyPlane {
                    0% { transform: translateX(-30px) rotate(0deg); }
                    50% { transform: translateX(30px) rotate(2deg); }
                    100% { transform: translateX(-30px) rotate(0deg); }
                }
                .cutting-animation {
                    text-align: center; padding: 20px;
                }
                .cutting-animation .plane {
                    animation: flyPlane 1.5s ease-in-out infinite;
                    display: inline-block; font-size: 2rem;
                }
                </style>
                <div class="cutting-animation">
                    <span class="plane">✈️</span>
                    <p style="margin-top: 10px; font-size: 1.1rem; color: #0D5F5D; font-weight: 600;">Analizando páginas del PDF...</p>
                </div>
                """, unsafe_allow_html=True)
                
                progress_bar = st.progress(0, text="Analizando páginas...")
                
                # Patrón: BASE FLEET CATEGORY (ej: "SVQ 320 JC" o "BCN 320 TC")
                base_pattern = re_mod.compile(r'^([A-Z]{3})\s+\d{3}\s+(JC|TC)\b')
                
                # Analizar TODAS las bases que aparecen en CADA página
                # Una página puede tener tripulantes de múltiples bases
                page_bases = []  # lista de sets, un set por página
                
                for i, page in enumerate(pdf_reader.pages):
                    progress_bar.progress((i + 1) / total_pages, text=f"Analizando página {i+1} de {total_pages}...")
                    
                    text = page.extract_text() or ""
                    lines = text.split('\n')
                    
                    bases_en_pagina = set()  # todas las bases en esta página
                    cats_en_pagina = {}  # base -> set of categories
                    
                    for line in lines:
                        m = base_pattern.match(line.strip())
                        if m:
                            b = m.group(1)
                            c = m.group(2)
                            bases_en_pagina.add(b)
                            if b not in cats_en_pagina:
                                cats_en_pagina[b] = set()
                            cats_en_pagina[b].add(c)
                    
                    page_bases.append({
                        'bases': bases_en_pagina,
                        'cats': cats_en_pagina
                    })
                
                progress_bar.progress(1.0, text="✅ Análisis completado")
                
                if base_seleccionada == 'BCN':
                    # Para BCN: generar 2 PDFs separados (JC y TCP)
                    # Incluir cualquier página donde aparezca al menos un tripulante BCN
                    writer_jc = PdfWriter()
                    writer_tcp = PdfWriter()
                    pages_jc = []
                    pages_tcp = []
                    
                    for i, pb in enumerate(page_bases):
                        if 'BCN' in pb['bases']:
                            bcn_cats = pb['cats'].get('BCN', set())
                            if 'JC' in bcn_cats:
                                writer_jc.add_page(pdf_reader.pages[i])
                                pages_jc.append(i + 1)
                            if 'TC' in bcn_cats:
                                writer_tcp.add_page(pdf_reader.pages[i])
                                pages_tcp.append(i + 1)
                    
                    if not pages_jc and not pages_tcp:
                        st.warning("⚠️ No se encontraron tripulantes de BCN en este PDF")
                    else:
                        st.success(f"✅ BCN: **{len(pages_jc)}** páginas con JC y **{len(pages_tcp)}** páginas con TCP")
                        
                        col_dl1, col_dl2 = st.columns(2)
                        
                        if pages_jc:
                            output_jc = io.BytesIO()
                            writer_jc.write(output_jc)
                            output_jc.seek(0)
                            with col_dl1:
                                st.download_button(
                                    label=f"📥 Descargar JC BCN ({len(pages_jc)} págs)",
                                    data=output_jc,
                                    file_name=f"BCN_JC.pdf",
                                    mime="application/pdf",
                                    type="primary",
                                    use_container_width=True
                                )
                        
                        if pages_tcp:
                            output_tcp = io.BytesIO()
                            writer_tcp.write(output_tcp)
                            output_tcp.seek(0)
                            with col_dl2:
                                st.download_button(
                                    label=f"📥 Descargar TCP BCN ({len(pages_tcp)} págs)",
                                    data=output_tcp,
                                    file_name=f"BCN_TCP.pdf",
                                    mime="application/pdf",
                                    type="primary",
                                    use_container_width=True
                                )
                else:
                    # Para otras bases: incluir CUALQUIER página donde aparezca
                    # al menos un tripulante de la base seleccionada
                    writer = PdfWriter()
                    pages_found = []
                    
                    for i, pb in enumerate(page_bases):
                        if base_seleccionada in pb['bases']:
                            writer.add_page(pdf_reader.pages[i])
                            pages_found.append(i + 1)
                    
                    if pages_found:
                        output = io.BytesIO()
                        writer.write(output)
                        output.seek(0)
                        
                        st.success(f"✅ Se encontraron **{len(pages_found)}** páginas con tripulantes de {base_seleccionada}")
                        
                        st.download_button(
                            label=f"📥 Descargar PDF de {base_seleccionada} ({len(pages_found)} págs)",
                            data=output,
                            file_name=f"PROGRAMACION_{base_seleccionada}.pdf",
                            mime="application/pdf",
                            type="primary",
                            use_container_width=True
                        )
                    else:
                        st.warning(f"⚠️ No se encontraron tripulantes de la base {base_seleccionada} en este PDF")
                        
            except Exception as e:
                st.error(f"❌ Error al procesar el PDF: {str(e)}")
    
    st.markdown('<p style="text-align:center;color:#94A3B8;font-size:0.8rem;font-weight:500;margin-top:40px;">✈️ Herramientas Sindicales © 2026</p>', unsafe_allow_html=True)
    st.stop()

# ===== SECCIÓN: PANEL DE ADMINISTRACIÓN (standalone) =====
if st.session_state['app_section'] == 'admin_panel':
    render_admin_sidebar()
    if not st.session_state.get('is_admin'):
        st.error("⚠️ No tienes permisos para acceder a esta sección")
        st.session_state['app_section'] = 'seleccion'
        st.rerun()
    
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
    * { font-family: 'Inter', sans-serif; }
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}

    /* Botones Secondary - glassmorphism elegante */
    div[data-testid="stButton"] > button[kind="secondary"],
    .stButton > button[kind="secondary"],
    button[kind="secondary"],
    button[data-testid="stBaseButton-secondary"] {
        background: rgba(255,255,255,0.8) !important;
        backdrop-filter: blur(10px) !important;
        color: #0D5F5D !important;
        border: 1.5px solid rgba(17,127,124,0.2) !important;
        font-weight: 600 !important;
        padding: 10px 20px !important;
        border-radius: 12px !important;
        box-shadow: 0 2px 10px rgba(17,127,124,0.08) !important;
        transition: all 0.3s ease !important;
    }
    div[data-testid="stButton"] > button[kind="secondary"]:hover,
    .stButton > button[kind="secondary"]:hover,
    button[kind="secondary"]:hover,
    button[data-testid="stBaseButton-secondary"]:hover {
        transform: translateY(-2px) !important;
        background: rgba(17,127,124,0.08) !important;
        border-color: #117F7C !important;
        box-shadow: 0 6px 20px rgba(17,127,124,0.15) !important;
        color: #117F7C !important;
    }
    </style>
    """, unsafe_allow_html=True)
    
    _usr_col1, _usr_col2, _usr_col3 = st.columns([1, 5, 1])
    with _usr_col1:
        if st.button("← Volver", key="back_admin"):
            st.session_state['app_section'] = 'seleccion'
            st.rerun()
    with _usr_col2:
        st.markdown(f"👤 Conectado como: **{st.session_state.get('username', '')}** 👑")
    with _usr_col3:
        if st.button("🚪 Cerrar Sesión", key="logout_admin", type="secondary"):
            update_logout_time()
            for key in ['authenticated', 'username', 'is_admin']:
                st.session_state[key] = False if key == 'authenticated' else None
            st.session_state['app_section'] = 'seleccion'
            st.rerun()
    
    st.markdown("""
    <div style="background: linear-gradient(135deg, #0A5C5A 0%, #0D6F6D 15%, #117F7C 35%, #1A9E9B 60%, #44BABC 85%, #61C1C3 100%); 
                padding: 35px 44px; 
                border-radius: 24px; 
                text-align: center;
                margin: 15px 0 30px 0;
                position: relative;
                overflow: hidden;
                box-shadow: 0 16px 50px rgba(17,127,124,0.3), 0 6px 20px rgba(11,132,127,0.15);">
        <div style="position:absolute;top:-50%;right:-20%;width:400px;height:400px;background:radial-gradient(circle,rgba(97,193,195,0.15) 0%,transparent 70%);border-radius:50%;"></div>
        <div style="position:absolute;bottom:0;left:0;right:0;height:1px;background:linear-gradient(90deg,transparent,rgba(255,255,255,0.2),transparent);"></div>
        <div style="position:relative;z-index:1;">
            <div style="font-size:2.8rem;margin-bottom:6px;filter:drop-shadow(0 3px 6px rgba(0,0,0,0.12));">👑</div>
            <h1 style="color: white; margin: 0; font-weight: 900; font-size: 2rem; text-shadow: 0 2px 8px rgba(0,0,0,0.08);">Panel de Administración</h1>
            <div style="width: 70px; height: 3px; background: linear-gradient(90deg, #61C1C3, rgba(255,255,255,0.5)); border-radius: 2px; margin: 14px auto;"></div>
            <p style="color: rgba(255,255,255,0.9); margin: 5px 0 0 0; font-size: 0.95rem; font-weight: 500;">Control y monitoreo del sistema</p>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    admin_tabs = st.tabs(["📊 Logs de Acceso", "📄 Informes Generados", "📧 Notificaciones Email", "👥 Gestión de Usuarios"])
    
    with admin_tabs[0]:
        st.markdown('<p class="section-title">📊 Logs de Acceso</p>', unsafe_allow_html=True)
        
        if os.path.exists(LOGS_FILE):
            try:
                with open(LOGS_FILE, 'r') as f:
                    logs = json.load(f)
            except:
                logs = []
            
            if logs:
                df_logs = pd.DataFrame(logs)
                # Mostrar columnas adicionales si existen
                display_cols = ['usuario', 'fecha', 'hora']
                if 'ip' in df_logs.columns:
                    display_cols.append('ip')
                if 'ubicacion' in df_logs.columns:
                    display_cols.append('ubicacion')
                if 'hora_desconexion' in df_logs.columns:
                    display_cols.append('hora_desconexion')
                
                df_logs_display = df_logs[display_cols].copy()
                col_rename = {
                    'usuario': 'Usuario', 'fecha': 'Fecha', 'hora': 'Hora Conexión',
                    'ip': 'IP', 'ubicacion': 'Ubicación', 'hora_desconexion': 'Hora Desconexión'
                }
                df_logs_display.columns = [col_rename.get(c, c) for c in df_logs_display.columns]
                
                # KPIs de acceso
                _ac1, _ac2, _ac3, _ac4 = st.columns(4)
                _ac1.markdown(f'<div class="kpi-box"><p>TOTAL ACCESOS</p><h1>{len(logs)}</h1></div>', unsafe_allow_html=True)
                _ac2.markdown(f'<div class="kpi-box"><p>USUARIOS ÚNICOS</p><h1>{df_logs["usuario"].nunique()}</h1></div>', unsafe_allow_html=True)
                _ac3.markdown(f'<div class="kpi-box accent"><p>ÚLTIMO USUARIO</p><h1 style="font-size:1.3rem;">{logs[-1]["usuario"]}</h1></div>', unsafe_allow_html=True)
                _ac4.markdown(f'<div class="kpi-box"><p>ÚLTIMO ACCESO</p><h1 style="font-size:1.3rem;">{logs[-1]["fecha"]}<br>{logs[-1]["hora"]}</h1></div>', unsafe_allow_html=True)
                
                st.markdown("<br>", unsafe_allow_html=True)
                
                # Tabla de logs (últimos 50)
                st.markdown("#### 📋 Últimos accesos (con IP y ubicación)")
                st.dataframe(df_logs_display.iloc[::-1].head(50), use_container_width=True, hide_index=True)
                
                # Accesos por usuario
                st.markdown("#### 📈 Accesos por usuario")
                accesos_por_usuario = df_logs['usuario'].value_counts().reset_index()
                accesos_por_usuario.columns = ['Usuario', 'Accesos']
                fig_accesos = px.bar(accesos_por_usuario, x='Usuario', y='Accesos',
                                     color_discrete_sequence=['#117F7C'])
                fig_accesos.update_layout(
                    plot_bgcolor='white',
                    paper_bgcolor='white',
                    font=dict(family='Inter'),
                    xaxis_title='',
                    yaxis_title='Nº Accesos'
                )
                st.plotly_chart(fig_accesos, use_container_width=True)
                
                # Accesos por día
                st.markdown("#### 📅 Accesos por día")
                accesos_por_dia = df_logs['fecha'].value_counts().sort_index().reset_index()
                accesos_por_dia.columns = ['Fecha', 'Accesos']
                fig_dias = px.line(accesos_por_dia, x='Fecha', y='Accesos',
                                   markers=True, color_discrete_sequence=['#117F7C'])
                fig_dias.update_layout(
                    plot_bgcolor='white',
                    paper_bgcolor='white',
                    font=dict(family='Inter'),
                    xaxis_title='',
                    yaxis_title='Nº Accesos'
                )
                st.plotly_chart(fig_dias, use_container_width=True)
                
                # Descargar logs
                st.markdown("---")
                st.download_button(
                    "📥 Descargar Logs completos (JSON)",
                    json.dumps(logs, indent=2),
                    "access_logs.json",
                    "application/json",
                    use_container_width=True
                )
            else:
                st.info("No hay logs de acceso aún")
        else:
            st.info("No hay logs de acceso aún")
    
    with admin_tabs[1]:
        st.markdown('<p class="section-title">📄 Informes Generados</p>', unsafe_allow_html=True)
        if os.path.exists(INFORMES_DIR):
            informes = sorted([f for f in os.listdir(INFORMES_DIR) if f.endswith('.json')], reverse=True)
            if informes:
                informes_data = []
                for inf_file in informes[:30]:
                    try:
                        with open(os.path.join(INFORMES_DIR, inf_file), 'r') as f:
                            inf = json.load(f)
                        informes_data.append({
                            'Usuario': inf.get('usuario', ''),
                            'Fecha': inf.get('fecha', ''),
                            'Base': inf.get('base', ''),
                            'Tripulantes': inf.get('tripulantes_analizados', 0)
                        })
                    except:
                        pass
                if informes_data:
                    st.dataframe(pd.DataFrame(informes_data), use_container_width=True, hide_index=True)
                else:
                    st.info("No hay informes generados aún")
            else:
                st.info("No hay informes generados aún")
        else:
            st.info("No hay informes generados aún")
    
    with admin_tabs[2]:
        st.markdown('<p class="section-title">📧 Notificaciones de Email Enviadas</p>', unsafe_allow_html=True)
        notif_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'email_notifications.json')
        if os.path.exists(notif_file):
            try:
                with open(notif_file, 'r') as f:
                    notifications = json.load(f)
                if notifications:
                    df_notif = pd.DataFrame(notifications)
                    df_notif_display = df_notif[['username', 'time', 'ip', 'location']].copy()
                    df_notif_display.columns = ['Usuario', 'Fecha/Hora', 'IP', 'Ubicación']
                    st.dataframe(df_notif_display.iloc[::-1].head(50), use_container_width=True, hide_index=True)
                else:
                    st.info("No hay notificaciones enviadas aún")
            except:
                st.info("No hay notificaciones enviadas aún")
        else:
            st.info("No hay notificaciones enviadas aún")
    
    with admin_tabs[3]:
        st.markdown('<p class="section-title">👥 Gestión de Usuarios</p>', unsafe_allow_html=True)
        
        st.markdown("""
        <div style="background: #F0F7F7; border-radius: 10px; padding: 14px 20px; margin-bottom: 18px; border-left: 4px solid #117F7C;">
            <span style="color: #0D5F5D; font-size: 0.92rem; font-weight: 500;">
                Aquí puedes ver todos los usuarios con acceso al sistema.
            </span>
        </div>
        """, unsafe_allow_html=True)
        
        # Mostrar usuarios actuales
        st.markdown("#### 📋 Usuarios Registrados")
        
        users_list = []
        for username in USERS.keys():
            status = "👑 Admin" if username == "admin" else "✅ Activo"
            display_name = username.replace('.', ' ').title()
            users_list.append({
                'Usuario': username,
                'Nombre': display_name,
                'Estado': status
            })
        
        df_users = pd.DataFrame(users_list)
        st.dataframe(df_users, use_container_width=True, hide_index=True)
        
        st.markdown(f"**Total usuarios:** {len(USERS)}")
        
        # (Sección de añadir usuario eliminada)
    
    st.markdown("---")
    st.markdown('<p style="text-align:center;color:#94A3B8;font-size:0.8rem;font-weight:500;">✈️ Herramientas Sindicales © 2026</p>', unsafe_allow_html=True)
    st.stop()

# ===== SECCIÓN: PDF GENERATOR (standalone) =====
if st.session_state['app_section'] == 'pdf_generator':
    # Renderizar sidebar de admin si corresponde
    render_admin_sidebar()
    
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
    * { font-family: 'Inter', sans-serif; }
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}

    /* Botones Secondary - glassmorphism elegante */
    div[data-testid="stButton"] > button[kind="secondary"],
    .stButton > button[kind="secondary"],
    button[kind="secondary"],
    button[data-testid="stBaseButton-secondary"] {
        background: rgba(255,255,255,0.8) !important;
        backdrop-filter: blur(10px) !important;
        color: #0D5F5D !important;
        border: 1.5px solid rgba(17,127,124,0.2) !important;
        font-weight: 600 !important;
        padding: 10px 20px !important;
        border-radius: 12px !important;
        box-shadow: 0 2px 10px rgba(17,127,124,0.08) !important;
        transition: all 0.3s ease !important;
    }
    div[data-testid="stButton"] > button[kind="secondary"]:hover,
    .stButton > button[kind="secondary"]:hover,
    button[kind="secondary"]:hover,
    button[data-testid="stBaseButton-secondary"]:hover {
        transform: translateY(-2px) !important;
        background: rgba(17,127,124,0.08) !important;
        border-color: #117F7C !important;
        box-shadow: 0 6px 20px rgba(17,127,124,0.15) !important;
        color: #117F7C !important;
    }
    </style>
    """, unsafe_allow_html=True)
    
    _usr_col1, _usr_col2, _usr_col3 = st.columns([1, 5, 1])
    with _usr_col1:
        if st.button("← Volver", key="back_pdf"):
            st.session_state['app_section'] = 'seleccion'
            st.rerun()
    with _usr_col2:
        st.markdown(f"👤 Conectado como: **{st.session_state.get('username', '')}**" + 
                    (" 👑" if st.session_state.get('is_admin') else ""))
    with _usr_col3:
        if st.button("🚪 Cerrar Sesión", key="logout_pdf", type="secondary"):
            update_logout_time()
            for key in ['authenticated', 'username', 'is_admin']:
                st.session_state[key] = False if key == 'authenticated' else None
            st.session_state['app_section'] = 'seleccion'
            st.rerun()
    
    st.markdown("""
    <div style="background: linear-gradient(135deg, #0A5C5A 0%, #0D6F6D 15%, #117F7C 35%, #1A9E9B 60%, #44BABC 85%, #61C1C3 100%); 
                padding: 35px 44px; border-radius: 24px; text-align: center; margin-bottom: 28px;
                position: relative; overflow: hidden;
                box-shadow: 0 16px 50px rgba(17,127,124,0.3), 0 6px 20px rgba(11,132,127,0.15);">
        <div style="position:absolute;top:-50%;right:-20%;width:400px;height:400px;background:radial-gradient(circle,rgba(97,193,195,0.15) 0%,transparent 70%);border-radius:50%;"></div>
        <div style="position:absolute;bottom:0;left:0;right:0;height:1px;background:linear-gradient(90deg,transparent,rgba(255,255,255,0.2),transparent);"></div>
        <div style="position:relative;z-index:1;">
            <div style="font-size:2.8rem;margin-bottom:6px;filter:drop-shadow(0 3px 6px rgba(0,0,0,0.12));">📄</div>
            <h1 style="color: white; margin: 0; font-size: 2rem; font-weight: 900; text-shadow: 0 2px 8px rgba(0,0,0,0.08);">Generador de PDFs</h1>
            <div style="width: 70px; height: 3px; background: linear-gradient(90deg, #61C1C3, rgba(255,255,255,0.5)); border-radius: 2px; margin: 14px auto;"></div>
            <p style="color: rgba(255,255,255,0.9); font-size: 0.95rem; margin: 0; font-weight: 500;">Genera PDFs con el formato de las programaciones iniciales desde archivos de roster (.TXT)</p>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("""
        <div style="background: #F0F7F7; border-radius: 10px; padding: 14px 20px; margin-bottom: 18px; border-left: 4px solid #117F7C;">
            <span style="color: #0D5F5D; font-size: 0.92rem; font-weight: 500;">
                📋 Selecciona la base y sube los archivos correspondientes para generar PDFs con el formato de las programaciones iniciciales.
            </span>
        </div>
    """, unsafe_allow_html=True)
    
    # Selector de base para el generador de PDFs CON opción por defecto
    pdf_base_options = [('', '🌍 Selecciona una base...')] + [('BCN', 'Barcelona (BCN)')] + [(b[0], b[1]) for b in BASES_ORDENADAS if b[0] != 'BCN']
    pdf_base_options.append(('RESTO', 'Genera un único PDF del resto de bases'))
    pdf_base_options.append(('DESPLAZADOS', '🔄 Desplazados / Segunda Residencia'))
    
    pdf_base_sel = st.selectbox(
        "🌍 Selecciona la base para generar PDF:",
        [b[0] for b in pdf_base_options],
        format_func=lambda x: dict(pdf_base_options).get(x, x),
        key="pdf_base_selector",
        index=0  # Seleccionar la primera opción (vacía) por defecto
    )
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Verificar que se ha seleccionado una base válida
    if not pdf_base_sel:
        st.info("👆 Por favor, selecciona una base para continuar.")
    elif pdf_base_sel == 'BCN':
        # BCN: 2 archivos TXT
        st.markdown("#### 📥 Barcelona (BCN) - Sube 2 archivos TXT")
        col1, col2 = st.columns(2)
        with col1:
            pdf_bcn_1 = st.file_uploader("📂 Primer archivo (TCPs)", type=["txt"], key="pdf_bcn_1")
        with col2:
            pdf_bcn_2 = st.file_uploader("📂 Segundo archivo (JCs)", type=["txt"], key="pdf_bcn_2")
        
        if pdf_bcn_1 and pdf_bcn_2:
            st.success("✅ Ambos archivos cargados")
            if st.button("🚀 Generar 2 PDFs para BCN", key="gen_pdf_bcn", use_container_width=True, type="primary"):
                with st.spinner("Procesando archivos de BCN..."):
                    try:
                        pdf_outputs = []
                        for idx, uploaded_file in enumerate([pdf_bcn_1, pdf_bcn_2], 1):
                            file_label = "TCPs" if idx == 1 else "JCs"
                            if uploaded_file.name.lower().endswith('.txt'):
                                raw_content = uploaded_file.read().decode("utf-8", errors="replace")
                                pages = extract_crew_pages_from_text(raw_content)
                                if pages:
                                    pdf_bytes, n_pages = generate_roster_pdf_bytes(pages)
                                    pdf_outputs.append((f"BCN_{file_label}.pdf", pdf_bytes, n_pages))
                                else:
                                    st.warning(f"⚠️ No se encontraron páginas válidas en el archivo {file_label}")
                            else:
                                # Para PDFs, indicar limitación actual
                                st.info(f"📄 Archivo PDF ({file_label}) detectado")
                                st.warning(f"⚠️ Para {file_label}: usa archivo .txt para generar PDF formateado. La conversión desde PDF está en desarrollo.")
                        
                        if pdf_outputs:
                            st.success(f"✅ {len(pdf_outputs)} PDF(s) generado(s)")
                            for filename, pdf_bytes, n_pages in pdf_outputs:
                                st.download_button(
                                    label=f"📥 Descargar {filename} ({n_pages} páginas)",
                                    data=pdf_bytes,
                                    file_name=filename,
                                    mime="application/pdf",
                                    key=f"download_{filename}"
                                )
                    except Exception as e:
                        st.error(f"❌ Error procesando archivos BCN: {str(e)}")
    
    elif pdf_base_sel == 'RESTO':
        # Resto de bases: 1 archivo general (excluye BCN)
        st.markdown("#### 📥 Resto de bases - Sube 1 archivo general")
        pdf_resto = st.file_uploader("📂 Archivo general (todas las bases excepto BCN)", type=["txt"], key="pdf_resto")
        
        if pdf_resto:
            if pdf_resto.name.endswith('.txt'):
                raw = pdf_resto.read().decode("utf-8", errors="replace")
                pages_with_base = extract_crew_pages_with_base(raw)
                
                # Mostrar resumen de bases detectadas (usando base OPERATIVA)
                bases_detected = {}
                desplazados_count = 0
                for page, base_of, base_op, es_desp in pages_with_base:
                    base_key = base_op if base_op else (base_of if base_of else 'Sin identificar')
                    bases_detected[base_key] = bases_detected.get(base_key, 0) + 1
                    if es_desp:
                        desplazados_count += 1
                
                if pages_with_base:
                    # Mostrar resumen
                    bases_summary = ", ".join([f"{b}: {c}" for b, c in sorted(bases_detected.items())])
                    st.success(f"✅ Se encontraron **{len(pages_with_base)}** páginas de tripulantes")
                    st.info(f"📊 Distribución por bases (operativa): {bases_summary}")
                    if desplazados_count > 0:
                        st.warning(f"🔄 **{desplazados_count}** tripulante(s) desplazado(s) detectado(s) (base oficial ≠ base operativa)")
                    
                    # Filtrar excluyendo BCN (usando base OPERATIVA)
                    filtered_pages = [page for page, base_of, base_op, es_desp in pages_with_base 
                                      if (base_op or base_of) != 'BCN']
                    
                    if filtered_pages:
                        if st.button("🚀 Generar PDF único del resto de bases", key="gen_pdf_resto", use_container_width=True, type="primary"):
                            with st.spinner("Generando PDF..."):
                                pdf_bytes, total_pages = generate_roster_pdf_bytes(filtered_pages)
                            if pdf_bytes:
                                st.success(f"✅ PDF generado correctamente · {len(filtered_pages)} páginas · {len(pdf_bytes)/1024:.1f} KB")
                                st.download_button(
                                    "📥 Descargar PDF",
                                    data=pdf_bytes,
                                    file_name=f"programacion_resto_bases.pdf",
                                    mime="application/pdf",
                                    use_container_width=True
                                )
                    else:
                        st.warning("⚠️ No se encontraron tripulantes fuera de BCN en el archivo")
                else:
                    st.warning("⚠️ No se encontraron páginas de tripulantes en el archivo")
            else:
                st.info("📄 Funcionalidad para archivos PDF en desarrollo. Por ahora usa archivos .txt")
    
    elif pdf_base_sel == 'DESPLAZADOS':
        # Nueva opción: Tripulantes desplazados / Segunda Residencia
        st.markdown("#### 📥 Tripulantes Desplazados / Segunda Residencia")
        st.markdown("""
            <div style="background: #FFF3E0; border-radius: 10px; padding: 14px 20px; margin-bottom: 18px; border-left: 4px solid #FF9800;">
                <span style="color: #E65100; font-size: 0.92rem; font-weight: 500;">
                    🔄 <strong>Desplazados:</strong> Tripulantes cuya base oficial difiere de donde operan la mayoría de días del mes.
                    Se asignan a la base operativa real y se marcan como desplazados.
                </span>
            </div>
        """, unsafe_allow_html=True)
        
        pdf_desplazados = st.file_uploader("📂 Archivo general con todas las bases", type=["txt"], key="pdf_desplazados")
        
        if pdf_desplazados:
            if pdf_desplazados.name.endswith('.txt'):
                raw = pdf_desplazados.read().decode("utf-8", errors="replace")
                pages_with_base = extract_crew_pages_with_base(raw)
                
                # Filtrar solo desplazados
                desplazados = [(page, base_of, base_op, es_desp) for page, base_of, base_op, es_desp in pages_with_base if es_desp]
                
                if desplazados:
                    st.success(f"✅ Se encontraron **{len(desplazados)}** tripulantes desplazados")
                    
                    # Mostrar desglose
                    desglose = {}
                    for page, base_of, base_op, es_desp in desplazados:
                        key = f"{base_of} → {base_op}"
                        desglose[key] = desglose.get(key, 0) + 1
                    
                    desglose_text = " | ".join([f"**{k}**: {v}" for k, v in sorted(desglose.items())])
                    st.info(f"📊 Movimientos: {desglose_text}")
                    
                    if st.button("🚀 Generar PDF de Desplazados", key="gen_pdf_desplazados", use_container_width=True, type="primary"):
                        with st.spinner("Generando PDF de desplazados..."):
                            desplazados_pages = [page for page, base_of, base_op, es_desp in desplazados]
                            pdf_bytes, total_pages = generate_roster_pdf_bytes(desplazados_pages)
                        if pdf_bytes:
                            st.success(f"✅ PDF generado correctamente · {len(desplazados_pages)} páginas · {len(pdf_bytes)/1024:.1f} KB")
                            st.download_button(
                                "📥 Descargar PDF Desplazados",
                                data=pdf_bytes,
                                file_name=f"programacion_DESPLAZADOS.pdf",
                                mime="application/pdf",
                                use_container_width=True
                            )
                else:
                    st.info("ℹ️ No se encontraron tripulantes desplazados en este archivo. Todos operan desde su base oficial.")
            else:
                st.info("📄 Funcionalidad para archivos PDF en desarrollo. Por ahora usa archivos .txt")
    
    else:
        # Otras bases: 1 archivo general, generar solo de esa base
        # MEJORA: Incluir tripulantes de la base + desplazados que operan desde ella
        base_nombre = dict(pdf_base_options).get(pdf_base_sel, pdf_base_sel)
        st.markdown(f"#### 📥 {base_nombre} - Sube 1 archivo general")
        pdf_otra_base = st.file_uploader(f"📂 Archivo general (se extraerá solo {pdf_base_sel})", type=["txt"], key=f"pdf_{pdf_base_sel}")
        
        if pdf_otra_base:
            if pdf_otra_base.name.endswith('.txt'):
                raw = pdf_otra_base.read().decode("utf-8", errors="replace")
                pages_with_base = extract_crew_pages_with_base(raw)
                
                # Mostrar resumen de bases detectadas (por base operativa)
                bases_detected = {}
                for page, base_of, base_op, es_desp in pages_with_base:
                    base_key = base_op if base_op else (base_of if base_of else 'Sin identificar')
                    bases_detected[base_key] = bases_detected.get(base_key, 0) + 1
                
                if pages_with_base:
                    # Mostrar resumen
                    bases_summary = ", ".join([f"{b}: {c}" for b, c in sorted(bases_detected.items())])
                    st.info(f"📊 Distribución de bases (operativa) en archivo: {bases_summary}")
                    
                    # Filtrar: base oficial = seleccionada O base operativa = seleccionada
                    # Primero los de base oficial = seleccionada, luego los desplazados
                    pages_base_oficial = [page for page, base_of, base_op, es_desp in pages_with_base 
                                          if base_of == pdf_base_sel and not es_desp]
                    pages_desplazados_aqui = [page for page, base_of, base_op, es_desp in pages_with_base 
                                              if es_desp and base_op == pdf_base_sel]
                    
                    filtered_pages = pages_base_oficial + pages_desplazados_aqui
                    
                    if filtered_pages:
                        msg = f"✅ Se encontraron **{len(pages_base_oficial)}** tripulantes de {pdf_base_sel}"
                        if pages_desplazados_aqui:
                            msg += f" + **{len(pages_desplazados_aqui)}** desplazado(s) operando desde {pdf_base_sel}"
                        st.success(msg)
                        
                        if st.button(f"🚀 Generar PDF de {pdf_base_sel}", key=f"gen_pdf_{pdf_base_sel}", use_container_width=True, type="primary"):
                            with st.spinner("Generando PDF..."):
                                pdf_bytes, total_pages = generate_roster_pdf_bytes(filtered_pages)
                            if pdf_bytes:
                                st.success(f"✅ PDF generado correctamente · {len(filtered_pages)} páginas · {len(pdf_bytes)/1024:.1f} KB")
                                st.download_button(
                                    "📥 Descargar PDF",
                                    data=pdf_bytes,
                                    file_name=f"programacion_{pdf_base_sel}.pdf",
                                    mime="application/pdf",
                                    use_container_width=True
                                )
                    else:
                        st.warning(f"⚠️ No se encontraron tripulantes de la base {pdf_base_sel} en el archivo")
                        if bases_detected:
                            st.info(f"📋 Bases disponibles en el archivo: {', '.join([k for k in bases_detected.keys() if k != 'Sin identificar'])}")
                else:
                    st.warning("⚠️ No se encontraron páginas de tripulantes en el archivo")
            else:
                st.info("📄 Funcionalidad para archivos PDF en desarrollo. Por ahora usa archivos .txt")
    
    st.markdown("---")
    st.markdown('<p style="text-align:center;color:#94A3B8;font-size:0.8rem;font-weight:500;">✈️ Herramientas Sindicales © 2026</p>', unsafe_allow_html=True)
    st.stop()

# =====================================================================
# SECCIÓN: TABLAS SALARIALES 2026
# =====================================================================
if st.session_state['app_section'] == 'tablas_salariales':
    render_admin_sidebar()
    _usr_col1, _usr_col2, _usr_col3 = st.columns([1, 5, 1])
    with _usr_col1:
        if st.button("⬅️ Volver", key="volver_tablas"):
            st.session_state['app_section'] = 'seleccion'
            st.rerun()
    with _usr_col3:
        if st.button("🚪 Cerrar Sesión", key="logout_tablas"):
            for key in ['authenticated', 'username', 'is_admin']:
                st.session_state[key] = False if key == 'authenticated' else None
            st.session_state['app_section'] = 'seleccion'
            st.rerun()

    # --- HEADER Tablas Salariales ---
    st.markdown("""
    <div style="background: linear-gradient(135deg, #0A5C5A 0%, #0D6F6D 15%, #117F7C 35%, #1A9E9B 60%, #44BABC 85%, #61C1C3 100%);
                padding: 35px 40px;
                border-radius: 24px;
                text-align: center;
                position: relative;
                overflow: hidden;
                box-shadow: 0 16px 50px rgba(17,127,124,0.3), 0 6px 20px rgba(11,132,127,0.15);
                margin-bottom: 28px;">
        <div style="position:absolute;top:-50%;right:-20%;width:400px;height:400px;background:radial-gradient(circle,rgba(97,193,195,0.15) 0%,transparent 70%);border-radius:50%;"></div>
        <div style="position:absolute;bottom:0;left:0;right:0;height:1px;background:linear-gradient(90deg,transparent,rgba(255,255,255,0.2),transparent);"></div>
        <div style="position:relative;z-index:1;">
            <div style="font-size: 3rem; margin-bottom: 8px; filter: drop-shadow(0 3px 6px rgba(0,0,0,0.12));">💶</div>
            <h1 style="color: white; font-size: 2rem; font-weight: 900; margin: 0 0 6px 0; letter-spacing: -0.5px; text-shadow: 0 2px 8px rgba(0,0,0,0.08);">Tablas Salariales 2026</h1>
            <div style="width:70px;height:3px;background:linear-gradient(90deg,#61C1C3,rgba(255,255,255,0.5));border-radius:2px;margin:12px auto;"></div>
            <p style="color: rgba(255,255,255,0.9); font-size: 0.95rem; font-weight: 500; margin: 0;">Retribuciones vigentes · Todos los niveles</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # --- CSS premium para tablas salariales ---
    st.markdown("""<style>
    .sal-table { width:100%; border-collapse:separate; border-spacing:0; font-size:0.72rem; margin-bottom:12px; border-radius:14px; overflow:hidden; box-shadow:0 4px 20px rgba(17,127,124,0.08), 0 1px 4px rgba(0,0,0,0.03); border:1px solid rgba(17,127,124,0.08); }
    .sal-table th { background:linear-gradient(135deg,#0D6F6D,#117F7C,#1A9E9B); color:#fff; padding:8px 8px; text-align:center; font-weight:700; font-size:0.7rem; letter-spacing:0.4px; border:none; text-transform:uppercase; border-bottom:2px solid rgba(255,255,255,0.15); }
    .sal-table th:first-child { text-align:left; min-width:155px; padding-left:14px; }
    .sal-table td { padding:6px 8px; text-align:center; border-bottom:1px solid rgba(17,127,124,0.06); font-size:0.72rem; transition: all 0.2s ease; }
    .sal-table td:first-child { text-align:left; font-weight:600; color:#0D5F5D; background:rgba(240,250,250,0.6); border-left:4px solid #44BABC; padding-left:12px; }
    .sal-table tr:nth-child(even) td { background:rgba(241,250,250,0.5); }
    .sal-table tr:nth-child(odd) td { background:rgba(255,255,255,0.8); }
    .sal-table tr:not(.total-row):hover td { background:rgba(68,186,188,0.1) !important; transform:scale(1.002); }
    .sal-table .total-row td { background:linear-gradient(135deg,#0D6F6D,#117F7C,#1A9E9B) !important; color:#fff !important; font-weight:800; font-size:0.76rem; border:none; padding:8px 8px; }
    .sal-table .total-row td:first-child { border-left:4px solid #FFD700; text-align:left; padding-left:12px; }
    .sal-section-title { color:#fff; font-size:0.95rem; font-weight:800; margin:16px 0 8px 0; padding:8px 18px; background:linear-gradient(135deg,#0D6F6D,#117F7C,#44BABC); border-radius:12px; display:inline-block; box-shadow:0 4px 14px rgba(17,127,124,0.2); letter-spacing:0.2px; }
    .sal-note { color:#64748B; font-size:0.72rem; margin-top:2px; margin-bottom:10px; padding:8px 14px; background:rgba(240,250,250,0.5); border-radius:10px; border-left:3px solid #44BABC; }
    .sal-divider { height:2px; background:linear-gradient(90deg,#117F7C,#44BABC,#61C1C3,transparent); margin:12px 0; border-radius:2px; }
    </style>""", unsafe_allow_html=True)

    def fmt(v):
        if isinstance(v, (int, float)):
            return f"{v:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")
        return str(v)

    niv_headers = '<th>Concepto</th><th>Niv.4</th><th>Niv.3</th><th>Niv.2</th><th>Niv.1</th><th>Niv.1A</th><th>Niv.1B</th><th>Niv.1C</th>'

    # ---- TABLA 1: RETRIBUCIÓN FIJA TCP ----
    # Valores: Anual del PDF / 12 = Mensual
    st.markdown('<div class="sal-section-title">💰 1. Retribución Fija Mensual — TCP</div>', unsafe_allow_html=True)
    tcp_data = [
        ('Salario Base',         522.11, 578.83, 672.16, 974.70, 1115.50, 1222.52, 1289.38),
        ('Prorrateo Paga Extra',  87.02,  96.47, 112.03, 162.45,  185.92,  203.75,  214.90),
        ('Plus Transporte',      132.08, 132.08, 132.08, 132.08,  134.15,  134.15,  134.15),
    ]
    tcp_totals = [sum(r[i] for r in tcp_data) for i in range(1, 8)]
    html = f'<table class="sal-table"><tr>{niv_headers}</tr>'
    for row in tcp_data:
        html += '<tr>' + f'<td>{row[0]}</td>' + ''.join(f'<td>{fmt(row[i])}</td>' for i in range(1,8)) + '</tr>'
    html += '<tr class="total-row"><td>✅ TOTAL MENSUAL</td>' + ''.join(f'<td>{fmt(t)}</td>' for t in tcp_totals) + '</tr></table>'
    st.markdown(html, unsafe_allow_html=True)

    # ---- TABLA 2: RETRIBUCIÓN FIJA SOBRECARGO ----
    st.markdown('<div class="sal-divider"></div>', unsafe_allow_html=True)
    st.markdown('<div class="sal-section-title">👑 2. Retribución Fija Mensual — Sobrecargo</div>', unsafe_allow_html=True)
    sc_data = [
        ('Salario Base',         522.11, 578.83, 672.16, 974.70, 1115.50, 1222.52, 1289.38),
        ('Prorrateo Paga Extra',  87.02,  96.47, 112.03, 162.45,  185.92,  203.75,  214.90),
        ('Plus Transporte',      132.08, 132.08, 132.08, 132.08,  134.15,  134.15,  134.15),
        ('Plus Sobrecargo',      146.39, 146.39, 211.33, 332.49,  337.43,  337.43,  337.43),
    ]
    sc_totals = [sum(r[i] for r in sc_data) for i in range(1, 8)]
    html = f'<table class="sal-table"><tr>{niv_headers}</tr>'
    for row in sc_data:
        html += '<tr>' + f'<td>{row[0]}</td>' + ''.join(f'<td>{fmt(row[i])}</td>' for i in range(1,8)) + '</tr>'
    html += '<tr class="total-row"><td>✅ TOTAL MENSUAL</td>' + ''.join(f'<td>{fmt(t)}</td>' for t in sc_totals) + '</tr></table>'
    st.markdown(html, unsafe_allow_html=True)

    # ---- TABLA 3: VARIABLES Y PLUSES ----
    st.markdown('<div class="sal-divider"></div>', unsafe_allow_html=True)
    st.markdown('<div class="sal-section-title">📊 3. Variables y Pluses (€/unidad)</div>', unsafe_allow_html=True)
    # Valores unitarios EXACTOS del PDF por nivel
    var_data = [
        ('Hora de Vuelo',              6.35,  6.88, 11.12, 11.12, 11.65, 11.65, 12.18),
        ('Imaginaria',                28.14, 28.14, 38.21, 38.31, 38.50, 38.50, 38.50),
        ('Imaginaria en Bloque',      38.00, 38.00, 51.59, 51.72, 51.98, 51.98, 51.98),
        ('Saltos de Inspección',      12.68, 12.68, 12.74, 12.77, 12.83, 12.83, 12.83),
        ('Instrucción en Tierra',    135.32,135.32,135.32,135.32,135.32,135.32,135.32),
        ('Actividad en Tierra',       60.57, 60.57, 97.84, 97.84,102.50,102.50,107.16),
        ('Flexiworking JC',          169.15,169.15,169.15,169.15,169.15,169.15,169.15),
        ('Flexiworking TCP',         112.77,112.77,112.77,112.77,112.77,112.77,112.77),
        ('Plus Nocturnidad',          11.28, 11.28, 11.28, 11.28, 11.28, 11.28, 11.28),
        ('Hora Posicional',            4.24,  4.24,  4.24,  4.24,  4.24,  4.24,  4.24),
        ('Forzoso Adicional',         56.38, 56.38, 56.38, 56.38, 56.38, 56.38, 56.38),
        ('Hora E-Learning',            4.76,  4.76,  4.76,  4.76,  4.76,  4.76,  4.76),
        ('Plus Festivo',              44.47, 44.47, 44.47, 44.47, 44.47, 44.47, 44.47),
        ('Invasión OFF (0:00-0:59)',  29.21, 29.21, 29.21, 29.21, 29.21, 29.21, 29.21),
        ('Invasión OFF (1:00-1:59)',  52.92, 52.92, 52.92, 52.92, 52.92, 52.92, 52.92),
        ('Invasión OFF (≥2:00)',      72.16, 72.16, 72.16, 72.16, 72.16, 72.16, 72.16),
        ('Compra del Día Libre JC',  211.77,211.77,211.77,211.77,211.77,211.77,211.77),
        ('Compra del Día Libre TCP', 169.42,169.42,169.42,169.42,169.42,169.42,169.42),
        ('Cambio Servicio Prog. (≥4)',15.88, 15.88, 15.88, 15.88, 15.88, 15.88, 15.88),
        ('Plus Extensión Actividad',  15.88, 15.88, 15.88, 15.88, 15.88, 15.88, 15.88),
        ('Días Sindicales',          108.95,108.95,108.95,108.95,108.95,108.95,108.95),
    ]
    html = f'<table class="sal-table"><tr>{niv_headers}</tr>'
    for row in var_data:
        html += '<tr>' + f'<td>{row[0]}</td>' + ''.join(f'<td>{fmt(row[i])}</td>' for i in range(1,8)) + '</tr>'
    html += '</table>'
    st.markdown(html, unsafe_allow_html=True)

    # ---- TABLA 4: COMPENSACIONES Y DIETAS ----
    st.markdown('<div class="sal-divider"></div>', unsafe_allow_html=True)
    st.markdown('<div class="sal-section-title">🍽️ 4. Compensaciones y Dietas (€)</div>', unsafe_allow_html=True)
    comp_data = [
        ('Venta Patrón (5-3/5-3→kP)',    190.60,190.60,190.60,190.60,190.60,190.60,190.60),
        ('Venta Patrón (5-3-5-4→kP)',    317.66,317.66,317.66,317.66,317.66,317.66,317.66),
        ('Comp. Pérdida Patrón 5-3/5-3',  84.71, 84.71, 84.71, 84.71, 84.71, 84.71, 84.71),
        ('Comp. Pérdida Patrón 5-3/5-4', 169.42,169.42,169.42,169.42,169.42,169.42,169.42),
        ('Dieta de Vuelo',                33.88, 33.88, 33.88, 33.88, 33.88, 33.88, 33.88),
        ('Dieta de Formación',            38.12, 38.12, 38.12, 38.12, 38.12, 38.12, 38.12),
        ('Dieta Formación Sujeta',        38.12, 38.12, 38.12, 38.12, 38.12, 38.12, 38.12),
        ('Dieta Form. Pernocta Nac.',     62.12, 62.12, 62.12, 62.12, 62.12, 62.12, 62.12),
        ('Dieta Form. Pernocta Inter.',   74.06, 74.06, 74.06, 74.06, 74.06, 74.06, 74.06),
        ('Dieta Pernocta Nacional',       70.03, 70.03, 70.03, 70.03, 70.03, 70.03, 70.03),
        ('Dieta Pernocta Internacional',  82.01, 82.01, 82.01, 82.01, 82.01, 82.01, 82.01),
        ('Destacamento',                  24.67, 24.67, 24.67, 24.67, 24.67, 24.67, 24.67),
        ('Pernocta Nacional Meeting',     24.19, 24.19, 24.19, 24.19, 24.19, 24.19, 24.19),
    ]
    html = f'<table class="sal-table"><tr>{niv_headers}</tr>'
    for row in comp_data:
        html += '<tr>' + f'<td>{row[0]}</td>' + ''.join(f'<td>{fmt(row[i])}</td>' for i in range(1,8)) + '</tr>'
    html += '</table>'
    st.markdown(html, unsafe_allow_html=True)

    # ---- TABLA 5: VACACIONES ----
    st.markdown('<div class="sal-divider"></div>', unsafe_allow_html=True)
    st.markdown('<div class="sal-section-title">🏖️ 5. Compensación por Vacaciones (€/día)</div>', unsafe_allow_html=True)
    vac_data = [('Vacaciones (diario)', 22.55, 22.55, 28.19, 33.83, 39.47, 39.47, 39.47)]
    html = f'<table class="sal-table"><tr>{niv_headers}</tr>'
    for row in vac_data:
        html += '<tr>' + f'<td>{row[0]}</td>' + ''.join(f'<td>{fmt(row[i])}</td>' for i in range(1,8)) + '</tr>'
    html += '</table>'
    st.markdown(html, unsafe_allow_html=True)

    st.markdown('<p class="sal-note">ℹ️ El % de retención IRPF y las cotizaciones dependen de la situación personal de cada trabajador. Las deducciones estándar son: Régimen General + MEI (4,83%) y D+F+P (1,65%) = Total 6,48%.</p>', unsafe_allow_html=True)

    # --- DESCARGAS ---
    st.markdown("<div style='margin-top:12px'></div>", unsafe_allow_html=True)
    dl_col1, dl_col2 = st.columns(2)

    with dl_col1:
        def generate_salary_excel():
            output = io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Tablas Salariales'
            hdr_fill = PatternFill(start_color='0B847F', end_color='0B847F', fill_type='solid')
            hdr_font = Font(bold=True, color='FFFFFF', size=10)
            tot_fill = PatternFill(start_color='0B847F', end_color='0B847F', fill_type='solid')
            tot_font = Font(bold=True, color='FFFFFF', size=10)
            sect_fill = PatternFill(start_color='E6F7F7', end_color='E6F7F7', fill_type='solid')
            sect_font = Font(bold=True, color='0B847F', size=11)
            brd = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                         top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
            num_fmt = '#,##0.00'
            headers_niv = ['Concepto','Nivel 4','Nivel 3','Nivel 2','Nivel 1','Nivel 1A','Nivel 1B','Nivel 1C']

            # Title
            ws.cell(row=1, column=1, value='Tablas Salariales 2026').font = Font(bold=True, size=14, color='0B847F')
            ws.merge_cells('A1:H1')
            ws.cell(row=2, column=1, value='Retribuciones vigentes · Todos los niveles').font = Font(size=9, color='64748B')
            ws.merge_cells('A2:H2')

            row = 4

            def write_section(title, data_rows, totals=None):
                nonlocal row
                ws.cell(row=row, column=1, value=title).font = sect_font
                ws.cell(row=row, column=1).fill = sect_fill
                for ci in range(1, 9):
                    ws.cell(row=row, column=ci).fill = sect_fill
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
                row += 1
                for ci, h in enumerate(headers_niv, 1):
                    c = ws.cell(row=row, column=ci, value=h)
                    c.fill = hdr_fill; c.font = hdr_font; c.border = brd
                    c.alignment = Alignment(horizontal='center' if ci>1 else 'left')
                row += 1
                for row_data in data_rows:
                    for ci, val in enumerate(row_data, 1):
                        c = ws.cell(row=row, column=ci, value=val)
                        c.border = brd
                        if ci > 1 and isinstance(val, (int, float)):
                            c.number_format = num_fmt
                            c.alignment = Alignment(horizontal='center')
                    row += 1
                if totals:
                    ws.cell(row=row, column=1, value='TOTAL').fill = tot_fill
                    ws.cell(row=row, column=1).font = tot_font
                    ws.cell(row=row, column=1).border = brd
                    for ci, val in enumerate(totals, 2):
                        c = ws.cell(row=row, column=ci, value=val)
                        c.fill = tot_fill; c.font = tot_font; c.border = brd
                        c.number_format = num_fmt; c.alignment = Alignment(horizontal='center')
                    row += 1
                row += 1  # blank separator

            write_section('1. Retribución Fija Mensual — TCP', tcp_data, tcp_totals)
            write_section('2. Retribución Fija Mensual — Sobrecargo', sc_data, sc_totals)
            write_section('3. Variables y Pluses (€/unidad)', var_data)
            write_section('4. Compensaciones y Dietas (€)', comp_data)
            write_section('5. Compensación por Vacaciones (€/día)', vac_data)

            # Note
            ws.cell(row=row, column=1, value='Deducciones estándar: Régimen General + MEI (4,83%) + D+F+P (1,65%) = Total 6,48%. El IRPF depende de la situación personal.').font = Font(size=8, color='64748B', italic=True)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)

            for col_cells in ws.columns:
                try: letter = col_cells[0].column_letter
                except AttributeError: continue
                mx = max(len(str(c.value or '')) for c in col_cells if hasattr(c, 'column_letter')) + 3
                ws.column_dimensions[letter].width = min(mx, 35)
            wb.save(output)
            return output.getvalue()

        st.download_button("📥 Descargar Excel", data=generate_salary_excel(),
                           file_name="Tablas_Salariales_2026.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)

    with dl_col2:
        def generate_salary_pdf():
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.units import mm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER
            from reportlab.pdfgen import canvas as rl_canvas
            import os

            buf = io.BytesIO()
            page_w, page_h = landscape(A4)

            # Watermark path
            watermark_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'watermark_tablas.png')

            def draw_watermark(canvas, doc):
                """Draw watermark on every page"""
                canvas.saveState()
                if os.path.exists(watermark_path):
                    canvas.setFillAlpha(0.06)
                    # Center watermark on page
                    wm_w = page_w * 0.6
                    wm_h = wm_w * (2000 / 1414)  # maintain aspect ratio
                    x = (page_w - wm_w) / 2
                    y = (page_h - wm_h) / 2
                    canvas.drawImage(watermark_path, x, y, width=wm_w, height=wm_h, mask='auto', preserveAspectRatio=True)
                # Footer
                canvas.setFillAlpha(1.0)
                canvas.setFont('Helvetica', 6)
                canvas.setFillColor(colors.HexColor('#94A3B8'))
                canvas.drawCentredString(page_w / 2, 8*mm, "Herramientas Sindicales © 2026")
                canvas.restoreState()

            doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=8*mm, rightMargin=8*mm, topMargin=8*mm, bottomMargin=12*mm)
            els = []
            styles = getSampleStyleSheet()
            teal = colors.HexColor('#0B847F'); lt = colors.HexColor('#f1fafa'); wh = colors.white
            gold = colors.HexColor('#FFD100')
            title_s = ParagraphStyle('TS', parent=styles['Title'], fontSize=13, textColor=teal, spaceAfter=2, alignment=TA_CENTER)
            sub_s = ParagraphStyle('SS', parent=styles['Normal'], fontSize=7, textColor=colors.HexColor('#64748B'), spaceAfter=3, alignment=TA_CENTER)
            sect_s = ParagraphStyle('SE', parent=styles['Normal'], fontSize=7, textColor=teal, spaceBefore=3, spaceAfter=0.5, fontName='Helvetica-Bold')

            els.append(Paragraph("Tablas Salariales 2026", title_s))
            els.append(Paragraph("Retribuciones vigentes · Todos los niveles", sub_s))

            avail_w = page_w - 16*mm
            concept_w = avail_w * 0.24
            niv_w = (avail_w - concept_w) / 7

            # Filter out "Días Sindicales" from var_data for PDF
            var_data_pdf = [r for r in var_data if 'Sindicales' not in r[0]]

            def add_pdf_table(title, rows, totals_row=None, emoji=""):
                els.append(Paragraph(f"{emoji} {title}", sect_s))
                hdr = ['Concepto','Niv.4','Niv.3','Niv.2','Niv.1','Niv.1A','Niv.1B','Niv.1C']
                tdata = [hdr]
                for r in rows:
                    tdata.append([r[0]] + [f"{r[i]:,.2f}" if isinstance(r[i], (int,float)) else str(r[i]) for i in range(1,8)])
                if totals_row:
                    tdata.append(['TOTAL'] + [f"{t:,.2f}" for t in totals_row])
                col_ws = [concept_w] + [niv_w]*7
                t = Table(tdata, repeatRows=1, colWidths=col_ws)
                sty = [
                    ('BACKGROUND', (0,0), (-1,0), teal), ('TEXTCOLOR', (0,0), (-1,0), wh),
                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,0), 5),
                    ('FONTSIZE', (0,1), (-1,-1), 5), ('ALIGN', (1,0), (-1,-1), 'CENTER'),
                    ('ALIGN', (0,0), (0,-1), 'LEFT'),
                    ('ROWBACKGROUNDS', (0,1), (-1,-1), [wh, lt]),
                    ('GRID', (0,0), (-1,-1), 0.25, colors.HexColor('#CCCCCC')),
                    ('TOPPADDING', (0,0), (-1,-1), 1), ('BOTTOMPADDING', (0,0), (-1,-1), 1),
                    ('LEFTPADDING', (0,0), (-1,-1), 2.5), ('RIGHTPADDING', (0,0), (-1,-1), 2.5),
                ]
                if totals_row:
                    lr = len(tdata)-1
                    sty += [('BACKGROUND', (0,lr), (-1,lr), teal), ('TEXTCOLOR', (0,lr), (-1,lr), wh),
                            ('FONTNAME', (0,lr), (-1,lr), 'Helvetica-Bold')]
                t.setStyle(TableStyle(sty))
                els.append(t)
                els.append(Spacer(1, 1*mm))

            add_pdf_table("1. Retribucion Fija Mensual - TCP", tcp_data, tcp_totals, "")
            add_pdf_table("2. Retribucion Fija Mensual - Sobrecargo", sc_data, sc_totals, "")
            add_pdf_table("3. Variables y Pluses (EUR/unidad)", var_data_pdf, emoji="")
            add_pdf_table("4. Compensaciones y Dietas (EUR)", comp_data, emoji="")
            add_pdf_table("5. Compensacion por Vacaciones (EUR/dia)", vac_data, emoji="")

            els.append(Spacer(1, 1.5*mm))
            note_s = ParagraphStyle('NT', parent=styles['Normal'], fontSize=5, textColor=colors.HexColor('#64748B'), alignment=TA_CENTER)
            els.append(Paragraph("Deducciones estandar: Regimen General + MEI (4,83%) + D+F+P (1,65%) = Total 6,48%. El IRPF depende de la situacion personal.", note_s))
            doc.build(els, onFirstPage=draw_watermark, onLaterPages=draw_watermark)
            return buf.getvalue()

        st.download_button("📥 Descargar PDF", data=generate_salary_pdf(),
                           file_name="Tablas_Salariales_2026.pdf", mime="application/pdf",
                           use_container_width=True)

    st.markdown("""
    <div style="text-align:center; margin-top:16px; padding:12px 0 4px 0;">
        <div style="width:80px;height:2px;background:linear-gradient(90deg,transparent,#44BABC,transparent);margin:0 auto 10px auto;border-radius:1px;"></div>
        <span style="color:#94A3B8; font-size:0.78rem; font-weight:500;">✈️ Herramientas Sindicales © 2026</span>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# =====================================================================
# SECCIÓN: CALCULADORA DE NÓMINAS (ZIP-EXACT REDESIGN)
# =====================================================================
if st.session_state['app_section'] == 'calculadora_nominas':
    render_admin_sidebar()

    # --- Top bar: Volver + Cerrar Sesión ---
    _usr_col1, _usr_col2, _usr_col3 = st.columns([1, 4, 1])
    with _usr_col1:
        if st.button("⬅️ Volver", key="volver_calc"):
            st.session_state['app_section'] = 'seleccion'
            st.rerun()
    with _usr_col3:
        if st.button("🚪 Cerrar Sesión", key="logout_calc"):
            for key in ['authenticated', 'username', 'is_admin']:
                st.session_state[key] = False if key == 'authenticated' else None
            st.session_state['app_section'] = 'seleccion'
            st.rerun()

    # ========== CSS ==========
    st.markdown("""<style>
    .stApp { background: linear-gradient(135deg, #f8fafc 0%, rgba(204,251,241,0.15) 50%, rgba(207,250,254,0.15) 100%) !important; }
    .calc-header {
        background: linear-gradient(to right, #0f766e, #0d9488, #0891b2);
        color: white; padding: 20px 24px; border-radius: 0;
        display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap;
        margin: -1rem -1rem 20px -1rem; box-shadow: 0 1px 3px rgba(0,0,0,0.1);
    }
    .calc-header-left { display: flex; align-items: center; gap: 12px; }
    .calc-header-icon {
        background: rgba(255,255,255,0.2); border-radius: 12px; padding: 10px;
        backdrop-filter: blur(8px); display: flex; align-items: center; justify-content: center;
    }
    .calc-header h1 { font-size: 1.5rem; font-weight: 700; margin: 0; letter-spacing: -0.5px; line-height: 1.2; }
    .calc-header p { font-size: 0.875rem; color: rgba(204,251,241,1); margin: 2px 0 0 0; }
    .calc-card {
        background: white; border-radius: 12px; padding: 16px 16px 4px 16px; margin-bottom: 16px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.06); border: none;
    }
    .calc-card-title {
        font-size: 1rem; font-weight: 600; color: #1e293b; margin-bottom: 12px;
        display: flex; align-items: center; gap: 8px; padding-bottom: 12px; border-bottom: 1px solid #f1f5f9;
    }
    div[data-testid="stNumberInput"] > div { max-width: 100% !important; min-width: 0 !important; }
    div[data-testid="stNumberInput"] input {
        padding: 8px 12px !important; font-size: 0.875rem !important; height: 40px !important;
        border-radius: 8px !important; border: 1px solid #e2e8f0 !important;
        text-align: right; font-weight: 500;
    }
    div[data-testid="stNumberInput"] input:focus {
        border-color: #14b8a6 !important; box-shadow: 0 0 0 3px rgba(20,184,166,0.1) !important;
    }
    div[data-testid="stNumberInput"] button {
        height: 40px !important; width: 36px !important; min-width: 36px !important; padding: 0 !important;
        border-radius: 6px !important;
    }
    div[data-testid="stNumberInput"] label {
        font-size: 0.85rem !important; font-weight: 600 !important; color: #475569 !important;
    }
    .fijos-banner {
        background: linear-gradient(to right, #0d9488, #0891b2); border-radius: 12px;
        padding: 20px; margin-bottom: 16px; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1); color: white;
    }
    .fijos-banner-top { display: flex; justify-content: space-between; align-items: flex-start; flex-wrap: wrap; gap: 8px; }
    .fijos-banner h3 { font-size: 1.125rem; font-weight: 700; margin: 0; }
    .fijos-banner .sub { font-size: 0.875rem; color: rgba(204,251,241,1); margin-top: 2px; }
    .fijos-banner .total { font-size: 1.875rem; font-weight: 700; letter-spacing: -0.5px; }
    .fijos-banner .total-sub { font-size: 0.75rem; color: rgba(204,251,241,1); text-align: right; }
    .fijos-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 12px; margin-top: 16px; }
    .fijos-item { background: rgba(255,255,255,0.15); border-radius: 8px; padding: 12px; backdrop-filter: blur(8px); }
    .fijos-item .label { font-size: 0.75rem; color: rgba(204,251,241,1); }
    .fijos-item .value { font-size: 0.875rem; font-weight: 700; margin-top: 4px; }
    .kpi-card-outline {
        border-radius: 12px; padding: 16px; margin-bottom: 12px; border: 1px solid #e2e8f0;
        background: white; transition: all 0.2s;
    }
    .kpi-card-outline:hover { box-shadow: 0 4px 12px rgba(0,0,0,0.08); }
    .kpi-card-outline .kpi-top { display: flex; justify-content: space-between; align-items: flex-start; }
    .kpi-card-outline .kpi-label { font-size: 0.75rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.05em; color: #64748b; }
    .kpi-card-outline .kpi-value { font-size: 1.5rem; font-weight: 700; color: #0f172a; margin-top: 4px; letter-spacing: -0.5px; }
    .kpi-card-outline .kpi-sub { font-size: 0.75rem; color: #94a3b8; margin-top: 2px; }
    .kpi-card-outline .kpi-icon { background: #f1f5f9; border-radius: 8px; padding: 8px; display: flex; }
    .kpi-card-outline .kpi-icon svg { width: 16px; height: 16px; stroke: #0d9488; fill: none; stroke-width: 2; }
    .kpi-card-primary {
        border-radius: 12px; padding: 16px; margin-bottom: 12px;
        background: linear-gradient(135deg, #0d9488, #0891b2); color: white;
    }
    .kpi-card-primary:hover { box-shadow: 0 4px 12px rgba(13,148,136,0.3); }
    .kpi-card-primary .kpi-top { display: flex; justify-content: space-between; align-items: flex-start; }
    .kpi-card-primary .kpi-label { font-size: 0.75rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.05em; color: rgba(255,255,255,0.8); }
    .kpi-card-primary .kpi-value { font-size: 1.5rem; font-weight: 700; color: white; margin-top: 4px; }
    .kpi-card-primary .kpi-icon { background: rgba(255,255,255,0.2); border-radius: 8px; padding: 8px; display: flex; }
    .kpi-card-primary .kpi-icon svg { width: 16px; height: 16px; stroke: white; fill: none; stroke-width: 2; }
    .kpi-card-warning {
        border-radius: 12px; padding: 16px; margin-bottom: 12px;
        background: linear-gradient(135deg, #f59e0b, #ea580c); color: white;
    }
    .kpi-card-warning .kpi-label { font-size: 0.75rem; font-weight: 600; text-transform: uppercase; color: rgba(255,255,255,0.8); }
    .kpi-card-warning .kpi-value { font-size: 1.5rem; font-weight: 700; color: white; margin-top: 4px; }
    .kpi-deduction {
        border-radius: 12px; padding: 16px; margin-bottom: 12px;
        background: #fef2f2; border: 1px solid #fecaca;
    }
    .kpi-deduction .kpi-label { font-size: 0.75rem; font-weight: 600; text-transform: uppercase; color: #dc2626; }
    .kpi-deduction .kpi-value { font-size: 1.5rem; font-weight: 700; color: #b91c1c; margin-top: 4px; }
    .kpi-card-success {
        border-radius: 12px; padding: 16px; margin-bottom: 12px;
        background: linear-gradient(135deg, #059669, #0d9488); color: white;
    }
    .kpi-card-success .kpi-top { display: flex; justify-content: space-between; align-items: flex-start; }
    .kpi-card-success .kpi-label { font-size: 0.75rem; font-weight: 600; text-transform: uppercase; color: rgba(255,255,255,0.8); }
    .kpi-card-success .kpi-value { font-size: 1.5rem; font-weight: 700; color: white; margin-top: 4px; }
    .kpi-card-success .kpi-icon { background: rgba(255,255,255,0.2); border-radius: 8px; padding: 8px; display: flex; }
    .kpi-card-success .kpi-icon svg { width: 16px; height: 16px; stroke: white; fill: none; stroke-width: 2; }
    .desglose-card {
        background: white; border-radius: 12px; box-shadow: 0 1px 3px rgba(0,0,0,0.06);
        border: none; overflow: hidden; margin-bottom: 12px;
    }
    .desglose-header { padding: 12px 16px; font-size: 1rem; font-weight: 600; border-bottom: 1px solid #f1f5f9; }
    .desglose-cols {
        display: grid; grid-template-columns: 1fr 70px 50px 80px; gap: 8px;
        padding: 8px 12px; font-size: 0.75rem; font-weight: 600; color: #64748b;
        border-bottom: 1px solid #e2e8f0;
    }
    .desglose-cols span:not(:first-child) { text-align: right; }
    .desglose-body { max-height: 400px; overflow-y: auto; }
    .desglose-row {
        display: grid; grid-template-columns: 1fr 70px 50px 80px; gap: 8px;
        padding: 8px 12px; font-size: 0.8rem; border-bottom: 1px solid #f8fafc;
    }
    .desglose-row span:first-child { font-weight: 500; }
    .desglose-row span:not(:first-child) { text-align: right; color: #64748b; }
    .desglose-row span:last-child { color: #0f172a; font-weight: 600; }
    .desglose-row.highlight {
        background: linear-gradient(135deg, #0d9488, #0891b2); color: white !important;
        font-weight: 600; border-radius: 8px; margin: 4px 8px;
    }
    .desglose-row.highlight span { color: white !important; }
    .calc-footer {
        border-top: 1px solid #e2e8f0; background: white; padding: 16px; text-align: center;
        margin: 20px -1rem -1rem -1rem;
    }
    .calc-footer p { color: #64748b; font-size: 0.875rem; margin: 0; }
    </style>""", unsafe_allow_html=True)

    # ========== HEADER ==========
    st.markdown('''
    <div class="calc-header">
        <div class="calc-header-left">
            <div class="calc-header-icon">
                <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" style="width:24px;height:24px;stroke:white;fill:none;stroke-width:2;stroke-linecap:round;stroke-linejoin:round;">
                    <rect width="16" height="20" x="4" y="2" rx="2"/><line x1="8" x2="16" y1="6" y2="6"/>
                    <line x1="16" x2="16" y1="14" y2="18"/><path d="M16 10h.01"/><path d="M12 10h.01"/>
                    <path d="M8 10h.01"/><path d="M12 14h.01"/><path d="M8 14h.01"/>
                    <path d="M12 18h.01"/><path d="M8 18h.01"/>
                </svg>
            </div>
            <div>
                <h1>Calculadora de Nóminas</h1>
                <p>Tablas Salariales 2026</p>
            </div>
        </div>
    </div>
    ''', unsafe_allow_html=True)

    # ========== DATA ==========
    CALC_DATA = {
        '4': {'salario_base':522.11,'prorrateo':87.02,'plus_transporte':132.08,'plus_sobrecargo':146.39,
              'hora_vuelo':6.35,'imaginaria':28.14,'dia_bloque_imag':38.00,'saltos_inspeccion':12.68,
              'instruccion_tierra':135.32,'actividad_tierra':60.57,'flexiworking_jc':169.15,'flexiworking_tcp':112.77,
              'plus_nocturnidad':11.28,'hora_posicional':4.24,'forzoso_adicional':56.38,'hora_elearning':4.76,
              'festivo':44.47,'invasion_0':29.21,'invasion_1':52.92,'invasion_2':72.16,
              'compra_libre_jc':211.77,'compra_libre_tcp':169.42,'cambio_servicio':15.88,'actividad_extendida':15.88,
              'dias_sindicales':108.95,
              'venta_5353':190.60,'venta_5453':317.66,'comp_5353':84.71,'comp_5453':169.42,
              'dieta_vuelo':33.88,'dieta_formacion':38.12,'dieta_form_sujeta':38.12,'dieta_form_nac':62.12,'dieta_form_int':74.06,
              'dieta_pernocta_nac':70.03,'dieta_pernocta_int':82.01,'destacamento':24.67,'pernocta_meeting':24.19,
              'vacaciones':22.55},
        '3': {'salario_base':578.83,'prorrateo':96.47,'plus_transporte':132.08,'plus_sobrecargo':146.39,
              'hora_vuelo':6.88,'imaginaria':28.14,'dia_bloque_imag':38.00,'saltos_inspeccion':12.68,
              'instruccion_tierra':135.32,'actividad_tierra':60.57,'flexiworking_jc':169.15,'flexiworking_tcp':112.77,
              'plus_nocturnidad':11.28,'hora_posicional':4.24,'forzoso_adicional':56.38,'hora_elearning':4.76,
              'festivo':44.47,'invasion_0':29.21,'invasion_1':52.92,'invasion_2':72.16,
              'compra_libre_jc':211.77,'compra_libre_tcp':169.42,'cambio_servicio':15.88,'actividad_extendida':15.88,
              'dias_sindicales':108.95,
              'venta_5353':190.60,'venta_5453':317.66,'comp_5353':84.71,'comp_5453':169.42,
              'dieta_vuelo':33.88,'dieta_formacion':38.12,'dieta_form_sujeta':38.12,'dieta_form_nac':62.12,'dieta_form_int':74.06,
              'dieta_pernocta_nac':70.03,'dieta_pernocta_int':82.01,'destacamento':24.67,'pernocta_meeting':24.19,
              'vacaciones':22.55},
        '2': {'salario_base':672.16,'prorrateo':112.03,'plus_transporte':132.08,'plus_sobrecargo':211.33,
              'hora_vuelo':11.12,'imaginaria':38.21,'dia_bloque_imag':51.59,'saltos_inspeccion':12.74,
              'instruccion_tierra':135.32,'actividad_tierra':97.84,'flexiworking_jc':169.15,'flexiworking_tcp':112.77,
              'plus_nocturnidad':11.28,'hora_posicional':4.24,'forzoso_adicional':56.38,'hora_elearning':4.76,
              'festivo':44.47,'invasion_0':29.21,'invasion_1':52.92,'invasion_2':72.16,
              'compra_libre_jc':211.77,'compra_libre_tcp':169.42,'cambio_servicio':15.88,'actividad_extendida':15.88,
              'dias_sindicales':108.95,
              'venta_5353':190.60,'venta_5453':317.66,'comp_5353':84.71,'comp_5453':169.42,
              'dieta_vuelo':33.88,'dieta_formacion':38.12,'dieta_form_sujeta':38.12,'dieta_form_nac':62.12,'dieta_form_int':74.06,
              'dieta_pernocta_nac':70.03,'dieta_pernocta_int':82.01,'destacamento':24.67,'pernocta_meeting':24.19,
              'vacaciones':28.19},
        '1': {'salario_base':974.70,'prorrateo':162.45,'plus_transporte':132.08,'plus_sobrecargo':332.49,
              'hora_vuelo':11.12,'imaginaria':38.31,'dia_bloque_imag':51.72,'saltos_inspeccion':12.77,
              'instruccion_tierra':135.32,'actividad_tierra':97.84,'flexiworking_jc':169.15,'flexiworking_tcp':112.77,
              'plus_nocturnidad':11.28,'hora_posicional':4.24,'forzoso_adicional':56.38,'hora_elearning':4.76,
              'festivo':44.47,'invasion_0':29.21,'invasion_1':52.92,'invasion_2':72.16,
              'compra_libre_jc':211.77,'compra_libre_tcp':169.42,'cambio_servicio':15.88,'actividad_extendida':15.88,
              'dias_sindicales':108.95,
              'venta_5353':190.60,'venta_5453':317.66,'comp_5353':84.71,'comp_5453':169.42,
              'dieta_vuelo':33.88,'dieta_formacion':38.12,'dieta_form_sujeta':38.12,'dieta_form_nac':62.12,'dieta_form_int':74.06,
              'dieta_pernocta_nac':70.03,'dieta_pernocta_int':82.01,'destacamento':24.67,'pernocta_meeting':24.19,
              'vacaciones':33.83},
        '1A': {'salario_base':1115.50,'prorrateo':185.92,'plus_transporte':134.15,'plus_sobrecargo':337.43,
              'hora_vuelo':11.65,'imaginaria':38.50,'dia_bloque_imag':51.98,'saltos_inspeccion':12.83,
              'instruccion_tierra':135.32,'actividad_tierra':102.50,'flexiworking_jc':169.15,'flexiworking_tcp':112.77,
              'plus_nocturnidad':11.28,'hora_posicional':4.24,'forzoso_adicional':56.38,'hora_elearning':4.76,
              'festivo':44.47,'invasion_0':29.21,'invasion_1':52.92,'invasion_2':72.16,
              'compra_libre_jc':211.77,'compra_libre_tcp':169.42,'cambio_servicio':15.88,'actividad_extendida':15.88,
              'dias_sindicales':108.95,
              'venta_5353':190.60,'venta_5453':317.66,'comp_5353':84.71,'comp_5453':169.42,
              'dieta_vuelo':33.88,'dieta_formacion':38.12,'dieta_form_sujeta':38.12,'dieta_form_nac':62.12,'dieta_form_int':74.06,
              'dieta_pernocta_nac':70.03,'dieta_pernocta_int':82.01,'destacamento':24.67,'pernocta_meeting':24.19,
              'vacaciones':39.47},
        '1B': {'salario_base':1222.52,'prorrateo':203.75,'plus_transporte':134.15,'plus_sobrecargo':337.43,
              'hora_vuelo':11.65,'imaginaria':38.50,'dia_bloque_imag':51.98,'saltos_inspeccion':12.83,
              'instruccion_tierra':135.32,'actividad_tierra':102.50,'flexiworking_jc':169.15,'flexiworking_tcp':112.77,
              'plus_nocturnidad':11.28,'hora_posicional':4.24,'forzoso_adicional':56.38,'hora_elearning':4.76,
              'festivo':44.47,'invasion_0':29.21,'invasion_1':52.92,'invasion_2':72.16,
              'compra_libre_jc':211.77,'compra_libre_tcp':169.42,'cambio_servicio':15.88,'actividad_extendida':15.88,
              'dias_sindicales':108.95,
              'venta_5353':190.60,'venta_5453':317.66,'comp_5353':84.71,'comp_5453':169.42,
              'dieta_vuelo':33.88,'dieta_formacion':38.12,'dieta_form_sujeta':38.12,'dieta_form_nac':62.12,'dieta_form_int':74.06,
              'dieta_pernocta_nac':70.03,'dieta_pernocta_int':82.01,'destacamento':24.67,'pernocta_meeting':24.19,
              'vacaciones':39.47},
        '1C': {'salario_base':1289.38,'prorrateo':214.90,'plus_transporte':134.15,'plus_sobrecargo':337.43,
              'hora_vuelo':12.18,'imaginaria':38.50,'dia_bloque_imag':51.98,'saltos_inspeccion':12.83,
              'instruccion_tierra':135.32,'actividad_tierra':107.16,'flexiworking_jc':169.15,'flexiworking_tcp':112.77,
              'plus_nocturnidad':11.28,'hora_posicional':4.24,'forzoso_adicional':56.38,'hora_elearning':4.76,
              'festivo':44.47,'invasion_0':29.21,'invasion_1':52.92,'invasion_2':72.16,
              'compra_libre_jc':211.77,'compra_libre_tcp':169.42,'cambio_servicio':15.88,'actividad_extendida':15.88,
              'dias_sindicales':108.95,
              'venta_5353':190.60,'venta_5453':317.66,'comp_5353':84.71,'comp_5453':169.42,
              'dieta_vuelo':33.88,'dieta_formacion':38.12,'dieta_form_sujeta':38.12,'dieta_form_nac':62.12,'dieta_form_int':74.06,
              'dieta_pernocta_nac':70.03,'dieta_pernocta_int':82.01,'destacamento':24.67,'pernocta_meeting':24.19,
              'vacaciones':39.47},
    }

    # ========== STATE INIT ==========
    if 'calc_nivel' not in st.session_state:
        st.session_state['calc_nivel'] = '4'
    if 'calc_tipo' not in st.session_state:
        st.session_state['calc_tipo'] = 'TCP'
    if 'calc_reset_counter' not in st.session_state:
        st.session_state['calc_reset_counter'] = 0
    _rc = st.session_state['calc_reset_counter']

    niveles_list = ['4', '3', '2', '1', '1A', '1B', '1C']

    # ========== NIVEL SELECTOR ==========
    _nv_card = '<div class="calc-card" style="padding:16px;">'
    _nv_card += '<div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;">'
    _nv_card += '<span style="font-size:0.875rem;font-weight:600;color:#475569;">Nivel:</span>'
    _nv_card += '</div></div>'
    st.markdown(_nv_card, unsafe_allow_html=True)
    nivel_cols = st.columns(len(niveles_list))
    for i, nv in enumerate(niveles_list):
        with nivel_cols[i]:
            btype = "primary" if nv == st.session_state['calc_nivel'] else "secondary"
            if st.button(nv, key=f"nbtn_{nv}", use_container_width=True, type=btype):
                if st.session_state['calc_nivel'] != nv:
                    st.session_state['calc_nivel'] = nv
                    st.session_state['calc_reset_counter'] = st.session_state.get('calc_reset_counter', 0) + 1
                st.rerun()

    # ========== ROL SELECTOR ==========
    _rol_card = '<div class="calc-card" style="padding:16px;">'
    _rol_card += '<div style="display:flex;align-items:center;gap:12px;">'
    _rol_card += '<span style="font-size:0.875rem;font-weight:600;color:#475569;">Rol:</span>'
    _rol_card += '</div></div>'
    st.markdown(_rol_card, unsafe_allow_html=True)
    tcp_selected = st.session_state.get('calc_tipo', 'TCP') == 'TCP'
    tipo_cols = st.columns([1, 1, 1, 3])
    with tipo_cols[0]:
        if st.button("👥 TCP", key="tipo_tcp_btn", use_container_width=True, type="primary" if tcp_selected else "secondary"):
            st.session_state['calc_tipo'] = 'TCP'
            st.rerun()
    with tipo_cols[1]:
        if st.button("👤 JC", key="tipo_jc_btn", use_container_width=True, type="primary" if not tcp_selected else "secondary"):
            st.session_state['calc_tipo'] = 'Sobrecargo'
            st.rerun()
    with tipo_cols[2]:
        if st.button("🔄 Nuevo Cálculo", key="reset_calc", use_container_width=True, type="secondary"):
            st.session_state['calc_reset_counter'] = st.session_state.get('calc_reset_counter', 0) + 1
            st.rerun()

    nivel = st.session_state['calc_nivel']
    tipo = st.session_state['calc_tipo']
    d = CALC_DATA[nivel]

    # --- CONCEPTOS FIJOS ---
    fijo_salario = d['salario_base']
    fijo_prorrateo = d['prorrateo']
    fijo_transporte = d['plus_transporte']
    fijo_sobrecargo = d['plus_sobrecargo'] if tipo == 'Sobrecargo' else 0.0
    total_fijos = fijo_salario + fijo_prorrateo + fijo_transporte + fijo_sobrecargo

    fijos_rows = [
        ('Salario Base', d['salario_base'], fijo_salario),
        ('Prorrateo Paga Extra', d['prorrateo'], fijo_prorrateo),
        ('Plus Transporte', d['plus_transporte'], fijo_transporte),
    ]
    if tipo == 'Sobrecargo':
        fijos_rows.append(('Plus Sobrecargo', d['plus_sobrecargo'], fijo_sobrecargo))

    def fmtc(v):
        if isinstance(v, (int, float)):
            return f"{v:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")
        return str(v)

    # ========== MAIN LAYOUT: Left (2/3) + Right (1/3) ==========
    calc_left, calc_right = st.columns([2, 1])

    with calc_left:
        # ---- Conceptos Fijos Banner ----
        plus_jc_html = ''
        if tipo == 'Sobrecargo':
            plus_jc_html = f'<div style="background:rgba(255,255,255,0.15);border-radius:8px;padding:12px;backdrop-filter:blur(8px);"><div style="font-size:0.75rem;color:rgba(204,251,241,1);">Plus JC</div><div style="font-size:0.875rem;font-weight:700;margin-top:4px;">{fmtc(d["plus_sobrecargo"])}</div></div>'

        _fijos_html = f'''
        <div class="fijos-banner" style="background:linear-gradient(to right,#0d9488,#0891b2);border-radius:12px;padding:20px;margin-bottom:16px;box-shadow:0 4px 6px -1px rgba(0,0,0,0.1);color:white;">
            <div class="fijos-banner-top" style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:8px;">
                <div>
                    <h3 style="font-size:1.125rem;font-weight:700;margin:0;">💰 Conceptos Fijos · Nivel {nivel}</h3>
                    <div style="font-size:0.875rem;color:rgba(204,251,241,1);margin-top:2px;">Rol: {tipo}</div>
                </div>
                <div style="text-align:right;">
                    <div style="font-size:1.875rem;font-weight:700;letter-spacing:-0.5px;">{fmtc(total_fijos)}</div>
                    <div style="font-size:0.75rem;color:rgba(204,251,241,1);text-align:right;">Total mensual fijo</div>
                </div>
            </div>
            <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:12px;margin-top:16px;">
                <div style="background:rgba(255,255,255,0.15);border-radius:8px;padding:12px;backdrop-filter:blur(8px);"><div style="font-size:0.75rem;color:rgba(204,251,241,1);">Salario Base</div><div style="font-size:0.875rem;font-weight:700;margin-top:4px;">{fmtc(d["salario_base"])}</div></div>
                <div style="background:rgba(255,255,255,0.15);border-radius:8px;padding:12px;backdrop-filter:blur(8px);"><div style="font-size:0.75rem;color:rgba(204,251,241,1);">Prorrateo</div><div style="font-size:0.875rem;font-weight:700;margin-top:4px;">{fmtc(d["prorrateo"])}</div></div>
                <div style="background:rgba(255,255,255,0.15);border-radius:8px;padding:12px;backdrop-filter:blur(8px);"><div style="font-size:0.75rem;color:rgba(204,251,241,1);">Plus Transporte</div><div style="font-size:0.875rem;font-weight:700;margin-top:4px;">{fmtc(d["plus_transporte"])}</div></div>
                {plus_jc_html}
            </div>
        </div>
        '''
        st.html(_fijos_html)

        # ---- CARD: Vuelo ----
        st.html('''<div style="background:white;border-radius:12px;padding:16px 16px 4px 16px;margin-bottom:16px;box-shadow:0 1px 3px rgba(0,0,0,0.06);">
            <div style="font-size:1rem;font-weight:600;color:#1e293b;margin-bottom:12px;display:flex;align-items:center;gap:8px;padding-bottom:12px;border-bottom:1px solid #f1f5f9;">
            <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" style="width:16px;height:16px;stroke:#0d9488;fill:none;stroke-width:2;stroke-linecap:round;stroke-linejoin:round;">
                <path d="M17.8 19.2 16 11l3.5-3.5C21 6 21.5 4 21 3c-1-.5-3 0-4.5 1.5L13 8 4.8 6.2c-.5-.1-.9.1-1.1.5l-.3.5c-.2.5-.1 1 .3 1.3L9 12l-2 3H4l-1 1 3 2 2 3 1-1v-3l3-2 3.5 5.3c.3.4.8.5 1.3.3l.5-.2c.4-.3.6-.7.5-1.2z"/>
            </svg> Vuelo
        </div></div>''')
        v1, v2, v3 = st.columns(3)
        with v1:
            n_horas_vuelo = st.number_input("Horas de vuelo", min_value=0.0, value=0.0, step=0.5, key=f'cv_hv_{_rc}', format="%.1f")
            n_hora_pos = st.number_input("Horas Posic.", min_value=0.0, value=0.0, step=0.5, key=f'cv_hpos_{_rc}')
        with v2:
            n_imaginarias = st.number_input("Imaginarias", min_value=0.0, value=0.0, step=1.0, key=f'cv_imag_{_rc}')
            n_nocturnidad = st.number_input("Nocturnidad", min_value=0.0, value=0.0, step=0.5, key=f'cv_noct_{_rc}')
        with v3:
            n_bloque_imag = st.number_input("Bloque Imag.", min_value=0.0, value=0.0, step=1.0, key=f'cv_blq_{_rc}')
            n_saltos = st.number_input("Saltos Inspec.", min_value=0.0, value=0.0, step=1.0, key=f'cv_salt_{_rc}')

        # ---- CARD: Extras ----
        st.html('<div style="background:white;border-radius:12px;padding:16px 16px 4px 16px;margin-bottom:16px;box-shadow:0 1px 3px rgba(0,0,0,0.06);"><div style="font-size:1rem;font-weight:600;color:#1e293b;margin-bottom:12px;display:flex;align-items:center;gap:8px;padding-bottom:12px;border-bottom:1px solid #f1f5f9;"><span style="color:#d97706;font-size:1.1rem;font-weight:700;">+</span> Extras</div></div>')
        e1, e2, e3 = st.columns(3)
        with e1:
            n_festivo = st.number_input("Festivos", min_value=0.0, value=0.0, step=1.0, key=f'cv_fest_{_rc}')
            n_inv0 = st.number_input("Inv. 0:00-0:59", min_value=0.0, value=0.0, step=1.0, key=f'cv_inv0_{_rc}')
            n_cambio_serv = st.number_input("Cambio Serv.", min_value=0.0, value=0.0, step=1.0, key=f'cv_cs_{_rc}')
        with e2:
            n_forzoso = st.number_input("Forzoso Adic.", min_value=0.0, value=0.0, step=1.0, key=f'cv_forz_{_rc}')
            n_inv1 = st.number_input("Inv. 1:00-1:59", min_value=0.0, value=0.0, step=1.0, key=f'cv_inv1_{_rc}')
            n_act_ext = st.number_input("Act. Extend.", min_value=0.0, value=0.0, step=1.0, key=f'cv_ae_{_rc}')
        with e3:
            n_compra_libre = st.number_input("Compra Libre", min_value=0.0, value=0.0, step=1.0, key=f'cv_cl_{_rc}')
            n_inv2 = st.number_input("Inv. ≥2:00", min_value=0.0, value=0.0, step=1.0, key=f'cv_inv2_{_rc}')
            n_flexiworking = st.number_input("Flexiworking", min_value=0.0, value=0.0, step=1.0, key=f'cv_flex_{_rc}')

        # ---- CARD: Dietas ----
        st.html('<div style="background:white;border-radius:12px;padding:16px 16px 4px 16px;margin-bottom:16px;box-shadow:0 1px 3px rgba(0,0,0,0.06);"><div style="font-size:1rem;font-weight:600;color:#1e293b;margin-bottom:12px;display:flex;align-items:center;gap:8px;padding-bottom:12px;border-bottom:1px solid #f1f5f9;"><span style="color:#ea580c;font-size:1.1rem;font-weight:700;">🍽</span> Dietas</div></div>')
        d1, d2, d3 = st.columns(3)
        with d1:
            n_dieta_vuelo = st.number_input("Dietas Vuelo", min_value=0.0, value=0.0, step=1.0, key=f'cv_dv_{_rc}')
            n_pernocta_nac = st.number_input("Pern. Nac.", min_value=0.0, value=0.0, step=1.0, key=f'cv_pn_{_rc}')
            n_destacamento = st.number_input("Destacamento", min_value=0.0, value=0.0, step=1.0, key=f'cv_dest_{_rc}')
        with d2:
            n_dieta_form = st.number_input("Dietas Form.", min_value=0.0, value=0.0, step=1.0, key=f'cv_df_{_rc}')
            n_pernocta_int = st.number_input("Pern. Int.", min_value=0.0, value=0.0, step=1.0, key=f'cv_pi_{_rc}')
            n_pernocta_meet = st.number_input("Pern. Meeting", min_value=0.0, value=0.0, step=1.0, key=f'cv_pm_{_rc}')
        with d3:
            n_dieta_form_nac = st.number_input("D.Form.Nac.", min_value=0.0, value=0.0, step=1.0, key=f'cv_dfn_{_rc}')
            n_dieta_form_int = st.number_input("D.Form.Int.", min_value=0.0, value=0.0, step=1.0, key=f'cv_dfi_{_rc}')
            n_comision_vab = st.number_input("Comisión VAB (€)", min_value=0.0, value=0.0, step=1.0, key=f'cv_vab_{_rc}')

        # ---- CARD: Otros ----
        st.html('<div style="background:white;border-radius:12px;padding:16px 16px 4px 16px;margin-bottom:16px;box-shadow:0 1px 3px rgba(0,0,0,0.06);"><div style="font-size:1rem;font-weight:600;color:#1e293b;margin-bottom:12px;display:flex;align-items:center;gap:8px;padding-bottom:12px;border-bottom:1px solid #f1f5f9;"><span style="color:#475569;font-size:1.1rem;font-weight:700;">📦</span> Otros</div></div>')
        o1, o2, o3 = st.columns(3)
        with o1:
            n_instruccion = st.number_input("Instr. Tierra", min_value=0.0, value=0.0, step=1.0, key=f'cv_inst_{_rc}')
            n_dias_sindicales = st.number_input("Días Sindic.", min_value=0.0, value=0.0, step=1.0, key=f'cv_dsind_{_rc}')
        with o2:
            n_actividad = st.number_input("Activ. Tierra", min_value=0.0, value=0.0, step=1.0, key=f'cv_act_{_rc}')
            n_vacaciones = st.number_input("Vacaciones", min_value=0.0, value=0.0, step=1.0, key=f'cv_vac_{_rc}')
        with o3:
            n_elearning = st.number_input("E-Learning", min_value=0.0, value=0.0, step=0.5, key=f'cv_elearn_{_rc}')

        # ---- CARD: Comp. Patrón ----
        st.html('<div style="background:white;border-radius:12px;padding:16px 16px 4px 16px;margin-bottom:16px;box-shadow:0 1px 3px rgba(0,0,0,0.06);"><div style="font-size:1rem;font-weight:600;color:#1e293b;margin-bottom:12px;display:flex;align-items:center;gap:8px;padding-bottom:12px;border-bottom:1px solid #f1f5f9;"><span style="color:#7c3aed;font-size:1.1rem;font-weight:700;">🏷</span> Comp. Patrón</div></div>')
        p1, p2, p3, p4 = st.columns(4)
        with p1:
            n_venta_5353 = st.number_input("Venta 5-3-5-3 (%)", min_value=0.0, max_value=100.0, value=0.0, step=1.0, key=f'cv_v53_{_rc}')
        with p2:
            n_venta_5453 = st.number_input("Venta 5-4-5-3 (%)", min_value=0.0, max_value=100.0, value=0.0, step=1.0, key=f'cv_v54_{_rc}')
        with p3:
            n_comp_5353 = st.number_input("Comp. 5-3-5-3 (%)", min_value=0.0, max_value=100.0, value=0.0, step=1.0, key=f'cv_c53_{_rc}')
        with p4:
            n_comp_5453 = st.number_input("Comp. 5-4-5-3 (%)", min_value=0.0, max_value=100.0, value=0.0, step=1.0, key=f'cv_c54_{_rc}')

        # ---- CARD: Especie ----
        st.html('<div style="background:white;border-radius:12px;padding:16px 16px 4px 16px;margin-bottom:16px;box-shadow:0 1px 3px rgba(0,0,0,0.06);"><div style="font-size:1rem;font-weight:600;color:#1e293b;margin-bottom:12px;display:flex;align-items:center;gap:8px;padding-bottom:12px;border-bottom:1px solid #f1f5f9;"><span style="color:#059669;font-size:1.1rem;font-weight:700;">💶</span> Especie</div></div>')
        s1, s2, s3, s4 = st.columns(4)
        with s1:
            n_beneficios_vuelo = st.number_input("Benef. Vuelo (€)", min_value=0.0, value=0.0, step=1.0, key=f'cv_bv_{_rc}')
        with s2:
            n_seguro_medico = st.number_input("Seguro Med. (€)", min_value=0.0, value=0.0, step=1.0, key=f'cv_sm_{_rc}')
        with s3:
            n_estudios = st.number_input("RF Estudios (€)", min_value=0.0, value=0.0, step=1.0, key=f'cv_est_{_rc}')
        with s4:
            n_guarderia = st.number_input("RF Guardería (€)", min_value=0.0, value=0.0, step=1.0, key=f'cv_guard_{_rc}')

        # ---- CARD: IRPF ----
        st.html('<div style="background:white;border-radius:12px;padding:16px 16px 4px 16px;margin-bottom:16px;box-shadow:0 1px 3px rgba(0,0,0,0.06);"><div style="font-size:1rem;font-weight:600;color:#1e293b;margin-bottom:12px;display:flex;align-items:center;gap:8px;padding-bottom:12px;border-bottom:1px solid #f1f5f9;"><span style="color:#dc2626;font-size:1.1rem;font-weight:700;">%</span> IRPF</div></div>')
        i1, i2 = st.columns(2)
        with i1:
            pct_irpf = st.number_input("IRPF General (%)", min_value=0.0, max_value=50.0, value=0.0, step=0.5, key=f'cv_irpf_{_rc}')
        with i2:
            pct_irpf_esp = st.number_input("IRPF Especie (%)", min_value=0.0, max_value=50.0, value=0.0, step=0.5, key=f'cv_irpf_esp_{_rc}')

    # ========== CÁLCULOS ==========
    flexiworking_rate = d['flexiworking_jc'] if tipo == 'Sobrecargo' else d['flexiworking_tcp']
    compra_libre_rate = d['compra_libre_jc'] if tipo == 'Sobrecargo' else d['compra_libre_tcp']

    variables_items = [
        ('Hora de Vuelo', d['hora_vuelo'], n_horas_vuelo),
        ('Imaginaria', d['imaginaria'], n_imaginarias),
        ('Día Bloque Imaginaria', d['dia_bloque_imag'], n_bloque_imag),
        ('Saltos de Inspección', d['saltos_inspeccion'], n_saltos),
        ('Instrucción en Tierra', d['instruccion_tierra'], n_instruccion),
        ('Actividad en Tierra', d['actividad_tierra'], n_actividad),
        ('Flexiworking', flexiworking_rate, n_flexiworking),
        ('Plus Nocturnidad', d['plus_nocturnidad'], n_nocturnidad),
        ('Hora Posicional', d['hora_posicional'], n_hora_pos),
        ('Forzoso Adicional', d['forzoso_adicional'], n_forzoso),
        ('Hora E-Learning', d['hora_elearning'], n_elearning),
        ('Festivo', d['festivo'], n_festivo),
        ('Invasión OFF (0:00-0:59)', d['invasion_0'], n_inv0),
        ('Invasión OFF (1:00-1:59)', d['invasion_1'], n_inv1),
        ('Invasión OFF (≥2:00)', d['invasion_2'], n_inv2),
        ('Compra del Día Libre', compra_libre_rate, n_compra_libre),
        ('Cambio Servicio Prog.', d['cambio_servicio'], n_cambio_serv),
        ('Plus Act. Extendida', d['actividad_extendida'], n_act_ext),
        ('Días Sindicales', d['dias_sindicales'], n_dias_sindicales),
    ]
    variables_items_calc = [(name, rate, qty, rate * qty) for name, rate, qty in variables_items]
    total_variables = sum(r[3] for r in variables_items_calc)

    dietas_items = [
        ('Venta Patrón 5-3-5-3', d['venta_5353'], n_venta_5353, d['venta_5353'] * n_venta_5353 / 100),
        ('Venta Patrón 5-4-5-3', d['venta_5453'], n_venta_5453, d['venta_5453'] * n_venta_5453 / 100),
        ('Compensación 5-3-5-3', d['comp_5353'], n_comp_5353, d['comp_5353'] * n_comp_5353 / 100),
        ('Compensación 5-4-5-3', d['comp_5453'], n_comp_5453, d['comp_5453'] * n_comp_5453 / 100),
        ('Dieta de Vuelo', d['dieta_vuelo'], n_dieta_vuelo, d['dieta_vuelo'] * n_dieta_vuelo),
        ('Dieta de Formación', d['dieta_formacion'], n_dieta_form, d['dieta_formacion'] * n_dieta_form),
        ('D. Form. Pernocta Nac.', d['dieta_form_nac'], n_dieta_form_nac, d['dieta_form_nac'] * n_dieta_form_nac),
        ('D. Form. Pernocta Int.', d['dieta_form_int'], n_dieta_form_int, d['dieta_form_int'] * n_dieta_form_int),
        ('Pernocta Nacional', d['dieta_pernocta_nac'], n_pernocta_nac, d['dieta_pernocta_nac'] * n_pernocta_nac),
        ('Pernocta Internacional', d['dieta_pernocta_int'], n_pernocta_int, d['dieta_pernocta_int'] * n_pernocta_int),
        ('Destacamento', d['destacamento'], n_destacamento, d['destacamento'] * n_destacamento),
        ('Pernocta Meeting', d['pernocta_meeting'], n_pernocta_meet, d['pernocta_meeting'] * n_pernocta_meet),
        ('Vacaciones', d['vacaciones'], n_vacaciones, d['vacaciones'] * n_vacaciones),
        ('Comisión Venta a Bordo', 0, 0, n_comision_vab),
    ]
    total_dietas = sum(r[3] for r in dietas_items)
    total_especie = n_beneficios_vuelo + n_seguro_medico + n_estudios + n_guarderia
    bruto = total_fijos + total_variables + total_dietas + total_especie

    ded_irpf = bruto * pct_irpf / 100
    ded_irpf_esp = total_especie * pct_irpf_esp / 100
    ded_regimen = bruto * 4.83 / 100
    ded_dfp = bruto * 1.65 / 100
    ded_seguro_medico = n_seguro_medico
    total_deducciones = ded_irpf + ded_irpf_esp + ded_regimen + ded_dfp + ded_seguro_medico
    neto = bruto - total_deducciones

    # ========== RIGHT PANEL: KPIs + Desglose ==========
    with calc_right:
        svg_euro_teal = '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" style="stroke:#0d9488;width:16px;height:16px;fill:none;stroke-width:2;"><path d="M4 10h12"/><path d="M4 14h9"/><path d="M19.17 5a12 12 0 0 0-8.28 14.68"/><path d="M17 5a12.08 12.08 0 0 1 2.76 7.62"/></svg>'
        st.markdown(f'''
        <div class="kpi-card-outline">
            <div class="kpi-top">
                <div><div class="kpi-label">Conceptos Fijos</div><div class="kpi-value">{fmtc(total_fijos)}</div></div>
                <div class="kpi-icon">{svg_euro_teal}</div>
            </div>
        </div>''', unsafe_allow_html=True)

        svg_plane_white = '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" style="stroke:white;width:16px;height:16px;fill:none;stroke-width:2;"><path d="M17.8 19.2 16 11l3.5-3.5C21 6 21.5 4 21 3c-1-.5-3 0-4.5 1.5L13 8 4.8 6.2c-.5-.1-.9.1-1.1.5l-.3.5c-.2.5-.1 1 .3 1.3L9 12l-2 3H4l-1 1 3 2 2 3 1-1v-3l3-2 3.5 5.3c.3.4.8.5 1.3.3l.5-.2c.4-.3.6-.7.5-1.2z"/></svg>'
        st.markdown(f'''
        <div class="kpi-card-primary">
            <div class="kpi-top">
                <div><div class="kpi-label">Variables + Dietas</div><div class="kpi-value">{fmtc(total_variables + total_dietas)}</div></div>
                <div class="kpi-icon">{svg_plane_white}</div>
            </div>
        </div>''', unsafe_allow_html=True)

        if total_especie > 0:
            st.markdown(f'''
            <div class="kpi-card-warning">
                <div class="kpi-label">Retrib. Especie</div>
                <div class="kpi-value">{fmtc(total_especie)}</div>
            </div>''', unsafe_allow_html=True)

        st.markdown(f'''
        <div class="kpi-card-outline">
            <div><div class="kpi-label">Bruto Total</div><div class="kpi-value">{fmtc(bruto)}</div><div class="kpi-sub">Antes de deducciones</div></div>
        </div>''', unsafe_allow_html=True)

        st.markdown(f'''
        <div class="kpi-deduction">
            <div class="kpi-label">Deducciones</div>
            <div class="kpi-value">-{fmtc(total_deducciones)}</div>
        </div>''', unsafe_allow_html=True)

        svg_calc_white = '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" style="stroke:white;width:16px;height:16px;fill:none;stroke-width:2;"><rect width="16" height="20" x="4" y="2" rx="2"/><line x1="8" x2="16" y1="6" y2="6"/><line x1="16" x2="16" y1="14" y2="18"/><path d="M16 10h.01"/><path d="M12 10h.01"/><path d="M8 10h.01"/><path d="M12 14h.01"/><path d="M8 14h.01"/><path d="M12 18h.01"/><path d="M8 18h.01"/></svg>'
        st.markdown(f'''
        <div class="kpi-card-success">
            <div class="kpi-top">
                <div><div class="kpi-label">Neto Estimado</div><div class="kpi-value">{fmtc(neto)}</div></div>
                <div class="kpi-icon">{svg_calc_white}</div>
            </div>
        </div>''', unsafe_allow_html=True)

        # ---- Desglose ----
        desglose_html = '<div class="desglose-card">'
        desglose_html += '<div class="desglose-header">📋 Desglose</div>'
        desglose_html += '<div class="desglose-cols"><span>Concepto</span><span>€/ud</span><span>Uds</span><span>Total</span></div>'
        desglose_html += '<div class="desglose-body">'

        for r in fijos_rows:
            desglose_html += f'<div class="desglose-row"><span>{r[0]}</span><span>{fmtc(r[1])}</span><span>100%</span><span>{fmtc(r[2])}</span></div>'

        for name, rate, qty, imp in variables_items_calc:
            if imp > 0:
                desglose_html += f'<div class="desglose-row"><span>{name}</span><span>{fmtc(rate)}</span><span>{qty:g}</span><span>{fmtc(imp)}</span></div>'

        for name, rate, qty, imp in dietas_items:
            if imp > 0:
                rate_str = fmtc(rate) if rate > 0 else "—"
                qty_str = f"{qty:g}" if qty > 0 else "—"
                desglose_html += f'<div class="desglose-row"><span>{name}</span><span>{rate_str}</span><span>{qty_str}</span><span>{fmtc(imp)}</span></div>'

        if total_especie > 0:
            desglose_html += f'<div class="desglose-row"><span>Retrib. Especie</span><span>—</span><span>—</span><span>{fmtc(total_especie)}</span></div>'

        desglose_html += f'<div class="desglose-row highlight"><span>BRUTO</span><span></span><span></span><span>{fmtc(bruto)}</span></div>'

        if ded_irpf > 0:
            desglose_html += f'<div class="desglose-row"><span>IRPF ({pct_irpf}%)</span><span></span><span></span><span>-{fmtc(ded_irpf)}</span></div>'
        if ded_irpf_esp > 0:
            desglose_html += f'<div class="desglose-row"><span>IRPF Esp. ({pct_irpf_esp}%)</span><span></span><span></span><span>-{fmtc(ded_irpf_esp)}</span></div>'
        desglose_html += f'<div class="desglose-row"><span>Rég.+MEI (4,83%)</span><span></span><span></span><span>-{fmtc(ded_regimen)}</span></div>'
        desglose_html += f'<div class="desglose-row"><span>D+F+P (1,65%)</span><span></span><span></span><span>-{fmtc(ded_dfp)}</span></div>'
        if ded_seguro_medico > 0:
            desglose_html += f'<div class="desglose-row"><span>🏥 Seguro Médico</span><span></span><span></span><span>-{fmtc(ded_seguro_medico)}</span></div>'

        desglose_html += f'<div class="desglose-row highlight"><span>NETO</span><span></span><span></span><span>{fmtc(neto)}</span></div>'
        desglose_html += '</div></div>'
        st.markdown(desglose_html, unsafe_allow_html=True)

        # ---- Download Buttons ----
        def generate_nomina_excel():
            output = io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Nomina Nivel {nivel}"
            hf = PatternFill(start_color='0B847F', end_color='0B847F', fill_type='solid')
            hfont = Font(bold=True, color='FFFFFF', size=10)
            bf = Font(bold=True, size=10)
            brd = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                         top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
            ws['A1'] = f'NÓMINA ESTIMADA — Nivel {nivel} — {tipo}'
            ws['A1'].font = Font(bold=True, size=13, color='0B847F')
            ws.merge_cells('A1:D1')
            row = 3
            for i, h in enumerate(['Concepto', '€/unidad', 'Cantidad', 'Importe'], 1):
                c = ws.cell(row=row, column=i, value=h); c.fill = hf; c.font = hfont; c.border = brd
            row += 1
            for r in fijos_rows:
                ws.cell(row=row, column=1, value=r[0]).border = brd
                ws.cell(row=row, column=2, value=r[1]).border = brd; ws.cell(row=row, column=2).number_format = '#,##0.00'
                ws.cell(row=row, column=3, value='100%').border = brd
                ws.cell(row=row, column=4, value=r[2]).border = brd; ws.cell(row=row, column=4).number_format = '#,##0.00'
                row += 1
            ws.cell(row=row, column=3, value='TOTAL FIJOS').font = bf
            ws.cell(row=row, column=4, value=total_fijos).font = bf; ws.cell(row=row, column=4).number_format = '#,##0.00'
            row += 2
            for name, rate, qty, imp in variables_items_calc:
                if imp > 0:
                    ws.cell(row=row, column=1, value=name).border = brd
                    ws.cell(row=row, column=2, value=rate).border = brd; ws.cell(row=row, column=2).number_format = '#,##0.00'
                    ws.cell(row=row, column=3, value=qty).border = brd
                    ws.cell(row=row, column=4, value=imp).border = brd; ws.cell(row=row, column=4).number_format = '#,##0.00'
                    row += 1
            for name, rate, qty, imp in dietas_items:
                if imp > 0:
                    ws.cell(row=row, column=1, value=name).border = brd
                    ws.cell(row=row, column=2, value=rate).border = brd; ws.cell(row=row, column=2).number_format = '#,##0.00'
                    ws.cell(row=row, column=3, value=qty).border = brd
                    ws.cell(row=row, column=4, value=imp).border = brd; ws.cell(row=row, column=4).number_format = '#,##0.00'
                    row += 1
            row += 1
            hl = PatternFill(start_color='E6F7F7', end_color='E6F7F7', fill_type='solid')
            ws.cell(row=row, column=3, value='BRUTO:').font = Font(bold=True, size=12, color='0B847F')
            ws.cell(row=row, column=4, value=bruto).font = Font(bold=True, size=12, color='0B847F')
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=3).fill = hl; ws.cell(row=row, column=4).fill = hl
            row += 2
            deductions_list = [(f'IRPF ({pct_irpf}%)', ded_irpf), (f'IRPF Esp. ({pct_irpf_esp}%)', ded_irpf_esp), ('Cotiz. Rég.+MEI (4,83%)', ded_regimen), ('Cotiz. D+F+P (1,65%)', ded_dfp)]
            if ded_seguro_medico > 0:
                deductions_list.append(('Seguro Médico', ded_seguro_medico))
            for label, val in deductions_list:
                ws.cell(row=row, column=1, value=label).border = brd
                ws.cell(row=row, column=4, value=-val).border = brd; ws.cell(row=row, column=4).number_format = '#,##0.00'
                row += 1
            row += 1
            ws.cell(row=row, column=3, value='NETO:').font = Font(bold=True, size=13, color='065A57')
            ws.cell(row=row, column=4, value=neto).font = Font(bold=True, size=13, color='065A57')
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=3).fill = hl; ws.cell(row=row, column=4).fill = hl
            for col_cells in ws.columns:
                try:
                    letter = col_cells[0].column_letter
                except AttributeError:
                    continue
                mx = max(len(str(c.value or '')) for c in col_cells if hasattr(c, 'column_letter')) + 3
                ws.column_dimensions[letter].width = min(mx, 35)
            wb.save(output)
            return output.getvalue()

        def generate_nomina_pdf():
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import mm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            buf = io.BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
            els = []
            styles = getSampleStyleSheet()
            teal = colors.HexColor('#0B847F'); lt = colors.HexColor('#f1fafa'); wh = colors.white
            title_s = ParagraphStyle('TN', parent=styles['Title'], fontSize=14, textColor=teal, spaceAfter=6)
            sub_s = ParagraphStyle('SN', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#64748B'), spaceAfter=8)
            els.append(Paragraph(f"Nómina Estimada — Nivel {nivel} — {tipo}", title_s))
            els.append(Paragraph("Simulación de nómina estimada", sub_s))
            els.append(Spacer(1, 4*mm))
            pdf_data = [['Concepto', '€/ud', 'Uds', 'Importe']]
            for r in fijos_rows:
                pdf_data.append([r[0], f"{r[1]:,.2f}", '100%', f"{r[2]:,.2f} €"])
            pdf_data.append(['TOTAL FIJOS', '', '', f"{total_fijos:,.2f} €"])
            for name, rate, qty, imp in variables_items_calc:
                if imp > 0:
                    pdf_data.append([name, f"{rate:,.2f}", f"{qty:g}", f"{imp:,.2f} €"])
            for name, rate, qty, imp in dietas_items:
                if imp > 0:
                    pdf_data.append([name, f"{rate:,.2f}" if rate > 0 else "—", f"{qty:g}", f"{imp:,.2f} €"])
            pdf_data.append(['BRUTO TOTAL', '', '', f"{bruto:,.2f} €"])
            if ded_irpf > 0:
                pdf_data.append([f'IRPF ({pct_irpf}%)', '', '', f"-{ded_irpf:,.2f} €"])
            if ded_irpf_esp > 0:
                pdf_data.append([f'IRPF Especies ({pct_irpf_esp}%)', '', '', f"-{ded_irpf_esp:,.2f} €"])
            pdf_data.append(['Cotiz. Régimen+MEI (4,83%)', '', '', f"-{ded_regimen:,.2f} €"])
            pdf_data.append(['Cotiz. D+F+P (1,65%)', '', '', f"-{ded_dfp:,.2f} €"])
            if ded_seguro_medico > 0:
                pdf_data.append(['Seguro Médico', '', '', f"-{ded_seguro_medico:,.2f} €"])
            pdf_data.append(['NETO ESTIMADO', '', '', f"{neto:,.2f} €"])
            bruto_idx = None; neto_idx = None; fijo_idx = None
            for i, row in enumerate(pdf_data):
                if row[0] == 'TOTAL FIJOS': fijo_idx = i
                if row[0] == 'BRUTO TOTAL': bruto_idx = i
                if row[0] == 'NETO ESTIMADO': neto_idx = i
            t = Table(pdf_data, colWidths=[200, 60, 50, 80])
            sty = [
                ('BACKGROUND', (0, 0), (-1, 0), teal), ('TEXTCOLOR', (0, 0), (-1, 0), wh),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 7.5), ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
                ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]
            if fijo_idx:
                sty += [('BACKGROUND', (0, fijo_idx), (-1, fijo_idx), lt), ('FONTNAME', (0, fijo_idx), (-1, fijo_idx), 'Helvetica-Bold')]
            if bruto_idx:
                sty += [('BACKGROUND', (0, bruto_idx), (-1, bruto_idx), teal), ('TEXTCOLOR', (0, bruto_idx), (-1, bruto_idx), wh), ('FONTNAME', (0, bruto_idx), (-1, bruto_idx), 'Helvetica-Bold')]
            if neto_idx:
                sty += [('BACKGROUND', (0, neto_idx), (-1, neto_idx), colors.HexColor('#065A57')), ('TEXTCOLOR', (0, neto_idx), (-1, neto_idx), wh), ('FONTNAME', (0, neto_idx), (-1, neto_idx), 'Helvetica-Bold'), ('FONTSIZE', (0, neto_idx), (-1, neto_idx), 9)]
            t.setStyle(TableStyle(sty))
            els.append(t)
            doc.build(els)
            return buf.getvalue()

        dcol1, dcol2 = st.columns(2)
        with dcol1:
            st.download_button("📥 Excel", data=generate_nomina_excel(),
                               file_name=f"Nomina_Nivel{nivel}_{tipo}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
        with dcol2:
            st.download_button("📥 PDF", data=generate_nomina_pdf(),
                               file_name=f"Nomina_Nivel{nivel}_{tipo}.pdf",
                               mime="application/pdf",
                               use_container_width=True)

    # Footer
    st.markdown('<div class="calc-footer"><p>✈️ Herramientas Sindicales © 2026</p></div>', unsafe_allow_html=True)

    st.stop()




# ===== SECCIÓN: ANALIZADOR =====
# Renderizar sidebar de admin si corresponde
render_admin_sidebar()

# Barra de usuario con logout y botón volver
_back_col, _usr_col1, _usr_col2, _usr_col3 = st.columns([1, 4, 2, 1])
with _back_col:
    if st.button("← Volver", key="back_analizador"):
        st.session_state['app_section'] = 'seleccion'
        st.rerun()
with _usr_col1:
    st.markdown(f"👤 Conectado como: **{st.session_state.get('username', '')}**" + 
                (" 👑" if st.session_state.get('is_admin') else ""))
with _usr_col3:
    if st.button("🚪 Cerrar Sesión", key="logout_btn", type="secondary"):
        for key in ['authenticated', 'username', 'is_admin']:
            st.session_state[key] = False if key == 'authenticated' else None
        st.session_state['app_section'] = 'seleccion'
        st.rerun()

st.markdown("""
    <div class="main-header">
        <h1>✈️ Analizador de Programaciones</h1>
        <div class="accent-line"></div>
        <p>Sistema de análisis de programaciones de tripulación</p>
    </div>
""", unsafe_allow_html=True)

# Selector de base - Orden oficial CON opción por defecto
base_codes = [''] + [b[0] for b in BASES_ORDENADAS]  # Añadir opción vacía al inicio
base_labels = {b[0]: b[1] for b in BASES_ORDENADAS}
base_labels[''] = '🌍 Selecciona una base...'  # Texto para opción por defecto
base_sel = st.selectbox(
    "🌍 Selecciona la Base:",
    base_codes,
    format_func=lambda x: base_labels.get(x, x),
    help="Selecciona la base de operaciones para analizar",
    index=0  # Seleccionar la primera opción (vacía) por defecto
)

st.markdown("---")

# Verificar que se ha seleccionado una base válida
if not base_sel:
    st.info("👆 Por favor, selecciona una base para continuar con el análisis.")
    st.stop()


# Carga de archivos — diseño personalizado sin botón "+"
if base_sel == 'BCN':
    st.markdown("#### 📂 Carga de Archivos (BCN requiere múltiples PDFs)")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("""<div class="upload-container">
            <div class="upload-header">📁 Programación INICIAL</div>
        """, unsafe_allow_html=True)
        i1 = st.file_uploader("Parte A", type="pdf", key="ini_a", label_visibility="collapsed")
        i2 = st.file_uploader("Parte B", type="pdf", key="ini_b", label_visibility="collapsed")
        st.markdown("</div>", unsafe_allow_html=True)
    with c2:
        st.markdown("""<div class="upload-container">
            <div class="upload-header">📁 Programación FINAL</div>
        """, unsafe_allow_html=True)
        f1 = st.file_uploader("Parte A", type="pdf", key="fin_a", label_visibility="collapsed")
        f2 = st.file_uploader("Parte B", type="pdf", key="fin_b", label_visibility="collapsed")
        st.markdown("</div>", unsafe_allow_html=True)
    ready = (i1 or i2) and (f1 or f2)
    f_init, f_final = [i1, i2], [f1, f2]
else:
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("""<div class="upload-container">
            <div class="upload-header"><span class="upload-header-icon">📁</span> Programación INICIAL</div>
        """, unsafe_allow_html=True)
        iu = st.file_uploader("📂 Arrastra tu archivo aquí o usa el botón de abajo", type="pdf", help="Límite: 2GB · PDF", key="inicial_pdf")
        st.markdown("</div>", unsafe_allow_html=True)
    with c2:
        st.markdown("""<div class="upload-container">
            <div class="upload-header"><span class="upload-header-icon">📁</span> Programación FINAL</div>
        """, unsafe_allow_html=True)
        fu = st.file_uploader("📂 Arrastra tu archivo aquí o usa el botón de abajo", type="pdf", help="Límite: 2GB · PDF", key="final_pdf")
        st.markdown("</div>", unsafe_allow_html=True)
    ready = iu and fu
    f_init, f_final = [iu], [fu]

# JavaScript para ocultar el botón "+" dinámicamente (refuerzo del CSS)
st.markdown("""
<script>
(function hideAddButtons() {
    // Función que busca y oculta botones "+" en file uploaders
    function hide() {
        // Buscar todos los file uploaders
        const uploaders = document.querySelectorAll('[data-testid="stFileUploader"]');
        uploaders.forEach(uploader => {
            // Si hay un archivo subido
            const hasFile = uploader.querySelector('[data-testid="stUploadedFileLayout"]') 
                         || uploader.querySelector('.uploadedFile')
                         || uploader.querySelector('li');
            if (hasFile) {
                // Ocultar el dropzone (que contiene el "+" o "Browse files")
                const dropzone = uploader.querySelector('[data-testid="stFileUploaderDropzone"]');
                if (dropzone) {
                    dropzone.style.display = 'none';
                    dropzone.style.height = '0';
                    dropzone.style.overflow = 'hidden';
                }
            }
            // Ocultar botones minimal (el "+") siempre
            const minBtns = uploader.querySelectorAll('[data-testid="stBaseButton-minimal"], button[data-testid="baseButton-minimal"]');
            minBtns.forEach(btn => {
                btn.style.display = 'none';
                btn.style.visibility = 'hidden';
            });
        });
    }
    // Ejecutar al cargar y observar cambios
    hide();
    setInterval(hide, 500);
    const observer = new MutationObserver(hide);
    observer.observe(document.body, { childList: true, subtree: true });
})();
</script>
""", unsafe_allow_html=True)

# Botón de procesamiento
if ready:
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if st.button("⚙️ Procesar y Analizar Programaciones", use_container_width=True, type="primary"):
            st.markdown("---")
            progress_bar = st.progress(0)
            status_text = st.empty()
            pct_text = st.empty()
            
            # ═══ CSS ANIMACIONES (se inyectan una sola vez) ═══
            st.markdown("""
            <style>
            @keyframes flyAcross {
                0% { transform: translateX(-80px) scaleX(1); }
                49% { transform: translateX(80px) scaleX(1); }
                50% { transform: translateX(80px) scaleX(-1); }
                99% { transform: translateX(-80px) scaleX(-1); }
                100% { transform: translateX(-80px) scaleX(1); }
            }
            @keyframes pulse {
                0%, 100% { opacity: 1; }
                50% { opacity: 0.5; }
            }
            .loading-container {
                text-align: center;
                padding: 25px 20px;
                margin: 10px 0;
            }
            .loading-plane {
                display: inline-block;
                font-size: 2.8rem;
                animation: flyAcross 2.5s ease-in-out infinite;
            }
            .loading-text {
                margin-top: 12px;
                font-size: 1.2rem;
                color: #0D5F5D;
                font-weight: 700;
                animation: pulse 1.5s ease-in-out infinite;
            }
            .loading-dots::after {
                content: '';
                animation: dots 1.5s steps(4, end) infinite;
            }
            @keyframes dots {
                0% { content: ''; }
                25% { content: '.'; }
                50% { content: '..'; }
                75% { content: '...'; }
            }
            .pct-display {
                text-align: center;
                font-size: 1.6rem;
                font-weight: 800;
                color: #117F7C;
                margin: 5px 0;
                letter-spacing: 1px;
            }
            </style>
            """, unsafe_allow_html=True)
            
            def _update_progress(pct, emoji, message):
                """Helper para actualizar barra + texto + porcentaje de forma consistente."""
                progress_bar.progress(pct / 100)
                status_text.markdown(f"""
                <div class="loading-container">
                    <div class="loading-plane">{emoji}</div>
                    <p class="loading-text">{message}<span class="loading-dots"></span></p>
                </div>
                """, unsafe_allow_html=True)
                pct_text.markdown(f'<div class="pct-display">{pct}%</div>', unsafe_allow_html=True)
            
            _update_progress(5, "✈️", "Extrayendo datos de programaciones")
            import time as _time
            for _smooth_pct in range(6, 11):
                _time.sleep(0.08)
                progress_bar.progress(_smooth_pct / 100)
                pct_text.markdown(f'<div class="pct-display">{_smooth_pct}%</div>', unsafe_allow_html=True)
            
            try:
                def _subprogress(done, total, msg, start_pct, end_pct, emoji="📄"):
                    total = max(total, 1)
                    frac = max(0.0, min(1.0, done / total))
                    pct = int(start_pct + (end_pct - start_pct) * frac)
                    _update_progress(pct, emoji, f"{msg} · Procesando página {done}/{total}")

                _safe_rewind_files(f_init)
                _update_progress(15, "📋", "Extrayendo programación inicial")
                r_i = extract_roster_stream(f_init, base_sel)
                _update_progress(32, "📋", "Programación inicial extraída")
                gc.collect()

                _safe_rewind_files(f_final)
                _update_progress(32, "📋", "Extrayendo programación final")
                r_f = extract_roster_stream(f_final, base_sel)
                _update_progress(48, "📋", "Programación final extraída")
                gc.collect()

                sectors_ini = {}
                sectors_fin = {}

                _safe_rewind_files(f_init)
                _update_progress(50, "🔍", "Analizando sectores (inicial)")
                try:
                    sectors_ini = extract_sectors_by_day_pdfplumber(f_init)
                except Exception:
                    sectors_ini = {}

                _safe_rewind_files(f_final)
                _update_progress(57, "🔍", "Analizando sectores (final)")
                try:
                    sectors_fin = extract_sectors_by_day_pdfplumber(f_final)
                except Exception:
                    sectors_fin = {}

                gc.collect()

                _safe_rewind_files(f_final)
                _update_progress(64, "⏱️", "Calculando long duties")
                try:
                    long_duty_data = extract_daily_duty_data(f_final)
                except Exception:
                    long_duty_data = {}

                _safe_rewind_files(f_final)
                _update_progress(74, "⏱️", "Long duties extraídos")
            except Exception as processing_error:
                progress_bar.empty()
                status_text.empty()
                pct_text.empty()
                _release_uploaded_files(f_init + f_final, clear_widget_keys=False)
                st.error("❌ Error durante el procesamiento. La sesión se mantiene activa para reintentar.")
                with st.expander("Detalle técnico del error", expanded=False):
                    st.code(traceback.format_exc())
                st.stop()
            
            if not r_f:
                progress_bar.empty()
                status_text.empty()
                pct_text.empty()
                _release_uploaded_files(f_init + f_final, clear_widget_keys=False)
                st.error(f"⚠️ **No se encontraron tripulantes** en la programación final.")
                st.warning(f"Verifica que los archivos PDF contienen programaciones válidas.")
                st.stop()
            
            if r_f:
                _update_progress(78, "📊", "Procesando análisis")
                
                # === MOSTRAR VALIDACIÓN DE EXTRACCIÓN ===
                meta_f = r_f.get('__meta__', {})
                val = meta_f.get('validation', {})
                if val:
                    total_unique = val.get('unique_ids_final', 0)
                    header_lines = val.get('crew_header_lines', 0)
                    with_id = val.get('crew_with_id', 0)
                    without_id = val.get('crew_without_id', 0)
                    overwritten = val.get('crew_overwritten', 0)
                    bases_found = val.get('bases_encontradas', {})
                    
                    # Detectar tripulantes desplazados (base oficial ≠ base seleccionada)
                    desplazados_en_pdf = {}
                    for _cid, _cdata in r_f.items():
                        if _cid == '__meta__':
                            continue
                        _cb = _cdata.get('Base', '')
                        if _cb and _cb != base_sel:
                            desplazados_en_pdf.setdefault(_cb, []).append((_cid, _cdata.get('Name', '?')))
                    
                    n_desplazados = sum(len(v) for v in desplazados_en_pdf.values())
                    n_base_oficial = total_unique - n_desplazados
                    
                    if n_desplazados > 0:
                        st.success(f"✅ **Extracción completada**: {total_unique} tripulantes únicos detectados "
                                   f"({n_base_oficial} de {base_sel} + {n_desplazados} desplazado(s))")
                    else:
                        st.success(f"✅ **Extracción completada**: {total_unique} tripulantes únicos detectados")
                    
                    with st.expander("🔍 Detalle de validación de extracción", expanded=False):
                        vc1, vc2, vc3, vc4 = st.columns(4)
                        vc1.metric("Líneas crew", header_lines)
                        vc2.metric("Con ID", with_id)
                        vc3.metric("Duplicados", overwritten)
                        vc4.metric("Únicos final", total_unique)
                        
                        if bases_found:
                            st.markdown("**Desglose por base:**")
                            bases_str = " | ".join([f"`{b}`: {c}" for b, c in sorted(bases_found.items())])
                            st.markdown(bases_str)
                        
                        if n_desplazados > 0:
                            st.info(f"🔄 **{n_desplazados} tripulante(s) desplazado(s)** detectados en el PDF "
                                    f"(base oficial diferente a {base_sel}):")
                            for _db, _dcrew in sorted(desplazados_en_pdf.items()):
                                names_str = ", ".join([f"{_did} ({_dn})" for _did, _dn in _dcrew[:5]])
                                if len(_dcrew) > 5:
                                    names_str += f" ... y {len(_dcrew)-5} más"
                                st.markdown(f"  - **{_db}** → {base_sel}: {len(_dcrew)} tripulante(s) - {names_str}")
                        
                        cats_found = val.get('categorias_encontradas', {})
                        if cats_found:
                            st.markdown(f"**Por categoría:** JC: {cats_found.get('JC', 0)} | TC: {cats_found.get('TC', 0)}")
                        
                        fleet_found = val.get('fleet_codes', {})
                        if fleet_found:
                            st.markdown(f"**Por flota:** {fleet_found}")
                        
                        if without_id > 0:
                            st.warning(f"⚠️ {without_id} tripulantes sin ID detectado - revisar PDFs")
                            warnings = val.get('warnings', [])
                            for w in warnings[:10]:
                                st.text(w)
                
                dias_mes = detectar_dias_mes(r_f)
                
                comp_data = []
                all_codes = set()
                categorias_resumen = defaultdict(lambda: {'tripulantes': set(), 'dias': 0, 'codigos': defaultdict(lambda: {'tripulantes': set(), 'dias': 0})})
                datos_por_dia = defaultdict(lambda: defaultdict(int))
                total_crews = len([_k for _k in r_f.keys() if _k != '__meta__'])
                processed_crews = 0

                for cid, data in r_f.items():
                    if cid == '__meta__':
                        continue

                    processed_crews += 1
                    if processed_crews % 80 == 0 or processed_crews == total_crews:
                        frac_crews = processed_crews / max(total_crews, 1)
                        pct_crews = int(74 + (79 - 74) * frac_crews)
                        _update_progress(pct_crews, "📊", f"Procesando tripulantes {processed_crews}/{total_crews}")

                    # CAMBIO CRÍTICO v3.1: NO filtrar por base oficial.
                    # Los PDFs de una base incluyen tanto tripulantes oficiales como
                    # desplazados que operan desde esa base. Filtrar por base oficial
                    # excluía a los desplazados (ej: BCN en PDF de PMI).
                    # El usuario ya seleccionó la base y subió PDFs específicos,
                    # así que TODOS los tripulantes del PDF deben analizarse.
                    crew_base = data.get('Base', '')
                    f_sched = data['Schedule']
                    
                    if len(f_sched) < dias_mes:
                        f_sched = f_sched + ['--'] * (dias_mes - len(f_sched))
                    elif len(f_sched) > dias_mes:
                        f_sched = f_sched[:dias_mes]
                    
                    # Obtener Block inicial si existe en r_i
                    block_inicial = '--'
                    if cid in r_i:
                        i_sched = list(r_i[cid]['Schedule'])
                        # Normalizar i_sched a dias_mes (igual que f_sched)
                        if len(i_sched) < dias_mes:
                            i_sched = i_sched + ['--'] * (dias_mes - len(i_sched))
                        elif len(i_sched) > dias_mes:
                            i_sched = i_sched[:dias_mes]
                        # Comparación mejorada: usa sectores detallados si están disponibles
                        i_sec = sectors_ini.get(cid, None)
                        f_sec = sectors_fin.get(cid, None)
                        mods = count_real_changes(i_sched, f_sched, dias_mes, i_sec, f_sec)
                        block_inicial = r_i[cid].get('Block', '--')
                    else:
                        mods = dias_mes
                    
                    row = {
                        'ID': cid,
                        'Nombre': data['Name'],
                        'Cat': data['Categoria'],
                        'Block_Inicial': block_inicial,  # Horas BLOCK de programación inicial
                        'Block': data.get('Block', '--'),  # Horas BLOCK de programación final
                        'Duty': data.get('Duty', '--'),
                        'Estabilidad': round(((dias_mes - mods) / dias_mes) * 100, 1),
                        'exact_sched': {idx + 1: code for idx, code in enumerate(f_sched)},
                        'mods': mods,
                        'Cambios': mods
                    }
                    
                    code_days = defaultdict(list)
                    for idx, code in enumerate(f_sched):
                        c = code.strip().upper()
                        code_days[c].append(idx + 1)
                        all_codes.add(c)
                        
                        cat_id = get_categoria_principal(c)
                        categorias_resumen[cat_id]['tripulantes'].add(cid)
                        categorias_resumen[cat_id]['dias'] += 1
                        categorias_resumen[cat_id]['codigos'][c]['tripulantes'].add(cid)
                        categorias_resumen[cat_id]['codigos'][c]['dias'] += 1
                        
                        datos_por_dia[idx + 1][cat_id] += 1
                    
                    row['code_days'] = dict(code_days)
                    comp_data.append(row)
                
                _update_progress(88, "📊", "Generando resumen ejecutivo")
                
                df_c = pd.DataFrame(comp_data)
                
                # Resumen por código
                resumen_codigos = []
                for code in sorted(all_codes):
                    afectados = []
                    total_dias = 0
                    for _, row in df_c.iterrows():
                        if code in row['code_days']:
                            dias = row['code_days'][code]
                            total_dias += len(dias)
                            afectados.append({
                                'ID': row['ID'],
                                'Nombre': row['Nombre'],
                                'Cat': row['Cat'],
                                'Dias': len(dias),
                                'Fechas': get_fechas_texto(dias)
                            })
                    
                    if afectados:
                        resumen_codigos.append({
                            'Codigo': code,
                            'Descripcion': get_codigo_descripcion(code),
                            'Categoria': get_categoria_principal(code),
                            'Afectados': len(afectados),
                            'Total_Dias': total_dias,
                            'Detalle': afectados
                        })
                
                df_r = pd.DataFrame(resumen_codigos)
                
                # Serializar categorías
                cat_resumen_serializable = {}
                for cat_id, cat_data in categorias_resumen.items():
                    cat_resumen_serializable[cat_id] = {
                        'tripulantes': len(cat_data['tripulantes']),
                        'dias': cat_data['dias'],
                        'codigos': {}
                    }
                    for code, code_data in cat_data['codigos'].items():
                        cat_resumen_serializable[cat_id]['codigos'][code] = {
                            'tripulantes': len(code_data['tripulantes']),
                            'dias': code_data['dias']
                        }
                
                # Serializar datos_por_dia
                datos_dia_serial = {d: dict(cats) for d, cats in datos_por_dia.items()}
                
                _update_progress(95, "✅", "Finalizando")
                import time as _time2
                for _smooth in range(96, 101):
                    _time2.sleep(0.05)
                    progress_bar.progress(_smooth / 100)
                    pct_text.markdown(f'<div class="pct-display">{_smooth}%</div>', unsafe_allow_html=True)
                status_text.empty()
                progress_bar.empty()
                pct_text.empty()
                
                jc_count = (df_c['Cat'] == 'JC').sum()
                tc_count = (df_c['Cat'] == 'TC').sum()
                
                st.session_state['df_c'] = df_c
                st.session_state['df_r'] = df_r
                st.session_state['dias_mes'] = dias_mes
                st.session_state['all_codes'] = sorted(all_codes)
                st.session_state['categorias_resumen'] = cat_resumen_serializable
                st.session_state['base_procesada'] = base_sel
                st.session_state['datos_por_dia'] = datos_dia_serial
                # Guardar mes y año extraídos de la cabecera del PDF
                # meta_f ya fue obtenido antes de la validación
                st.session_state['header_mes'] = meta_f.get('header_mes')
                st.session_state['header_anio'] = meta_f.get('header_anio')
                st.session_state['validation_data'] = meta_f.get('validation', {})
                # Calcular FTG, UNFIT, Bajas
                _ftg_trips = set()
                _unfit_trips = set()
                _baja_trips = set()
                _ftg_codes = ['FTG']
                _unfit_codes = ['UNFT', 'UNFD']
                _baja_codes = ['MED', 'ASEP', 'SICK', 'SICD', 'LSIC', 'LSCK', 'NJSK']
                for _, _row in df_c.iterrows():
                    _cd = _row.get('code_days', {})
                    for _fc in _ftg_codes:
                        if _fc in _cd and len(_cd[_fc]) > 0:
                            _ftg_trips.add(_row['ID'])
                    for _uc in _unfit_codes:
                        if _uc in _cd and len(_cd[_uc]) > 0:
                            _unfit_trips.add(_row['ID'])
                    for _bc in _baja_codes:
                        if _bc in _cd and len(_cd[_bc]) > 0:
                            _baja_trips.add(_row['ID'])
                
                # Calcular block medios por categoría (inicial y final)
                _jc_mask = df_c['Cat'] == 'JC'
                _tc_mask = df_c['Cat'] == 'TC'
                
                def _avg_block(mask, col):
                    vals = df_c.loc[mask, col].apply(parse_hours)
                    vals = vals[vals > 0]
                    return vals.mean() if len(vals) > 0 else 0
                
                _block_medio_jc_ini = _avg_block(_jc_mask, 'Block_Inicial')
                _block_medio_jc_fin = _avg_block(_jc_mask, 'Block')
                _block_medio_tc_ini = _avg_block(_tc_mask, 'Block_Inicial')
                _block_medio_tc_fin = _avg_block(_tc_mask, 'Block')
                
                def _fmt_hm(hours_dec):
                    """Formatea horas decimales como HH:MM"""
                    if hours_dec <= 0:
                        return "0:00"
                    h = int(hours_dec)
                    m = int(round((hours_dec - h) * 60))
                    if m == 60:
                        h += 1
                        m = 0
                    return f"{h}:{m:02d}"
                
                st.session_state['kpis'] = {
                    'total': len(df_c),
                    'jc': jc_count,
                    'tc': tc_count,
                    'mod': df_c['mods'].sum(),
                    'stab': df_c['Estabilidad'].mean(),
                    'ftg': len(_ftg_trips),
                    'unfit': len(_unfit_trips),
                    'bajas': len(_baja_trips),
                    'block_medio_jc_ini': _block_medio_jc_ini,
                    'block_medio_jc_fin': _block_medio_jc_fin,
                    'block_medio_tc_ini': _block_medio_tc_ini,
                    'block_medio_tc_fin': _block_medio_tc_fin,
                    'block_medio_jc_ini_fmt': _fmt_hm(_block_medio_jc_ini),
                    'block_medio_jc_fin_fmt': _fmt_hm(_block_medio_jc_fin),
                    'block_medio_tc_ini_fmt': _fmt_hm(_block_medio_tc_ini),
                    'block_medio_tc_fin_fmt': _fmt_hm(_block_medio_tc_fin),
                }
                st.session_state['long_duty_data'] = long_duty_data
                # Reducir memoria en session_state: conservar solo metadatos del roster final.
                st.session_state['r_f'] = {'__meta__': r_f.get('__meta__', {})}
                st.session_state['processed_data'] = {
                    'df_c': df_c,
                    'df_r': df_r,
                    'long_duty_data': long_duty_data,
                    'meta': r_f.get('__meta__', {}),
                }
                st.session_state['analysis_complete'] = True
                st.session_state['procesado'] = True
                # Generar informe automático
                try:
                    generate_report(
                        st.session_state.get('username', 'unknown'),
                        base_sel,
                        len([k for k in r_f if k != '__meta__'])
                    )
                except:
                    pass

                # Liberación explícita de memoria para bases grandes (BCN)
                try:
                    del comp_data, resumen_codigos, sectors_ini, sectors_fin, r_i, r_f
                except Exception:
                    pass
                _release_uploaded_files(f_init + f_final, clear_widget_keys=True)
                gc.collect()

                st.rerun()


# --- 6. VISUALIZACIÓN DE RESULTADOS ---
if st.session_state.get('procesado'):
    k = st.session_state['kpis']
    df_c = st.session_state['df_c']
    df_r = st.session_state['df_r']
    dias_mes = st.session_state['dias_mes']
    all_codes = st.session_state['all_codes']
    categorias_resumen = st.session_state.get('categorias_resumen', {})
    base_procesada = st.session_state.get('base_procesada', 'N/A')
    datos_por_dia = st.session_state.get('datos_por_dia', {})
    long_duty_data = st.session_state.get('long_duty_data', {})
    r_f = st.session_state.get('r_f', {})
    total_plantilla = k['total']
    
    # Obtener mes y año del análisis
    _h_mes_banner = st.session_state.get('header_mes')
    _h_anio_banner = st.session_state.get('header_anio')
    MESES_ES = {1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL', 5: 'MAYO', 6: 'JUNIO',
                7: 'JULIO', 8: 'AGOSTO', 9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'}
    mes_texto = ""
    if _h_mes_banner and _h_anio_banner:
        mes_texto = f"{MESES_ES.get(_h_mes_banner, '')} {_h_anio_banner} - "
    
    # Banner - Block horas medias por categoría
    _bm_jc_ini = k.get('block_medio_jc_ini_fmt', '0:00')
    _bm_jc_fin = k.get('block_medio_jc_fin_fmt', '0:00')
    _bm_tc_ini = k.get('block_medio_tc_ini_fmt', '0:00')
    _bm_tc_fin = k.get('block_medio_tc_fin_fmt', '0:00')
    
    st.markdown(f"""
        <div class="base-banner">
            <span class="base-name">✈️ {mes_texto}{base_procesada}</span>
            <span class="stat">{total_plantilla} tripulantes</span>
            <span class="stat">👤 JC: <span class="stat-num">{k.get('jc', 0)}</span></span>
            <span class="stat">👥 TCP: <span class="stat-num">{k.get('tc', 0)}</span></span>
        </div>
    """, unsafe_allow_html=True)
    

    
    # Pestañas principales - SIN panel de admin (ahora es sección separada)
    _tab_names = [
        "📊 Resumen Ejecutivo",
        "📈 Impacto por Categorías",
        "🔍 Buscador Individual",
        "👥 Relación de Empleados",
        "⏰ Long Duties",
        "⚠️ Actividad Continua >5 Días",
        "📋 Códigos Específicos",
        "📅 Calendario Diario",
        "🗄️ Base de Datos"
    ]
    
    _tabs = st.tabs(_tab_names)
    t1, t2, t3, t3b, t_ld, t_rot, t4, t5, t6 = _tabs[0], _tabs[1], _tabs[2], _tabs[3], _tabs[4], _tabs[5], _tabs[6], _tabs[7], _tabs[8]
    
    # ═══════════════════════════════════════════════════════════════
    # TAB 1: RESUMEN EJECUTIVO - ÉNFASIS EN PORCENTAJE
    # ═══════════════════════════════════════════════════════════════
    with t1:
        st.markdown('<p class="section-title" style="font-size: 1.5rem;">📊 Resumen Ejecutivo</p>', unsafe_allow_html=True)
        
        c1, c2, c3, c4 = st.columns(4)
        c1.markdown(f'<div class="kpi-box"><p>TRIPULANTES</p><h1>{k["total"]}</h1></div>', unsafe_allow_html=True)
        c2.markdown(f'<div class="kpi-box"><p>JC / TCP</p><h1>{k["jc"]} / {k["tc"]}</h1></div>', unsafe_allow_html=True)
        c3.markdown(f'<div class="kpi-box accent"><p>CAMBIOS TOTALES</p><h1>{int(k["mod"])}</h1></div>', unsafe_allow_html=True)
        c4.markdown(f'<div class="kpi-box"><p>CÓDIGOS DETECTADOS</p><h1>{len(all_codes)}</h1></div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        
        # Estabilidad Global
        stab_val = k['stab']
        stab_color = '#27AE60' if stab_val >= 80 else ('#E67E22' if stab_val >= 60 else '#E74C3C')
        stab_label = 'ALTA' if stab_val >= 80 else ('MEDIA' if stab_val >= 60 else 'BAJA')
        
        col_g1, col_g2 = st.columns([2, 3])
        with col_g1:
            st.markdown(f"""
                <div style="background: white; border-radius: 14px; padding: 30px; text-align: center;
                            box-shadow: 0 6px 22px rgba(0,0,0,0.08); border-top: 4px solid {stab_color};">
                    <p style="color: #64748B; font-size: 0.82rem; font-weight: 700; text-transform: uppercase; margin: 0; letter-spacing: 0.8px;">
                        ESTABILIDAD GLOBAL DE LA BASE</p>
                    <h1 style="font-size: 3.5rem; color: {stab_color}; margin: 10px 0 5px 0; font-weight: 900; letter-spacing: -1px;">{stab_val:.1f}%</h1>
                    <span style="background: {stab_color}; color: white; padding: 5px 18px; border-radius: 20px; 
                                 font-size: 0.8rem; font-weight: 800;">{stab_label}</span>
                    <p style="color: #94A3B8; font-size: 0.75rem; margin-top: 12px; font-weight: 500;">
                        📅 {dias_mes} días analizados</p>
                    <p style="color: #B0BEC5; font-size: 0.65rem; margin-top: 4px; font-weight: 400;">
                        Comparación programación inicial vs final</p>
                </div>
            """, unsafe_allow_html=True)
        
        with col_g2:
            fig_pie = go.Figure(data=[go.Pie(
                labels=['JC', 'TCP'],
                values=[k['jc'], k['tc']],
                hole=0.5,
                marker=dict(colors=['#0D5F5D', '#44BABC']),
                textinfo='label+percent',
                textfont=dict(size=14, family='Inter')
            )])
            fig_pie.update_layout(
                height=280,
                margin=dict(t=30, b=20, l=20, r=20),
                title=dict(text="Distribución de Plantilla", font=dict(size=14, color='#0D5F5D', family='Inter')),
                showlegend=False
            )
            st.plotly_chart(fig_pie, use_container_width=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # PORCENTAJES GLOBALES POR CONCEPTO - ÉNFASIS EN % PLANTILLA AFECTADA
        st.markdown('<p class="section-title" style="font-size: 1.5rem;">📊 Porcentajes Globales por Concepto</p>', unsafe_allow_html=True)
        
        total_dias_programados = total_plantilla * dias_mes
        
        st.markdown(f"""<div style="background: #F0F7F7; padding: 12px 20px; border-radius: 10px; margin-bottom: 15px; border-left: 4px solid #117F7C;">
<span style="color: #0D5F5D; font-size: 1rem; font-weight: 500;">
El <strong>% de plantilla afectada</strong> muestra el porcentaje de tripulantes con al menos un día de ese concepto (de {total_plantilla} total).
</span></div>""", unsafe_allow_html=True)
        
        for concepto_nombre, concepto_info in CONCEPTOS_GLOBALES.items():
            dias_concepto = 0
            trip_set = set()
            for code in concepto_info['codigos']:
                matching = df_r[df_r['Codigo'] == code] if not df_r.empty else pd.DataFrame()
                if not matching.empty:
                    dias_concepto += matching.iloc[0]['Total_Dias']
                    for det in matching.iloc[0]['Detalle']:
                        trip_set.add(det['ID'])
            
            pct_dias = round((dias_concepto / total_dias_programados * 100), 2) if total_dias_programados > 0 else 0
            n_trip = len(trip_set)
            pct_plantilla = round((n_trip / total_plantilla * 100), 1) if total_plantilla > 0 else 0
            promedio_dias = round(dias_concepto / n_trip, 1) if n_trip > 0 else 0
            
            if dias_concepto > 0 or n_trip > 0:
                # NUEVO DISEÑO: Porcentaje de plantilla EN GRANDE
                st.markdown(f"""
                    <div class="concepto-card" style="border-left-color: {concepto_info['color']}; padding: 18px 24px;">
                        <div style="flex: 1;">
                            <span style="font-size: 1.15rem; font-weight: 800; color: #0D5F5D;">{concepto_info['icono']} {concepto_nombre}</span>
                        </div>
                        <div style="text-align: right;">
                            <span style="font-size: 2.2rem; font-weight: 900; color: {concepto_info['color']}; line-height: 1;">{pct_plantilla}%</span>
                            <span style="display: block; font-size: 0.85rem; color: #64748B; font-weight: 600;">de plantilla afectada</span>
                            <span style="display: block; font-size: 0.95rem; color: #0D5F5D; font-weight: 700; margin-top: 4px;">{n_trip} tripulantes (de {total_plantilla} total)</span>
                            <span style="display: block; font-size: 0.8rem; color: #94A3B8; font-weight: 500;">{dias_concepto} días totales · ≈ {promedio_dias} días/trip.</span>
                        </div>
                    </div>
                """, unsafe_allow_html=True)
    
    # ═══════════════════════════════════════════════════════════════
    # TAB 2: IMPACTO POR CATEGORÍAS
    # ═══════════════════════════════════════════════════════════════
    with t2:
        st.markdown('<p class="section-title">📈 Impacto por Categorías Principales</p>', unsafe_allow_html=True)
        st.markdown(f"""
<div style="background: linear-gradient(135deg, #117F7C 0%, #44BABC 100%); 
            padding: 15px 25px; 
            border-radius: 12px; 
            box-shadow: 0 4px 12px rgba(17, 127, 124, 0.25);
            text-align: center;
            margin: 20px 0;">
    <span style="color: white; font-size: 1.2rem; font-weight: 600;">
        📊 Plantilla total: {total_plantilla} tripulantes
    </span>
</div>
""", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        
        cat_list = list(CATEGORIAS_PRINCIPALES.items())
        
        if 'otros' in categorias_resumen:
            cat_list.append(('otros', {'nombre': '⚠️ Otros (No catalogados)', 'icono': '⚠️', 'color': 'cat-otros', 'codigos': []}))
        
        for idx, (cat_id, cat_info) in enumerate(cat_list):
            col_idx = idx % 3
            # Crear nueva fila de columnas cada 3 tarjetas
            if col_idx == 0:
                cat_cols = st.columns(3)
            
            cat_data = categorias_resumen.get(cat_id, {'tripulantes': 0, 'dias': 0, 'codigos': {}})
            num_trip = cat_data.get('tripulantes', 0)
            num_dias = cat_data.get('dias', 0)
            porcentaje = round((num_trip / total_plantilla * 100), 1) if total_plantilla > 0 else 0
            promedio_dias = round(num_dias / num_trip, 1) if num_trip > 0 else 0
            
            codigos_detectados = list(cat_data.get('codigos', {}).keys())
            codigos_text = ", ".join(sorted(codigos_detectados)[:15])
            if len(codigos_detectados) > 15:
                codigos_text += f" (+{len(codigos_detectados)-15} más)"
            
            with cat_cols[col_idx]:
                st.markdown(
                    f"""<div class="category-card {cat_info['color']}">
                        <h2>{cat_info['icono']} {cat_info['nombre']}</h2>
                        <div class="cat-trip-row">
                            <span class="cat-number">{num_trip}</span>
                            <span class="cat-trip-label">Tripulantes</span>
                        </div>
                        <p class="cat-info"><strong>{porcentaje}%</strong> de la plantilla</p>
                        <p class="cat-info">📅 {num_dias} días totales · ≈ {promedio_dias} días/trip.</p>
                        <div class="cat-codes">Códigos: {codigos_text if codigos_text else "—"}</div>
                    </div>""",
                    unsafe_allow_html=True
                )
    
    # ═══════════════════════════════════════════════════════════════
    # TAB 3: BUSCADOR INDIVIDUAL - CON BOTÓN LIMPIAR
    # ═══════════════════════════════════════════════════════════════
    with t3:
        st.markdown('<p class="section-title">🔍 Buscador Individual</p>', unsafe_allow_html=True)
        st.markdown("""
            <div style="background: #F0F7F7; border-radius: 10px; padding: 14px 20px; margin-bottom: 18px; border-left: 4px solid #117F7C;">
                <span style="color: #0D5F5D; font-size: 0.92rem; font-weight: 500;">
                    Aquí podrás ver el resumen completo de la programación de cada tripulante, incluyendo su calendario mensual, horas de vuelo y estabilidad.
                </span>
            </div>
        """, unsafe_allow_html=True)
        
        # Inicializar banderas para controlar visibilidad de detalles y limpieza
        if 'show_crew_details' not in st.session_state:
            st.session_state.show_crew_details = False
        # CSS para alinear perfectamente los botones Buscar y Limpiar
        st.markdown("""
        <style>
        /* Alinear botones de búsqueda verticalmente */
        div[data-testid="column"]:has(button) {
            display: flex !important;
            align-items: flex-end !important;
        }
        /* Botón Buscar - estilo primario */
        .search-buttons button[kind="primary"] {
            height: 42px !important;
            min-height: 42px !important;
        }
        /* Botón Limpiar del buscador - estilo visual consistente con Buscar */
        div[data-testid="stButton"] > button[kind="secondary"],
        button[data-testid="stBaseButton-secondary"] {
            height: 42px !important;
            min-height: 42px !important;
            background: linear-gradient(135deg, #64748B 0%, #94A3B8 100%) !important;
            color: white !important;
            border: none !important;
            font-weight: 700 !important;
            border-radius: 10px !important;
            font-size: 0.95rem !important;
            box-shadow: 0 4px 12px rgba(100, 116, 139, 0.2) !important;
            transition: all 0.2s ease !important;
        }
        div[data-testid="stButton"] > button[kind="secondary"]:hover,
    button[data-testid="stBaseButton-secondary"]:hover {
            background: linear-gradient(135deg, #475569 0%, #64748B 100%) !important;
            box-shadow: 0 6px 16px rgba(100, 116, 139, 0.4) !important;
            transform: translateY(-1px) !important;
        }
        </style>
        """, unsafe_allow_html=True)
        
        # Callback para limpiar búsqueda del buscador individual
        def _clear_individual_search():
            st.session_state['search_input'] = ""
            if 'selected_crew' in st.session_state:
                del st.session_state['selected_crew']
            st.session_state.show_crew_details = False
        
        # Barra de búsqueda con botones Buscar y Limpiar alineados
        col_search, col_btn, col_clear = st.columns([6, 2, 2])
        with col_search:
            search_term = st.text_input(
                "Buscar tripulante:",
                placeholder="Introduce ID o nombre",
                label_visibility="collapsed",
                key="search_input"
            )
        with col_btn:
            search_clicked = st.button("🔍 Buscar", use_container_width=True, type="primary")
        with col_clear:
            st.button("🗑️ Limpiar", use_container_width=True, 
                       help="Limpia la búsqueda y cierra la ficha del tripulante",
                       type="secondary", on_click=_clear_individual_search)
        
        selected_crew_id = st.session_state.get('selected_crew', None)
        
        show_individual = False
        selected_row = None
        
        if selected_crew_id:
            matches = df_c[df_c['ID'] == selected_crew_id]
            if len(matches) > 0:
                with st.spinner("Cargando información del tripulante..."):
                    import time
                    time.sleep(0.5)
                selected_row = matches.iloc[0]
                show_individual = True
        
        if (search_term or search_clicked) and not show_individual:
            term = search_term.strip().upper() if search_term else ""
            if term:
                matches = df_c[
                    (df_c['ID'].astype(str).str.contains(term)) | 
                    (df_c['Nombre'].str.upper().str.contains(term))
                ]
                
                if len(matches) == 0:
                    st.warning(f"❌ No se encontró ningún tripulante con '{term}'")
                elif len(matches) == 1:
                    with st.spinner("Cargando información del tripulante..."):
                        import time
                        time.sleep(0.5)
                    selected_row = matches.iloc[0]
                    show_individual = True
                else:
                    st.info(f"Se encontraron **{len(matches)}** coincidencias. Selecciona:")
                    for _, row in matches.iterrows():
                        label = f"[{row['ID']}] {row['Nombre']} — {row['Cat']}"
                        if st.button(label, key=f"sel_{row['ID']}"):
                            st.session_state['selected_crew'] = row['ID']
                            st.rerun()
        
        if show_individual and selected_row is not None:
            row = selected_row
            if 'selected_crew' in st.session_state and selected_crew_id:
                if st.button("← Volver al listado", key="back_btn"):
                    del st.session_state['selected_crew']
                    st.rerun()
            
            # Result card - JC / TCP terminology
            rol_label = "JC" if row['Cat'] == 'JC' else "TCP"
            rol_icon = "🧑‍✈️" if row['Cat'] == 'JC' else "👥"
            st.markdown(f"""
                <div class="result-card">
                    <div style="display: flex; align-items: center; gap: 12px; margin-bottom: 12px;">
                        <span style="font-size: 1.8rem;">{rol_icon}</span>
                        <div>
                            <h3 style="margin: 0; color: #0D5F5D; font-size: 1.4rem; font-weight: 800;">{row['Nombre']}</h3>
                            <span style="color: #64748B; font-size: 0.85rem; font-weight: 500;">ID: {row['ID']}</span>
                        </div>
                    </div>
                    <span class="badge {'badge-jc' if row['Cat'] == 'JC' else 'badge-tc'}">{rol_icon} {rol_label}</span>
                    <span class="badge badge-block">✈️ Block: {row['Block']}</span>
                    <span class="badge badge-duty">⏱️ Duty: {row['Duty']}</span>
                    <span class="badge badge-estab">📊 Estabilidad: {row['Estabilidad']:.1f}%</span>
                </div>
            """, unsafe_allow_html=True)
            
            # Calendario mensual
            st.markdown("#### 📅 Programación del Mes:")
            
            import calendar as cal_mod
            # Usar mes/año de la cabecera del PDF si está disponible
            _h_mes = st.session_state.get('header_mes')
            _h_anio = st.session_state.get('header_anio')
            if _h_mes and _h_anio:
                mes_ref = _h_mes
                anio_ref = _h_anio
            else:
                now = datetime.now()
                anio_ref = now.year
                mes_ref = now.month - 1 if now.month > 1 else 12
                if mes_ref == 0:
                    mes_ref = 12
                    anio_ref -= 1
                for m_try in range(1, 13):
                    if cal_mod.monthrange(anio_ref, m_try)[1] == dias_mes:
                        mes_ref = m_try
                        break
            
            dias_semana = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
            
            try:
                from datetime import date
                primer_dia = date(anio_ref, mes_ref, 1)
                offset = primer_dia.weekday()
            except:
                offset = 0
            
            # Build calendar as HTML - BARRA DE DÍAS MEJORADA
            cal_html = '<div style="display:grid;grid-template-columns:repeat(7,1fr);gap:6px;max-width:100%;margin:0 auto;">'
            
            # Header row - MEJORADO: más visible y atractivo
            for d_idx, d_name in enumerate(dias_semana):
                is_weekend = d_idx >= 5
                bg_color = 'linear-gradient(135deg, #117F7C 0%, #44BABC 100%)' if is_weekend else 'linear-gradient(135deg, #0D5F5D 0%, #117F7C 100%)'
                cal_html += f'''<div style="text-align:center;font-weight:900;color:white;font-size:1rem;padding:12px 6px;
                                background:{bg_color};border-radius:10px;box-shadow:0 3px 10px rgba(17,127,124,0.3);
                                text-transform:uppercase;letter-spacing:1px;">{d_name}</div>'''
            
            # Empty cells for offset
            for _ in range(offset):
                cal_html += '<div style="min-height:90px;"></div>'
            
            # Day cells
            for dia in range(1, dias_mes + 1):
                code = row['exact_sched'].get(dia, '--')
                estilo, icono, desc = get_visual_style(code)
                
                cal_html += f'<div class="day-card {estilo}"><span class="day-number">Día {dia}</span><span class="day-icon">{icono}</span><span class="day-code">{code}</span><span class="day-desc">{desc}</span></div>'
            
            # Fill remaining cells
            total_cells = offset + dias_mes
            remainder = total_cells % 7
            if remainder > 0:
                for _ in range(7 - remainder):
                    cal_html += '<div style="min-height:90px;"></div>'
            
            cal_html += '</div>'
            st.markdown(cal_html, unsafe_allow_html=True)
            
            # Resumen de actividades
            st.markdown("#### 📊 Resumen de Actividades:")
            if hasattr(row, 'code_days') and row['code_days']:
                code_summary_items = sorted(row['code_days'].items(), key=lambda x: -len(x[1]))
                
                summary_html = '<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;">'
                for cd, dias_list in code_summary_items[:12]:
                    if cd not in ['--', '-', '']:
                        desc_code = get_codigo_descripcion(cd)
                        cat_color = get_color_categoria(get_categoria_principal(cd))
                        summary_html += f'<div style="background:white;border-radius:10px;padding:12px;box-shadow:0 3px 12px rgba(0,0,0,0.06);border-left:3px solid {cat_color};text-align:center;transition:transform 0.2s;"><strong style="font-size:1.1rem;color:#0D5F5D;">{cd}</strong><br><span style="font-size:0.73rem;color:#64748B;font-weight:500;">{desc_code[:30]}</span><br><strong style="color:#117F7C;font-size:1rem;">{len(dias_list)} días</strong></div>'
                summary_html += '</div>'
                st.markdown(summary_html, unsafe_allow_html=True)
        
    
    # ═══════════════════════════════════════════════════════════════
    # TAB 3B: CENSO DE EMPLEADOS
    # ═══════════════════════════════════════════════════════════════
    with t3b:
        st.markdown('<p class="section-title">👥 Relación de Empleados en Base</p>', unsafe_allow_html=True)
        
        # KPIs
        kpi_cols = st.columns(5)
        
        total_block = sum(parse_hours(h) for h in df_c['Block'])
        total_duty = sum(parse_hours(h) for h in df_c['Duty'])
        
        with kpi_cols[0]:
            st.markdown(f'<div class="mini-kpi"><h2>{total_plantilla}</h2><p>Total Tripulantes</p></div>', unsafe_allow_html=True)
        with kpi_cols[1]:
            st.markdown(f'<div class="mini-kpi"><h2>{k["jc"]}</h2><p>JC</p></div>', unsafe_allow_html=True)
        with kpi_cols[2]:
            st.markdown(f'<div class="mini-kpi"><h2>{k["tc"]}</h2><p>TCP</p></div>', unsafe_allow_html=True)
        with kpi_cols[3]:
            st.markdown(f'<div class="mini-kpi"><h2>{total_block:.0f}h</h2><p>Total Horas Block</p></div>', unsafe_allow_html=True)
        with kpi_cols[4]:
            st.markdown(f'<div class="mini-kpi"><h2>{total_duty:.0f}h</h2><p>Total Horas Duty</p></div>', unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # ── Calcular horas block medias por rol (sobre dataset completo) ──
        _jc_mask_full = df_c['Cat'] == 'JC'
        _tc_mask_full = df_c['Cat'] == 'TC'
        
        jc_blocks_ini = df_c.loc[_jc_mask_full, 'Block_Inicial'].apply(parse_hours)
        jc_blocks_fin = df_c.loc[_jc_mask_full, 'Block'].apply(parse_hours)
        tcp_blocks_ini = df_c.loc[_tc_mask_full, 'Block_Inicial'].apply(parse_hours)
        tcp_blocks_fin = df_c.loc[_tc_mask_full, 'Block'].apply(parse_hours)
        
        jc_blocks_ini_pos = jc_blocks_ini[jc_blocks_ini > 0]
        jc_blocks_fin_pos = jc_blocks_fin[jc_blocks_fin > 0]
        tcp_blocks_ini_pos = tcp_blocks_ini[tcp_blocks_ini > 0]
        tcp_blocks_fin_pos = tcp_blocks_fin[tcp_blocks_fin > 0]
        
        jc_avg_ini = jc_blocks_ini_pos.mean() if len(jc_blocks_ini_pos) > 0 else 0
        jc_avg_fin = jc_blocks_fin_pos.mean() if len(jc_blocks_fin_pos) > 0 else 0
        tcp_avg_ini = tcp_blocks_ini_pos.mean() if len(tcp_blocks_ini_pos) > 0 else 0
        tcp_avg_fin = tcp_blocks_fin_pos.mean() if len(tcp_blocks_fin_pos) > 0 else 0
        
        def _fmt_hm_rel(hours):
            h = int(hours)
            m = int((hours - h) * 60)
            return f"{h}:{m:02d}h"
        
        jc_ini_str = _fmt_hm_rel(jc_avg_ini)
        jc_fin_str = _fmt_hm_rel(jc_avg_fin)
        tcp_ini_str = _fmt_hm_rel(tcp_avg_ini)
        tcp_fin_str = _fmt_hm_rel(tcp_avg_fin)
        
        # Banner de horas block medias - Tarjetas visuales
        _rel_diff_jc = jc_avg_fin - jc_avg_ini
        _rel_diff_tc = tcp_avg_fin - tcp_avg_ini
        
        def _rel_fmt_diff(val):
            sign = "+" if val > 0.001 else ("" if val < -0.001 else "")
            h = int(abs(val))
            m = int((abs(val) - h) * 60)
            prefix = "-" if val < -0.001 else sign
            return f"{prefix}{h}:{m:02d}h"
        
        _rel_diff_jc_str = _rel_fmt_diff(_rel_diff_jc)
        _rel_diff_tc_str = _rel_fmt_diff(_rel_diff_tc)
        # Colores suaves para la diferencia, acordes al banner (blanco sobre fondo tintado)
        _rel_diff_jc_col = "rgba(255,255,255,0.95)"
        _rel_diff_tc_col = "rgba(255,255,255,0.95)"
        _rel_diff_jc_bg = "rgba(255,255,255,0.22)" if abs(_rel_diff_jc) > 0.001 else "rgba(255,255,255,0.12)"
        _rel_diff_tc_bg = "rgba(255,255,255,0.22)" if abs(_rel_diff_tc) > 0.001 else "rgba(255,255,255,0.12)"
        
        _col_bm_r1, _col_bm_r2 = st.columns(2)
        with _col_bm_r1:
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #117F7C 0%, #0B847F 100%);
                        padding: 20px 22px; border-radius: 14px; color: white;
                        box-shadow: 0 4px 16px rgba(17,127,124,0.25); margin: 12px 0;">
                <div style="font-size: 0.82rem; font-weight: 700; letter-spacing: 1px; opacity: 0.85; margin-bottom: 12px;">
                    👤 JCS — Block Medio
                </div>
                <div style="display: flex; gap: 20px; align-items: flex-end; flex-wrap: wrap;">
                    <div>
                        <div style="font-size: 0.72rem; opacity: 0.7;">Prog. Inicial</div>
                        <div style="font-size: 1.7rem; font-weight: 800; line-height: 1.1;">{jc_ini_str}</div>
                    </div>
                    <div style="font-size: 1.2rem; opacity: 0.5; padding-bottom: 3px;">→</div>
                    <div>
                        <div style="font-size: 0.72rem; opacity: 0.7;">Prog. Final</div>
                        <div style="font-size: 1.7rem; font-weight: 800; line-height: 1.1;">{jc_fin_str}</div>
                    </div>
                    <div style="margin-left: auto; text-align: right;">
                        <div style="font-size: 0.68rem; opacity: 0.7;">Diferencia</div>
                        <div style="font-size: 1.1rem; font-weight: 700; color: {_rel_diff_jc_col};
                                    background: {_rel_diff_jc_bg}; padding: 3px 12px; border-radius: 18px;">
                            {_rel_diff_jc_str}
                        </div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
        with _col_bm_r2:
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #489999 0%, #44BABC 100%);
                        padding: 20px 22px; border-radius: 14px; color: white;
                        box-shadow: 0 4px 16px rgba(68,186,188,0.25); margin: 12px 0;">
                <div style="font-size: 0.82rem; font-weight: 700; letter-spacing: 1px; opacity: 0.85; margin-bottom: 12px;">
                    👥 TCPS — Block Medio
                </div>
                <div style="display: flex; gap: 20px; align-items: flex-end; flex-wrap: wrap;">
                    <div>
                        <div style="font-size: 0.72rem; opacity: 0.7;">Prog. Inicial</div>
                        <div style="font-size: 1.7rem; font-weight: 800; line-height: 1.1;">{tcp_ini_str}</div>
                    </div>
                    <div style="font-size: 1.2rem; opacity: 0.5; padding-bottom: 3px;">→</div>
                    <div>
                        <div style="font-size: 0.72rem; opacity: 0.7;">Prog. Final</div>
                        <div style="font-size: 1.7rem; font-weight: 800; line-height: 1.1;">{tcp_fin_str}</div>
                    </div>
                    <div style="margin-left: auto; text-align: right;">
                        <div style="font-size: 0.68rem; opacity: 0.7;">Diferencia</div>
                        <div style="font-size: 1.1rem; font-weight: 700; color: {_rel_diff_tc_col};
                                    background: {_rel_diff_tc_bg}; padding: 3px 12px; border-radius: 18px;">
                            {_rel_diff_tc_str}
                        </div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        # ── Buscador por ID de tripulante (debajo de banners) ──
        st.markdown("#### 🔎 Buscador por ID de Tripulante")
        
        # Callback para limpiar la búsqueda
        def _clear_crew_search():
            st.session_state['search_crew_id'] = ""
        
        col_search, col_btn = st.columns([4, 1])
        with col_search:
            search_id = st.text_input(
                "Buscar tripulante",
                placeholder="Introduce ID o Nombre",
                key="search_crew_id",
                label_visibility="collapsed"
            )
        with col_btn:
            st.button("🗑️ Limpiar", key="btn_clear_crew_search", on_click=_clear_crew_search)
        
        # ── Filtros ──
        filter_cols = st.columns([2, 2, 4])
        with filter_cols[0]:
            cat_filter = st.selectbox("Filtrar por rol:", ["Todos", "JC", "TCP"])
        with filter_cols[1]:
            sort_by = st.selectbox("Ordenar por:", ["ID", "Nombre", "Horas Block", "Horas Duty", "Estabilidad"])
        
        df_display = df_c.copy()
        # Añadir columnas numéricas para ordenamiento
        df_display['Horas_Block_Inicial_Num'] = df_display['Block_Inicial'].apply(parse_hours)
        df_display['Horas_Block_Num'] = df_display['Block'].apply(parse_hours)
        df_display['Horas_Duty_Num'] = df_display['Duty'].apply(parse_hours)
        
        if cat_filter == "JC":
            df_display = df_display[df_display['Cat'] == 'JC']
        elif cat_filter == "TCP":
            df_display = df_display[df_display['Cat'] == 'TC']
        
        # Ordenamiento
        sort_map = {
            "ID": ("ID", True), "Nombre": ("Nombre", True),
            "Horas Block": ("Horas_Block_Num", False),
            "Horas Duty": ("Horas_Duty_Num", False),
            "Estabilidad": ("Estabilidad", False)
        }
        sort_col, sort_asc = sort_map[sort_by]
        df_display = df_display.sort_values(sort_col, ascending=sort_asc)
        
        # Aplicar búsqueda por ID o Nombre
        if search_id and search_id.strip():
            _q = search_id.strip()
            _mask = (
                df_display['ID'].astype(str).str.contains(_q, case=False, na=False, regex=False)
                | df_display['Nombre'].astype(str).str.contains(_q, case=False, na=False, regex=False)
            )
            df_display = df_display[_mask]
            if df_display.empty:
                st.warning(f"⚠️ No se encontró ningún tripulante que contenga '{_q}'")
        
        # Tabla HTML - CON DOS COLUMNAS HORAS BLOCK (Inicial, Final), DIFERENCIA y CAMBIOS
        rows_html = []
        for _, row in df_display.iterrows():
            estab = row['Estabilidad']
            progress_class = "progress-green" if estab >= 85 else ("progress-yellow" if estab >= 70 else "progress-red")
            cat_badge = "badge-jc" if row['Cat'] == 'JC' else "badge-tc"
            rol_icon = "🧑‍✈️" if row['Cat'] == 'JC' else "👥"
            rol_text = "JC" if row['Cat'] == 'JC' else "TCP"
            horas_block_inicial = row.get('Horas_Block_Inicial_Num', parse_hours(row['Block_Inicial']))
            horas_block_final = row.get('Horas_Block_Num', parse_hours(row['Block']))
            
            # Calcular diferencia con precisión de minutos
            diferencia = horas_block_final - horas_block_inicial
            diff_text = format_diff_hm(diferencia)
            if diferencia > 0.001:
                diff_color = "#27AE60"  # Verde
            elif diferencia < -0.001:
                diff_color = "#FF9800"  # Naranja suave (antes rojo chillón)
            else:
                diff_color = "#94A3B8"  # Gris
            
            # Formato visual círculo para Cambios
            cambios = int(row.get('Cambios', 0))
            if cambios == 0:
                _c_bg = "#4CAF50"  # Verde
            elif cambios <= 5:
                _c_bg = "#FFC107"  # Amarillo
            elif cambios <= 10:
                _c_bg = "#FF9800"  # Naranja
            else:
                _c_bg = "#FF6B6B"  # Rojo suave
            _c_txt_col = "#fff" if cambios > 0 else "#fff"
            cambios_html = (
                f"<div style='display:inline-flex;align-items:center;justify-content:center;"
                f"background:{_c_bg};color:{_c_txt_col};border-radius:50%;width:36px;height:36px;"
                f"font-weight:700;font-size:0.85rem;box-shadow:0 2px 6px rgba(0,0,0,0.15);'>"
                f"{cambios}</div>"
            )
            
            row_html = (
                f"<tr>"
                f"<td><strong>{row['ID']}</strong></td>"
                f"<td>{row['Nombre']}</td>"
                f"<td><span class='badge {cat_badge}'>{rol_icon} {rol_text}</span></td>"
                f"<td style='color:#6B7280; font-weight:600;'>{row['Block_Inicial']}</td>"
                f"<td style='color:#117F7C; font-weight:700;'>{row['Block']}</td>"
                f"<td style='color:{diff_color}; font-weight:700;'>{diff_text}</td>"
                f"<td><strong style='color:#0D5F5D'>{row['Duty']}</strong></td>"
                f"<td>{cambios_html}</td>"
                f"<td><div class='progress-bar'><div class='progress-fill {progress_class}' style='width:{estab}%'></div></div><small style='font-weight:700;'>{estab:.1f}%</small></td>"
                f"</tr>"
            )
            rows_html.append(row_html)
        
        table_html = (
            "<div class='censo-table-wrapper'><table class='censo-table'>"
            "<tr><th>ID</th><th>Nombre</th><th>Rol</th><th>Block Inicial</th><th>Block Final</th><th>Diferencia</th><th>Duty</th><th>Cambios</th><th>Estabilidad</th></tr>"
            + "".join(rows_html)
            + "</table></div>"
        )
        st.markdown(table_html, unsafe_allow_html=True)

    
    # ═══════════════════════════════════════════════════════════════
    # TAB LONG DUTIES: ROTACIONES SUPERIORES A 9H 40MIN
    # ═══════════════════════════════════════════════════════════════
    with t_ld:
        st.markdown('<p class="section-title">⏰ Long Duties - Rotaciones Superiores a 9h 40min</p>', unsafe_allow_html=True)
        
        # Construir DataFrame de Long Duties
        # ═══ FIX: Filtro estricto por base y exclusión de Desconocidos ═══
        # Solo incluir tripulantes que:
        # 1. Existen en df_c (procesados desde el PDF final de la base seleccionada)
        # 2. O existen en r_f (roster final) 
        # 3. NUNCA incluir "DESCONOCIDO" (tripulantes de otras bases o no identificados)
        ld_rows = []
        if long_duty_data:
            for cid, ld_info in long_duty_data.items():
                if cid == '__meta__':
                    continue
                long_days = ld_info.get('long_duty_days', [])
                if not long_days:
                    continue
                
                # Buscar nombre y rol desde df_c
                crew_row = df_c[df_c['ID'] == cid]
                if not crew_row.empty:
                    nombre = crew_row.iloc[0]['Nombre']
                    rol = crew_row.iloc[0]['Cat']
                else:
                    # Buscar desde rosters
                    if cid in r_f and cid != '__meta__':
                        nombre = r_f[cid].get('Name', 'DESCONOCIDO')
                        rol = r_f[cid].get('Categoria', '--')
                    else:
                        nombre = 'DESCONOCIDO'
                        rol = '--'
                
                # ═══ FIX: Excluir tripulantes "DESCONOCIDO" ═══
                # Si el nombre es DESCONOCIDO, el tripulante no existe en el PDF
                # de la base seleccionada → es de otra base → NO mostrar
                if nombre == 'DESCONOCIDO' or nombre.strip() == '':
                    continue
                
                # Formatear fechas afectadas
                meta = r_f.get('__meta__', {})
                h_mes = meta.get('header_mes')
                h_anio = meta.get('header_anio')
                
                fechas_str_list = []
                for ld_entry in long_days:
                    day_num = ld_entry[0]
                    duty_h = ld_entry[1]
                    ruta_str = ld_entry[2] if len(ld_entry) > 2 else ''
                    if h_mes and h_anio:
                        try:
                            fecha_str = f"{day_num:02d}/{h_mes:02d}/{h_anio}"
                        except:
                            fecha_str = f"Día {day_num}"
                    else:
                        fecha_str = f"Día {day_num}"
                    ruta_part = f" ({ruta_str})" if ruta_str else ""
                    fechas_str_list.append(f"{fecha_str} ({format_duration(duty_h)}){ruta_part}")
                
                ld_rows.append({
                    'ID': cid,
                    'Nombre': nombre,
                    'Rol': rol,
                    'Fechas_Afectadas': ', '.join(fechas_str_list),
                    'N_Dias': len(long_days),
                    '_long_days_raw': long_days,
                })
        
        if ld_rows:
            df_ld = pd.DataFrame(ld_rows).sort_values('N_Dias', ascending=False)
            
            n_afectados = len(df_ld)
            pct_plantilla = (n_afectados / max(total_plantilla, 1)) * 100
            total_dias_ld = df_ld['N_Dias'].sum()
            
            # --- KPI CARDS ---
            n_jc_ld = len(df_ld[df_ld['Rol'] == 'JC'])
            n_tcp_ld = len(df_ld[df_ld['Rol'] == 'TC'])
            
            kpi_cols = st.columns(4)
            kpi_data = [
                ("📊 Plantilla afectada", f"{pct_plantilla:.1f}%", f"{n_afectados} de {total_plantilla}"),
                ("📅 Total días >9h40min", f"{total_dias_ld}", "días acumulados"),
                ("🧑‍✈️ JC afectados", f"{n_jc_ld}", f"Jefes de Cabina"),
                ("👥 TCP afectados", f"{n_tcp_ld}", f"tripulantes de cabina"),
            ]
            for col, (title, value, subtitle) in zip(kpi_cols, kpi_data):
                with col:
                    st.markdown(f"""
                    <div style="background: white; border-radius: 12px; padding: 16px; text-align: center;
                                box-shadow: 0 2px 8px rgba(0,0,0,0.08); border-top: 3px solid #0D9488;">
                        <div style="font-size: 0.8rem; color: #64748B; font-weight: 600;">{title}</div>
                        <div style="font-size: 1.8rem; font-weight: 800; color: #0D5F5D; margin: 4px 0;">{value}</div>
                        <div style="font-size: 0.75rem; color: #94A3B8;">{subtitle}</div>
                    </div>
                    """, unsafe_allow_html=True)
            
            st.markdown("<br>", unsafe_allow_html=True)
            
            # --- FILTROS INTERACTIVOS ---
            filter_cols = st.columns(2)
            with filter_cols[0]:
                filtro_rol_ld = st.selectbox(
                    "Filtrar por rol:",
                    ["Todos", "JC", "TCP"],
                    key="filtro_rol_ld"
                )
            with filter_cols[1]:
                orden_ld = st.selectbox(
                    "Ordenar por:",
                    ["ID", "Chequeo", "Long Duties", "Mismos Long Duties (tripulantes)"],
                    key="orden_ld"
                )
            
            # Aplicar filtro de rol
            df_ld_filtered = df_ld.copy()
            if filtro_rol_ld == "JC":
                df_ld_filtered = df_ld_filtered[df_ld_filtered['Rol'] == 'JC']
            elif filtro_rol_ld == "TCP":
                df_ld_filtered = df_ld_filtered[df_ld_filtered['Rol'] == 'TC']
            
            # Aplicar orden
            if orden_ld == "ID":
                df_ld_filtered = df_ld_filtered.sort_values('ID', ascending=True)
            elif orden_ld == "Chequeo":
                df_ld_filtered = df_ld_filtered.sort_values('Nombre', ascending=True)
            elif orden_ld == "Long Duties":
                df_ld_filtered = df_ld_filtered.sort_values('N_Dias', ascending=False)
            elif orden_ld == "Mismos Long Duties (tripulantes)":
                # ═══ IMPLEMENTACIÓN: Agrupar TODOS los tripulantes por RUTA ESPECÍFICA ═══
                # Objetivo: Ver la dotación completa del avión en cada ruta de long duty.
                # Incluye TODOS los tripulantes (TCP + JC) independientemente del filtro de rol.
                # Ejemplo:  RUTA BCN-DSS-BCN (Día 15): JC García (10h55), TCP López (10h50), TCP Ruiz (10h45)
                
                # 1. Construir índice usando df_ld COMPLETO (sin filtro de rol)
                #    para mostrar siempre la dotación completa del avión
                def normalize_route_for_grouping(ruta_str):
                    """Normaliza ruta para agrupar JCs y TCPs del mismo vuelo.
                    
                    NORMALIZACIÓN ESTRICTA: Solo limpia diferencias cosméticas,
                    NUNCA elimina información que diferencia rutas reales.
                    
                    Operaciones permitidas:
                    1. Quitar emojis (🌙)
                    2. Quitar prefijo * de aeropuertos (indicador de escala)
                    3. Filtrar códigos no-aeropuerto (TBD, etc.)
                    4. Normalizar a mayúsculas
                    
                    NO se eliminan duplicados consecutivos porque representan
                    tramos reales de ida y vuelta (ej: BIO-LHR-BIO-VLC-BIO).
                    """
                    r = ruta_str.replace('🌙', '').strip()
                    # Quitar * de los códigos de aeropuerto
                    segments = r.split('-')
                    clean_segments = []
                    for s in segments:
                        s = s.replace('*', '').strip().upper()
                        if not s:
                            continue
                        # Filtrar códigos no-aeropuerto (TBD, etc.)
                        if s in ('TBD', 'UNK', 'XXX', 'N/A'):
                            continue
                        # Solo incluir segmentos que parezcan códigos IATA (3 letras)
                        if len(s) == 3 and s.isalpha():
                            clean_segments.append(s)
                    # NO eliminar consecutivos duplicados - representan tramos reales
                    return '-'.join(clean_segments) if clean_segments else r
                
                route_day_groups = {}
                for _, row in df_ld.iterrows():
                    raw = row.get('_long_days_raw', [])
                    if not raw:
                        continue
                    for ld_entry in raw:
                        day_num = ld_entry[0]
                        duty_h = ld_entry[1]
                        ruta = ld_entry[2] if len(ld_entry) > 2 else ''
                        if not ruta:
                            ruta = 'Ruta no disponible'
                        # Normalizar ruta para agrupar (JCs y TCPs del mismo vuelo)
                        ruta_clean = normalize_route_for_grouping(ruta)
                        key = (ruta_clean, day_num)
                        if key not in route_day_groups:
                            route_day_groups[key] = []
                        route_day_groups[key].append({
                            'ID': row['ID'],
                            'Nombre': row['Nombre'],
                            'Rol': row['Rol'],
                            'Horas': duty_h,
                            'Ruta_Display': ruta  # Con 🌙 si aplica
                        })
                
                # 2. Mostrar agrupación por ruta (TODAS las rutas, sin filtro mínimo)
                meta_ld = r_f.get('__meta__', {})
                h_mes_ld = meta_ld.get('header_mes')
                h_anio_ld = meta_ld.get('header_anio')
                
                groups_html_parts = []
                for (ruta, day_num), crew_list in sorted(route_day_groups.items(), key=lambda x: (-len(x[1]), x[0][1], x[0][0])):
                    
                    if h_mes_ld and h_anio_ld:
                        try:
                            fecha_label = f"{day_num:02d}/{h_mes_ld:02d}/{h_anio_ld}"
                        except:
                            fecha_label = f"Día {day_num}"
                    else:
                        fecha_label = f"Día {day_num}"
                    
                    # Contar JCs y TCPs en esta ruta
                    n_jc = sum(1 for c in crew_list if c['Rol'] == 'JC')
                    n_tcp = sum(1 for c in crew_list if c['Rol'] != 'JC')
                    dotacion_info = []
                    if n_jc > 0:
                        dotacion_info.append(f"🧑‍✈️ {n_jc} JC")
                    if n_tcp > 0:
                        dotacion_info.append(f"👥 {n_tcp} TCP")
                    dotacion_str = " + ".join(dotacion_info) if dotacion_info else ""
                    
                    crew_rows = ""
                    # Ordenar: primero JCs, luego TCPs, dentro de cada grupo por horas desc
                    crew_sorted = sorted(crew_list, key=lambda x: (0 if x['Rol'] == 'JC' else 1, -x['Horas']))
                    for c in crew_sorted:
                        rol_badge = "badge-jc" if c['Rol'] == 'JC' else "badge-tc"
                        rol_icon = "🧑‍✈️" if c['Rol'] == 'JC' else "👥"
                        rol_text = "JC" if c['Rol'] == 'JC' else "TCP"
                        crew_rows += (
                            f"<tr>"
                            f"<td><strong>{c['ID']}</strong></td>"
                            f"<td>{c['Nombre']}</td>"
                            f"<td><span class='badge {rol_badge}'>{rol_icon} {rol_text}</span></td>"
                            f"<td style='text-align:center; font-weight:700; color:#0D5F5D;'>{format_duration(c['Horas'])}</td>"
                            f"</tr>"
                        )
                    
                    groups_html_parts.append(f"""
                    <div style="background:white; border-radius:12px; padding:16px 20px; margin:12px 0;
                                box-shadow:0 2px 8px rgba(0,0,0,0.08); border-left:4px solid #0D9488;">
                        <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:12px;">
                            <div>
                                <span style="font-size:1.1rem; font-weight:800; color:#0D5F5D;">✈️ {ruta}</span>
                                <span style="margin-left:12px; background:#E0F2F1; color:#0D5F5D; padding:4px 12px;
                                       border-radius:20px; font-size:0.85rem; font-weight:600;">📅 {fecha_label}</span>
                            </div>
                            <span style="background:linear-gradient(135deg,#0D9488,#0D5F5D); color:white; padding:4px 14px;
                                   border-radius:20px; font-size:0.9rem; font-weight:700;">
                                {len(crew_list)} trip. ({dotacion_str})
                            </span>
                        </div>
                        <table class="censo-table" style="margin:0;">
                            <tr><th>ID</th><th>Nombre</th><th>Rol</th><th>Duty</th></tr>
                            {crew_rows}
                        </table>
                    </div>
                    """)
                
                if groups_html_parts:
                    st.markdown("".join(groups_html_parts), unsafe_allow_html=True)
                else:
                    st.info("ℹ️ No se encontraron rutas de long duty en este período.")
                
                # Mantener también la tabla normal debajo
                df_ld_filtered = df_ld_filtered.sort_values('N_Dias', ascending=False)
            
            # Banner informativo sobre el umbral
            st.markdown("""
            <div style="
                background: linear-gradient(135deg, rgba(17, 127, 124, 0.08) 0%, rgba(68, 186, 188, 0.08) 100%);
                border: 1px solid rgba(17, 127, 124, 0.3);
                border-left: 4px solid #117F7C;
                border-radius: 10px;
                padding: 16px 20px;
                margin: 15px 0 20px 0;
                display: flex;
                align-items: flex-start;
                gap: 14px;
                box-shadow: 0 2px 4px rgba(0,0,0,0.05);
            ">
                <span style="font-size: 1.4rem; flex-shrink: 0; margin-top: 2px;">⚠️</span>
                <div style="flex: 1;">
                    <div style="color: #117F7C; font-weight: 700; font-size: 0.95rem; margin-bottom: 8px; letter-spacing: 0.3px;">
                        RECUERDA: AJUSTE DE TIEMPOS POST-VUELO
                    </div>
                    <div style="color: #2c5f5d; font-weight: 400; font-size: 0.9rem; line-height: 1.5;">
                        Al total de horas volcado por el sistema, se deben sumar 20 minutos tras calzos en los casos que corresponda. 
                        El sistema registra la actividad de inicio a fin, pero excluye este periodo adicional obligatorio.
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            st.markdown("<br>", unsafe_allow_html=True)
            
            # Tabla de Long Duties con chips grandes para fechas
            rows_ld_html = []
            for idx, (_, row) in enumerate(df_ld_filtered.iterrows()):
                cat_badge = "badge-jc" if row['Rol'] == 'JC' else "badge-tc"
                rol_icon = "🧑‍✈️" if row['Rol'] == 'JC' else "👥"
                rol_text = "JC" if row['Rol'] == 'JC' else "TCP"
                severity_color = '#DC2626' if row['N_Dias'] > 3 else ('#EA580C' if row['N_Dias'] > 1 else '#F59E0B')
                
                # Crear chips grandes para cada fecha
                raw_days = row['_long_days_raw'] if '_long_days_raw' in row.index else []
                fecha_chips = []
                meta = r_f.get('__meta__', {})
                h_mes = meta.get('header_mes')
                h_anio = meta.get('header_anio')
                for i, ld_entry in enumerate(raw_days):
                    day_num = ld_entry[0]
                    duty_h = ld_entry[1]
                    ruta_str = ld_entry[2] if len(ld_entry) > 2 else ''
                    if h_mes and h_anio:
                        try:
                            fecha_label = f"{day_num:02d}/{h_mes:02d}"
                        except:
                            fecha_label = f"D{day_num}"
                    else:
                        fecha_label = f"D{day_num}"
                    # Color según severidad de horas
                    if duty_h >= 12:
                        chip_bg = '#DC2626'; chip_border = '#B91C1C'
                    elif duty_h >= 11:
                        chip_bg = '#EA580C'; chip_border = '#C2410C'
                    else:
                        chip_bg = '#0D9488'; chip_border = '#0F766E'
                    ruta_chip = f" &middot; {ruta_str}" if ruta_str else ""
                    fecha_chips.append(
                        f"<span style='display:inline-block; background:{chip_bg}; color:white; "
                        f"padding:6px 14px; border-radius:20px; font-size:0.88rem; font-weight:700; "
                        f"margin:3px 4px; white-space:nowrap; letter-spacing:0.3px; "
                        f"border:2px solid {chip_border}; "
                        f"box-shadow:0 1px 3px rgba(0,0,0,0.15);'>"
                        f"📅 {fecha_label} &middot; {format_duration(duty_h)}{ruta_chip}</span>"
                    )
                fechas_html = " ".join(fecha_chips)
                
                row_html = (
                    f"<tr>"
                    f"<td><strong>{row['ID']}</strong></td>"
                    f"<td>{row['Nombre']}</td>"
                    f"<td><span class='badge {cat_badge}'>{rol_icon} {rol_text}</span></td>"
                    f"<td>{fechas_html}</td>"
                    f"<td style='text-align:center;'>"
                    f"<span style='display:inline-block; background:linear-gradient(135deg, {severity_color}, {severity_color}dd); "
                    f"color:white; min-width:38px; height:38px; line-height:38px; border-radius:50%; "
                    f"font-weight:800; font-size:1.1rem; text-align:center; "
                    f"box-shadow:0 2px 6px {severity_color}55;'>"
                    f"{row['N_Dias']}</span></td>"
                    f"</tr>"
                )
                rows_ld_html.append(row_html)
            
            table_ld_html = (
                "<div class='censo-table-wrapper'><table class='censo-table'>"
                "<tr><th>ID</th><th>Nombre</th><th>Rol</th><th>Fechas Afectadas</th><th>Nº Días &gt;9h40</th></tr>"
                + "".join(rows_ld_html)
                + "</table></div>"
            )
            st.markdown(table_ld_html, unsafe_allow_html=True)
            
            st.markdown("<br>", unsafe_allow_html=True)
            
            # === BOTONES DE EXPORTACIÓN ===
            # --- EXPORTAR A EXCEL ---
            try:
                buffer_ld_excel = io.BytesIO()
                wb_ld = openpyxl.Workbook()
                ws_ld = wb_ld.active
                ws_ld.title = "Long Duties"
                
                # Estilos
                header_fill_ld = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
                header_font_ld = Font(color="FFFFFF", bold=True, size=11)
                totals_fill_ld = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
                totals_font_ld = Font(bold=True, size=11)
                border_ld = Border(
                    left=Side(style='thin', color='B2DFDB'),
                    right=Side(style='thin', color='B2DFDB'),
                    top=Side(style='thin', color='B2DFDB'),
                    bottom=Side(style='thin', color='B2DFDB')
                )
                
                headers_ld = ['ID', 'Nombre', 'Rol', 'Fechas Afectadas', 'Nº Días >9h40']
                for col, header in enumerate(headers_ld, 1):
                    cell = ws_ld.cell(row=1, column=col, value=header)
                    cell.fill = header_fill_ld
                    cell.font = header_font_ld
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border_ld
                
                for row_idx, (_, row) in enumerate(df_ld_filtered.iterrows(), 2):
                    ws_ld.cell(row=row_idx, column=1, value=row['ID']).border = border_ld
                    ws_ld.cell(row=row_idx, column=2, value=row['Nombre']).border = border_ld
                    rol_val = 'JC' if row['Rol'] == 'JC' else 'TCP'
                    ws_ld.cell(row=row_idx, column=3, value=rol_val).border = border_ld
                    ws_ld.cell(row=row_idx, column=4, value=row['Fechas_Afectadas']).border = border_ld
                    ws_ld.cell(row=row_idx, column=5, value=row['N_Dias']).border = border_ld
                    ws_ld.cell(row=row_idx, column=5).alignment = Alignment(horizontal='center')
                
                # Fila de totales
                total_row = len(df_ld_filtered) + 2
                ws_ld.cell(row=total_row, column=1, value='TOTAL')
                ws_ld.cell(row=total_row, column=2, value=f'{n_afectados} tripulantes ({pct_plantilla:.1f}% plantilla)')
                ws_ld.cell(row=total_row, column=5, value=total_dias_ld)
                for col in range(1, 6):
                    cell = ws_ld.cell(row=total_row, column=col)
                    cell.fill = totals_fill_ld
                    cell.font = totals_font_ld
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border_ld
                
                # Ajustar anchos
                ws_ld.column_dimensions['A'].width = 10
                ws_ld.column_dimensions['B'].width = 25
                ws_ld.column_dimensions['C'].width = 8
                ws_ld.column_dimensions['D'].width = 55
                ws_ld.column_dimensions['E'].width = 15
                
                wb_ld.save(buffer_ld_excel)
                excel_ld_data = buffer_ld_excel.getvalue()
                
                st.download_button(
                    "📊 Descargar Excel",
                    data=excel_ld_data,
                    file_name=f"long_duties_{base_procesada}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            except Exception as e:
                st.error(f"Error generando Excel: {e}")
        else:
            # No hay long duties
            st.markdown("""
            <div style="text-align: center; padding: 40px; background: #F0FDF4; border-radius: 16px; border: 2px solid #86EFAC;">
                <h2 style="color: #16A34A; margin: 0;">✅ Sin Long Duties detectados</h2>
                <p style="color: #15803D; margin-top: 10px;">No se han encontrado rotaciones superiores a 9h 40min en esta programación.</p>
                <small style="color: #94A3B8;">Esto puede deberse a que el PDF procesado no contiene datos detallados de tiempos por día, 
                o a que efectivamente no hay duties que superen las 9h 40min.</small>
            </div>
            """, unsafe_allow_html=True)

    # ═══════════════════════════════════════════════════════════════
    # TAB ROTACIONES LARGAS >5 DÍAS  (ACTIVIDAD CONTINUA)
    # ═══════════════════════════════════════════════════════════════
    with t_rot:
        st.markdown('<p class="section-title">⚠️ Actividad Continua >5 Días — Tripulantes con más de 5 días seguidos de actividad</p>', unsafe_allow_html=True)

        # ── Códigos que representan días libres (cortan la rotación) ──
        _OFF_CODES = {'OFF', 'SROF', 'ROFF', 'NOFF', 'NROF', 'EOFF', 'FOFF', 'DOFF', 'SOFF',
                      'REST', 'RRES', 'LFR', 'RF', 'EF', 'XOF1', 'XOF2', 'XOF3', 'XSOF', 'CDOF', 'DEOF'}
        # NOTA: F y F2 NO están en _OFF_CODES — son "Disponibilidad Empresa" (recurso), NO descanso.
        # ── Códigos que cortan la rotación (no-actividad) ──
        _EXCLUDE_CODES = {'VAC', 'RVAC', 'NVAC', 'PVAC',
                          'PER', 'RPER', 'BOD', 'MUD', 'RPUR',
                          'PAT', 'MAT', 'EMB', 'LAC', 'ELAC', 'ELAT', 'PPA', 'PPAI',
                          'PT', 'PTF', 'PTD', 'PTI',
                          'SICK', 'SICD', 'LSIC', 'LSCK', 'NJSK',
                          'MED', 'ASEP', 'UNFT', 'UNFD', 'FTG',
                          'ILT', 'BAJA', 'ENF', 'RVCC',
                          'NQ', 'NANQ', 'NAVL'}
        # ── Códigos FW: CORTAN la rotación si aparecen solos (se consideran OFF) ──
        # Si un día tiene SOLO FW → es OFF (breaker). Si FW + actividad → cuenta.
        _FW_CODES = {'FW', 'FLEX', 'FLEXIWORKING'}
        _NON_ACTIVITY_CODES = _OFF_CODES | _EXCLUDE_CODES
        # ── Patrón horario HH:MM (ej: 3:00, 18:35) - artefacto del PDF ──
        _re_time_pattern = re.compile(r'^\d{1,2}:\d{2}$')

        def _is_blank_day(code):
            """Determina si un día está en blanco (sin codificación real)."""
            if not code:
                return True
            c = code.strip()
            if c in ('--', '-', '', 'None'):
                return True
            return False

        def _is_time_token(code):
            """Detecta si un token es un horario HH:MM (artefacto del PDF, no un código real).
            Ejemplos: '3:00', '18:35', '12:50'
            """
            if not code:
                return False
            return bool(_re_time_pattern.match(code.strip()))

        def _is_rotation_breaker(code):
            """
            Determina si un código CORTA la rotación de actividad.
            
            CORTAN (True):  OFF y variantes, VAC, SICK, NQ, FW solo, días en blanco,
                            y demás códigos de exclusión.
            NO CORTAN (False): Vuelos, SBY, actividad, horarios HH:MM.
            
            CAMBIO CRÍTICO (Abril 2026): Los días en blanco AHORA SÍ cortan la rotación.
            Un día sin código asignado significa que el tripulante NO tiene actividad ese día,
            por lo que la secuencia de actividad continuada se interrumpe.
            Caso verificado: Tripulante 14594 en Marzo - días 21-24 en blanco deben
            romper la secuencia entre días 16-20 y 25-29.
            """
            if _is_blank_day(code):
                return True   # Blancos SÍ cortan — sin código = sin actividad = ruptura
            c = code.strip().upper()
            if c in _OFF_CODES:
                return True   # OFF y variantes SÍ cortan
            if c in _EXCLUDE_CODES:
                return True   # VAC, SICK, NQ, etc. SÍ cortan
            if c in _FW_CODES:
                return True   # FW solo SÍ corta (se considera OFF si no hay más actividad ese día)
            if _is_time_token(c):
                return False  # Horarios HH:MM NO cortan (artefacto PDF, indica presentación)
            return False      # Todo lo demás (vuelos, SBY, etc.) NO corta

        def _is_activity_day(code):
            """
            Determina si un código cuenta como actividad real.
            
            ✅ SÍ cuenta: vuelos (4 dígitos), SBY, IMAG, OCC, PQP, ART, horarios HH:MM, etc.
            ❌ NO cuenta: OFF y variantes, VAC, SICK, NQ, FW solo, días en blanco
            
            Nota: Horarios HH:MM (ej: '3:00') se cuentan como actividad porque indican
            que el tripulante tiene hora de presentación ese día (actividad en líneas detalle del PDF).
            """
            if _is_blank_day(code):
                return False
            c = code.strip().upper()
            if c in _OFF_CODES:
                return False
            if c in _EXCLUDE_CODES:
                return False
            if c in _FW_CODES:
                # FW solo = NO actividad directa, pero NO corta rotación
                return False
            # Horarios HH:MM = actividad (indica presentación / standby con hora)
            if _is_time_token(c):
                return True
            return True
        
        meta_rot = r_f.get('__meta__', {})
        _h_mes_rot = meta_rot.get('header_mes')
        _h_anio_rot = meta_rot.get('header_anio')
        
        # ── Recopilar períodos de actividad continua por tripulante ──
        crew_periods = {}  # {cid: {'nombre': str, 'cat': str, 'periods': [{'start': int, 'end': int, 'days': int}]}}
        
        for _, row in df_c.iterrows():
            cid = row['ID']
            nombre = row['Nombre']
            cat = row['Cat']
            sched = row.get('exact_sched', {})
            
            # ── Paso 1: Clasificar cada día ──
            day_types = {}  # {day: 'activity'|'breaker'|'neutral'}
            for d in range(1, dias_mes + 1):
                code = sched.get(d, '--').strip().upper()
                if _is_rotation_breaker(code):
                    day_types[d] = 'breaker'
                elif _is_activity_day(code):
                    day_types[d] = 'activity'
                else:
                    # Neutral: blanco - no rompe rotación pero no es actividad directa
                    # (FW ahora es breaker, así que no llega aquí)
                    day_types[d] = 'neutral'
            
            # ── Paso 2: Detectar períodos de actividad continua ──
            # Regla: Desde el último OFF/BLANK/EXCLUDE hasta el siguiente OFF/BLANK/EXCLUDE.
            # El período va desde el PRIMER día de actividad tras un breaker
            # hasta el DÍA ANTERIOR al siguiente breaker.
            # CAMBIO Abril 2026: Días en blanco AHORA cortan la rotación (son breakers).
            # Solo horarios HH:MM (artefactos PDF) se tratan como neutral.
            rotaciones = []
            period_start = None       # Primer día de actividad del período actual
            last_activity_day = None  # Último día de actividad confirmada
            
            for d in range(1, dias_mes + 1):
                dtype = day_types[d]
                
                if dtype == 'breaker':
                    # OFF/VAC/EXCLUDE → cierra el período activo
                    if period_start is not None and last_activity_day is not None:
                        # SOLO contar días con actividad REAL, NO días en blanco/neutral
                        period_days = [dd for dd in range(period_start, last_activity_day + 1)
                                       if day_types[dd] == 'activity']
                        rotaciones.append(period_days)
                    period_start = None
                    last_activity_day = None
                elif dtype == 'activity':
                    if period_start is None:
                        period_start = d  # Primer día de actividad tras OFF
                    last_activity_day = d
                else:
                    # Neutral (blanco): no inicia período, no lo cierra,
                    # pero NO se cuenta como día de actividad
                    pass
            
            # Última rotación al final del mes (sin breaker después)
            # Solo contar días con actividad REAL
            if period_start is not None and last_activity_day is not None:
                period_days = [dd for dd in range(period_start, last_activity_day + 1)
                               if day_types[dd] == 'activity']
                rotaciones.append(period_days)
            
            # Filtrar períodos >5 días
            long_periods = []
            for rot in rotaciones:
                if len(rot) > 5:
                    long_periods.append({
                        'start': rot[0],
                        'end': rot[-1],
                        'days': len(rot),
                    })
            
            if long_periods:
                crew_periods[cid] = {
                    'nombre': nombre,
                    'cat': cat,
                    'periods': long_periods,
                }
        
        if crew_periods:
            # ── Construir DataFrame para KPIs (una fila por período) ──
            rot_rows_flat = []
            for cid, data in crew_periods.items():
                for p in data['periods']:
                    if _h_mes_rot and _h_anio_rot:
                        fecha_ini = f"{p['start']:02d}/{_h_mes_rot:02d}/{_h_anio_rot}"
                        fecha_fin = f"{p['end']:02d}/{_h_mes_rot:02d}/{_h_anio_rot}"
                    else:
                        fecha_ini = f"Día {p['start']}"
                        fecha_fin = f"Día {p['end']}"
                    rot_rows_flat.append({
                        'ID': cid,
                        'Nombre': data['nombre'],
                        'Cat': data['cat'],
                        'Días': p['days'],
                        'Inicio': fecha_ini,
                        'Fin': fecha_fin,
                        'Dia_Inicio_Num': p['start'],
                        'Dia_Fin_Num': p['end'],
                    })
            
            df_rot = pd.DataFrame(rot_rows_flat)
            
            # ── KPIs ──
            n_trips_afectados = df_rot['ID'].nunique()
            pct_plantilla_rot = (n_trips_afectados / max(total_plantilla, 1)) * 100
            max_dias = df_rot['Días'].max()
            total_rotaciones = len(df_rot)
            n_jc_rot = df_rot[df_rot['Cat'] == 'JC']['ID'].nunique()
            n_tcp_rot = df_rot[df_rot['Cat'] == 'TC']['ID'].nunique()
            
            kpi_cols_rot = st.columns(4)
            kpi_rot_data = [
                ("📊 Plantilla afectada", f"{pct_plantilla_rot:.1f}%", f"{n_trips_afectados} de {total_plantilla}"),
                ("📅 Total períodos >5 días", f"{total_rotaciones}", "períodos detectados"),
                ("👤 JC afectados", f"{n_jc_rot}", "JCS"),
                ("👥 TCP afectados", f"{n_tcp_rot}", "TCPS"),
            ]
            for col, (title, value, subtitle) in zip(kpi_cols_rot, kpi_rot_data):
                with col:
                    st.markdown(f"""
                    <div style="background: white; border-radius: 12px; padding: 16px; text-align: center;
                                box-shadow: 0 2px 8px rgba(0,0,0,0.08); border-top: 3px solid #0D9488;">
                        <div style="font-size: 0.8rem; color: #64748B; font-weight: 600;">{title}</div>
                        <div style="font-size: 1.8rem; font-weight: 800; color: #0D5F5D; margin: 4px 0;">{value}</div>
                        <div style="font-size: 0.75rem; color: #94A3B8;">{subtitle}</div>
                    </div>
                    """, unsafe_allow_html=True)
            
            st.markdown("<br>", unsafe_allow_html=True)
            
            # ── FILTROS INTERACTIVOS ──
            filter_cols_rot = st.columns(2)
            with filter_cols_rot[0]:
                cat_filter_rot = st.selectbox("Filtrar por rol:", ["Todos", "JC", "TCP"], key="rot_cat_filter")
            with filter_cols_rot[1]:
                sort_rot = st.selectbox("Ordenar por:", ["Días (mayor)", "Días (menor)", "ID", "Inicio"], key="rot_sort")
            
            # Filtrar crew_periods según selección
            filtered_crew = {}
            for cid, data in crew_periods.items():
                if cat_filter_rot == "JC" and data['cat'] != 'JC':
                    continue
                if cat_filter_rot == "TCP" and data['cat'] != 'TC':
                    continue
                filtered_crew[cid] = data
            
            # Ordenar
            if sort_rot == "Días (mayor)":
                sorted_crew = sorted(filtered_crew.items(), key=lambda x: max(p['days'] for p in x[1]['periods']), reverse=True)
            elif sort_rot == "Días (menor)":
                sorted_crew = sorted(filtered_crew.items(), key=lambda x: max(p['days'] for p in x[1]['periods']))
            elif sort_rot == "ID":
                sorted_crew = sorted(filtered_crew.items(), key=lambda x: x[0])
            elif sort_rot == "Inicio":
                sorted_crew = sorted(filtered_crew.items(), key=lambda x: x[1]['periods'][0]['start'])
            else:
                sorted_crew = list(filtered_crew.items())
            
            st.markdown("<br>", unsafe_allow_html=True)
            
            # ── TABLA HTML: Una fila por período (vertical), con rowspan para ID/Nombre/Rol ──
            rows_rot_html = []
            for cid, data in sorted_crew:
                cat_badge = "badge-jc" if data['cat'] == 'JC' else "badge-tc"
                rol_icon = '👤' if data['cat'] == 'JC' else '👥'
                rol_text = "JC" if data['cat'] == 'JC' else "TCP"
                n_periods = len(data['periods'])
                
                for idx_p, p in enumerate(data['periods']):
                    if _h_mes_rot and _h_anio_rot:
                        f_ini = f"{p['start']:02d}/{_h_mes_rot:02d}"
                        f_fin = f"{p['end']:02d}/{_h_mes_rot:02d}"
                    else:
                        f_ini = f"Día {p['start']}"
                        f_fin = f"Día {p['end']}"
                    
                    dias_val = p['days']
                    severity_color = '#DC2626' if dias_val >= 8 else ('#EA580C' if dias_val >= 7 else '#0D9488')
                    
                    period_chip = (
                        f"<span style='display:inline-block; background:{severity_color}; color:white; "
                        f"padding:6px 14px; border-radius:20px; font-size:0.88rem; font-weight:700; "
                        f"white-space:nowrap; letter-spacing:0.3px; "
                        f"border:2px solid {severity_color}; "
                        f"box-shadow:0 1px 3px rgba(0,0,0,0.15);'>"
                        f"📅 {f_ini} – {f_fin}</span>"
                    )
                    
                    day_circle = (
                        f"<span style='display:inline-block; background:linear-gradient(135deg, {severity_color}, {severity_color}dd); "
                        f"color:white; min-width:38px; height:38px; line-height:38px; border-radius:50%; "
                        f"font-weight:800; font-size:1.1rem; text-align:center; "
                        f"box-shadow:0 2px 6px {severity_color}55;'>"
                        f"{dias_val}</span>"
                    )
                    
                    if idx_p == 0:
                        # Primera fila del tripulante: incluir ID, Nombre, Rol con rowspan
                        rowspan_attr = f" rowspan='{n_periods}'" if n_periods > 1 else ""
                        row_html = (
                            f"<tr>"
                            f"<td{rowspan_attr} style='vertical-align:middle;'><strong>{cid}</strong></td>"
                            f"<td{rowspan_attr} style='vertical-align:middle; white-space:nowrap;'>{data['nombre']}</td>"
                            f"<td{rowspan_attr} style='vertical-align:middle;'><span class='badge {cat_badge}'>{rol_icon} {rol_text}</span></td>"
                            f"<td>{period_chip}</td>"
                            f"<td style='text-align:center;'>{day_circle}</td>"
                            f"</tr>"
                        )
                    else:
                        # Filas adicionales: solo período y días
                        row_html = (
                            f"<tr>"
                            f"<td>{period_chip}</td>"
                            f"<td style='text-align:center;'>{day_circle}</td>"
                            f"</tr>"
                        )
                    rows_rot_html.append(row_html)
            
            table_rot_html = (
                "<div class='censo-table-wrapper'><table class='censo-table'>"
                "<tr><th>ID</th><th>Nombre</th><th>Rol</th><th>Período</th><th>Días</th></tr>"
                + "".join(rows_rot_html)
                + "</table></div>"
            )
            st.markdown(table_rot_html, unsafe_allow_html=True)
            
            st.markdown("<br>", unsafe_allow_html=True)
            
            # ── Excel download (una fila por período para detalle completo) ──
            try:
                buf_rot = io.BytesIO()
                with pd.ExcelWriter(buf_rot, engine='openpyxl') as writer:
                    df_rot_export = df_rot[['ID', 'Nombre', 'Cat', 'Días', 'Inicio', 'Fin']].copy()
                    df_rot_export.to_excel(writer, index=False, sheet_name='Actividad Continua')
                    ws_rot = writer.sheets['Actividad Continua']
                    from openpyxl.styles import PatternFill, Font, Alignment
                    hf = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
                    hfont = Font(color="FFFFFF", bold=True)
                    for cell in ws_rot[1]:
                        cell.fill = hf
                        cell.font = hfont
                        cell.alignment = Alignment(horizontal='center')
                    for col_cells in ws_rot.columns:
                        max_len = max(len(str(c.value or '')) for c in col_cells) + 2
                        ws_rot.column_dimensions[col_cells[0].column_letter].width = max(max_len, 12)
                
                st.download_button(
                    "📥 Descargar Excel - Actividad Continua >5 Días",
                    data=buf_rot.getvalue(),
                    file_name=f"actividad_continua_{base_procesada}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            except Exception as e:
                st.error(f"Error generando Excel: {e}")
        
        else:
            st.markdown("""
            <div style="text-align: center; padding: 40px; background: #F0FDF4; border-radius: 16px; border: 2px solid #86EFAC;">
                <h2 style="color: #16A34A; margin: 0;">✅ Sin períodos de actividad continua >5 días</h2>
                <p style="color: #15803D; margin-top: 10px;">No se han encontrado tripulantes con más de 5 días consecutivos de actividad.</p>
            </div>
            """, unsafe_allow_html=True)

    
    # ═══════════════════════════════════════════════════════════════
    # TAB 4: CÓDIGOS ESPECÍFICOS - DOBLE DESPLEGABLE
    # ═══════════════════════════════════════════════════════════════
    with t4:
        st.markdown('<p class="section-title">📋 Impacto Desglosado por Código</p>', unsafe_allow_html=True)
        
        if not df_r.empty:
            # All activity codes: detected + full dictionary
            codigos_detectados_actividad = set(c for c in df_r['Codigo'].tolist() 
                                               if c not in ['--', '-', ''] and not (c.isdigit() and len(c) == 4))
            codigos_diccionario = set(CODIGO_DESCRIPCIONES.keys())
            codigos_actividad_todos = sorted(codigos_detectados_actividad | codigos_diccionario)
            
            codigos_vuelos = sorted([c for c in df_r['Codigo'].tolist() 
                                    if c.isdigit() and len(c) == 4])
            
            col_cod, col_vuelo = st.columns(2)
            
            with col_cod:
                if codigos_actividad_todos:
                    cod_opciones = {c: f"{c} — {get_codigo_descripcion(c)}" for c in codigos_actividad_todos}
                    selected_activity = st.selectbox(
                        "🏷️ Filtrar por código de actividad:",
                        options=[None] + codigos_actividad_todos,
                        format_func=lambda c: "-- Seleccionar código --" if c is None else cod_opciones.get(c, c),
                        key="code_activity_filter"
                    )
                else:
                    selected_activity = None
                    st.info("No se detectaron códigos de actividad.")
            
            with col_vuelo:
                if codigos_vuelos:
                    vuelo_opciones = {c: f"VY{c}" for c in codigos_vuelos}
                    selected_vuelo = st.selectbox(
                        "✈️ Filtrar por número de vuelo:",
                        options=[None] + codigos_vuelos,
                        format_func=lambda c: "-- Seleccionar vuelo --" if c is None else vuelo_opciones.get(c, c),
                        key="code_flight_filter"
                    )
                else:
                    selected_vuelo = None
                    st.info("No se detectaron vuelos en las programaciones.")
            
            selected_code = None
            if selected_activity:
                selected_code = selected_activity
            elif selected_vuelo:
                selected_code = selected_vuelo
            
            if selected_code:
                code_matches = df_r[df_r['Codigo'] == selected_code]
                if code_matches.empty:
                    st.info(f"ℹ️ El código **{selected_code}** ({get_codigo_descripcion(selected_code)}) no se ha detectado en las programaciones de este mes.")
                
                if not code_matches.empty:
                    code_data = code_matches.iloc[0]
                    
                    cat_color = get_color_categoria(code_data['Categoria'])
                    pct_trip = round(code_data['Afectados'] / total_plantilla * 100, 2) if total_plantilla > 0 else 0
                    
                    st.markdown("<br>", unsafe_allow_html=True)
                    k1, k2, k3, k4 = st.columns(4)
                    
                    with k1:
                        st.markdown(f"""
                            <div class="kpi-box" style="border-top-color: {cat_color};">
                                <p>TRIPULANTES AFECTADOS</p>
                                <h1>{code_data['Afectados']}</h1>
                                <span style="font-size:0.8rem; color:#64748B; font-weight:600;">({pct_trip}% de la plantilla)</span>
                            </div>
                        """, unsafe_allow_html=True)
                    
                    with k2:
                        st.markdown(f"""
                            <div class="kpi-box accent">
                                <p>% SOBRE PLANTILLA</p>
                                <h1>{pct_trip}%</h1>
                            </div>
                        """, unsafe_allow_html=True)
                    
                    with k3:
                        st.markdown(f"""
                            <div class="kpi-box" style="border-top-color: {cat_color};">
                                <p>TOTAL DÍAS ASIGNADOS</p>
                                <h1>{code_data['Total_Dias']}</h1>
                            </div>
                        """, unsafe_allow_html=True)
                    
                    with k4:
                        desc_texto = code_data['Descripcion']
                        is_vuelo_code = selected_code.isdigit() and len(selected_code) == 4
                        label_code = f"VY{selected_code}" if is_vuelo_code else selected_code
                        st.markdown(f"""
                            <div class="kpi-box">
                                <p>{label_code}</p>
                                <h1 style="font-size:0.95rem; line-height:1.3;">{desc_texto}</h1>
                            </div>
                        """, unsafe_allow_html=True)
                    
                    st.markdown(f"#### 📋 Tripulantes con código **{label_code}** ({code_data['Afectados']} tripulantes · {pct_trip}% de la plantilla):")
                    
                    detail_rows = []
                    for item in code_data['Detalle']:
                        r_icon = "🧑‍✈️" if item['Cat'] == 'JC' else "👥"
                        r_text = "JC" if item['Cat'] == 'JC' else "TCP"
                        detail_rows.append(
                            f"<tr><td><strong>{item['ID']}</strong></td>"
                            f"<td>{item['Nombre']}</td>"
                            f"<td>{r_icon} {r_text}</td>"
                            f"<td><strong>{item['Dias']}</strong></td>"
                            f"<td>{item['Fechas']}</td></tr>"
                        )
                    
                    detail_html = (
                        "<div class='censo-table-wrapper'><table class='censo-table'>"
                        "<tr><th>ID</th><th>Nombre</th><th>Rol</th><th>Nº Días</th><th>Días del mes</th></tr>"
                        + "".join(detail_rows)
                        + "</table></div>"
                    )
                    st.markdown(detail_html, unsafe_allow_html=True)
            
            st.markdown("---")
            
            # RESUMEN de todos los códigos
            st.markdown("#### 📊 Resumen de todos los códigos detectados:")
            
            if not df_r.empty:
                cat_groups = defaultdict(list)
                for _, row in df_r.iterrows():
                    if row['Codigo'] not in ['--', '-', '']:
                        cat_groups[row['Categoria']].append(row)
                
                for cat_id in list(CATEGORIAS_PRINCIPALES.keys()) + ['otros']:
                    if cat_id in cat_groups:
                        if cat_id == 'otros':
                            cat_name = '⚠️ Otros (No catalogados)'
                            cat_icon = '⚠️'
                        else:
                            cat_name = CATEGORIAS_PRINCIPALES[cat_id]['nombre']
                            cat_icon = CATEGORIAS_PRINCIPALES[cat_id]['icono']
                        
                        items = cat_groups[cat_id]
                        n_codes = len(items)
                        total_trip = sum(r['Afectados'] for r in items)
                        
                        with st.expander(f"{cat_icon} {cat_name} — {n_codes} códigos · {total_trip} registros", expanded=False):
                            for r in sorted(items, key=lambda x: -x['Afectados']):
                                pct = round(r['Afectados'] / total_plantilla * 100, 2) if total_plantilla > 0 else 0
                                is_vuelo = r['Codigo'].isdigit() and len(r['Codigo']) == 4
                                label = f"VY{r['Codigo']}" if is_vuelo else r['Codigo']
                                
                                st.markdown(f"""
                                    <div class="code-summary-card" style="border-left-color: {get_color_categoria(cat_id)};">
                                        <div>
                                            <strong style="color:#0D5F5D;font-weight:800;">{label}</strong> — {r['Descripcion'][:45]}
                                        </div>
                                        <div style="display:flex; gap:8px; align-items:center;">
                                            <span class="badge" style="background:#E0F2F1; color:#0D5F5D; font-weight:700;">{r['Afectados']} trip. ({pct}%)</span>
                                            <span class="badge" style="background:#E0F2F1; color:#117F7C; font-weight:700;">{r['Total_Dias']} días</span>
                                        </div>
                                    </div>
                                """, unsafe_allow_html=True)
                                
                                with st.expander(f"👁️ Ver {r['Afectados']} tripulantes de {label}", expanded=False):
                                    crew_rows = []
                                    for item in r['Detalle']:
                                        r_ic = "🧑‍✈️" if item['Cat'] == 'JC' else "👥"
                                        r_tx = "JC" if item['Cat'] == 'JC' else "TCP"
                                        crew_rows.append(
                                            f"<tr><td><strong>{item['ID']}</strong></td>"
                                            f"<td>{item['Nombre']}</td>"
                                            f"<td>{r_ic} {r_tx}</td>"
                                            f"<td>{item['Dias']}</td>"
                                            f"<td>{item['Fechas']}</td></tr>"
                                        )
                                    crew_html = (
                                        "<div class='censo-table-wrapper'><table class='censo-table'>"
                                        "<tr><th>ID</th><th>Nombre</th><th>Rol</th><th>Días</th><th>Fechas</th></tr>"
                                        + "".join(crew_rows)
                                        + "</table></div>"
                                    )
                                    st.markdown(crew_html, unsafe_allow_html=True)
    
    # ═══════════════════════════════════════════════════════════════
    # TAB 5: CALENDARIO RESUMEN DIARIO - BARRA MEJORADA
    # ═══════════════════════════════════════════════════════════════
    with t5:
        st.markdown('<p class="section-title" style="font-size: 1.5rem;">📅 Calendario Resumen Diario</p>', unsafe_allow_html=True)
        st.markdown(f"*Distribución de actividades por día para toda la plantilla ({total_plantilla} tripulantes)*")
        
        import calendar as cal_mod
        # Usar mes/año de la cabecera del PDF si está disponible
        _h_mes_cal = st.session_state.get('header_mes')
        _h_anio_cal = st.session_state.get('header_anio')
        if _h_mes_cal and _h_anio_cal:
            mes_cal = _h_mes_cal
            anio_cal = _h_anio_cal
        else:
            now = datetime.now()
            anio_cal = now.year
            mes_cal = now.month - 1 if now.month > 1 else 12
            if mes_cal == 0:
                mes_cal = 12
                anio_cal -= 1
            for m_try in range(1, 13):
                if cal_mod.monthrange(anio_cal, m_try)[1] == dias_mes:
                    mes_cal = m_try
                    break
        
        cat_nombres_cortos = {
            'vuelo': ('✈️', 'Vuelos'),
            'guardias': ('⏰', 'Guardias'),
            'libres': ('🟢', 'Libres'),
            'formacion': ('📚', 'Formación'),
            'oficina': ('🏢', 'Oficina'),
            'medico': ('🏥', 'Médico'),
            'permisos': ('👤', 'Permisos'),
            'vacaciones': ('🌴', 'Vacaciones'),
            'parttime': ('⏸️', 'Part-Time'),
            'incidencias': ('⚠️', 'Incidencias'),
            'otros': ('📋', 'Otros')
        }
        
        dias_semana_cal = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
        
        try:
            from datetime import date
            primer_dia_cal = date(anio_cal, mes_cal, 1)
            offset_cal = primer_dia_cal.weekday()
        except:
            offset_cal = 0
        
        # Build calendar as single HTML grid - BARRA MEJORADA
        grid_html = '<div style="display:grid;grid-template-columns:repeat(7,1fr);gap:6px;max-width:100%;margin:0 auto;">'
        
        # Header row - MEJORADO: más grande, atractivo y con colores vibrantes
        for d_idx, d_name in enumerate(dias_semana_cal):
            is_weekend = d_idx >= 5
            bg_color = 'linear-gradient(135deg, #117F7C 0%, #44BABC 100%)' if is_weekend else 'linear-gradient(135deg, #0D5F5D 0%, #117F7C 100%)'
            grid_html += f'''<div style="text-align:center;font-weight:900;color:white;font-size:0.95rem;padding:14px 6px;
                            background:{bg_color};border-radius:12px;box-shadow:0 4px 12px rgba(17,127,124,0.35);
                            text-transform:uppercase;letter-spacing:0.5px;">{d_name}</div>'''
        
        # Empty cells for offset
        for _ in range(offset_cal):
            grid_html += '<div style="min-height:120px;"></div>'
        
        for dia in range(1, dias_mes + 1):
            col_idx = (dia - 1 + offset_cal) % 7
            is_weekend = col_idx >= 5
            weekend_class = "cal-weekend" if is_weekend else ""
            
            dia_data = datos_por_dia.get(dia, {})
            
            items_html = ""
            for cat_id in ['vuelo', 'guardias', 'libres', 'formacion', 'medico', 'vacaciones', 'permisos', 'parttime', 'oficina', 'incidencias', 'otros']:
                count = dia_data.get(cat_id, 0)
                if count > 0:
                    icono, nombre = cat_nombres_cortos.get(cat_id, ('📋', cat_id))
                    items_html += f'<div class="cal-item"><span>{icono} {nombre}</span><strong>{count}</strong></div>'
            
            no_data_div = '<div style="color:#94A3B8;font-size:0.65rem;">Sin datos</div>'
            content = items_html if items_html else no_data_div
            grid_html += f'<div class="cal-day {weekend_class}"><div class="cal-day-num">{dia}</div>{content}</div>'
        
        # Fill remaining cells
        total_cells = offset_cal + dias_mes
        remainder = total_cells % 7
        if remainder > 0:
            for _ in range(7 - remainder):
                grid_html += '<div style="min-height:120px;"></div>'
        
        grid_html += '</div>'
        st.markdown(grid_html, unsafe_allow_html=True)

    
    # === SECCIÓN 6: FUNCIONES DE EXPORTACIÓN ===
    # ═══════════════════════════════════════════════════════════════
    # TAB 6: BASE DE DATOS / EXCEL
    # ═══════════════════════════════════════════════════════════════
    with t6:
        st.markdown('<p class="section-title">🗄️ Base de Datos Completa</p>', unsafe_allow_html=True)
        
        df_export = df_c[['ID', 'Nombre', 'Cat', 'Block_Inicial', 'Block', 'Duty', 'Estabilidad', 'Cambios']].copy()
        df_export = df_export.rename(columns={
            'Cat': 'Rol',
            'Block_Inicial': 'Block Inicial',
            'Block': 'Block Final'
        })
        
        # Añadir columna de Diferencia
        df_export['Horas_Inicial_Num'] = df_c['Block_Inicial'].apply(parse_hours)
        df_export['Horas_Final_Num'] = df_c['Block'].apply(parse_hours)
        df_export['Diferencia'] = df_export['Horas_Final_Num'] - df_export['Horas_Inicial_Num']
        df_export['Diferencia'] = df_export['Diferencia'].apply(format_diff_hm)
        
        # Columnas especiales: FTG, UNFIT, Bajas
        _ftg_codes_xl = ['FTG']
        _unfit_codes_xl = ['UNFT', 'UNFD']
        _baja_codes_xl = ['MED', 'ASEP', 'SICK', 'SICD', 'LSIC', 'LSCK', 'NJSK']
        
        def _count_days_for_codes(row_cd, codes_list):
            total = 0
            for c in codes_list:
                total += len(row_cd.get(c, []))
            return total
        
        df_export['FTG_Dias'] = [_count_days_for_codes(row['code_days'], _ftg_codes_xl) for _, row in df_c.iterrows()]
        df_export['UNFIT_Dias'] = [_count_days_for_codes(row['code_days'], _unfit_codes_xl) for _, row in df_c.iterrows()]
        df_export['Bajas_Dias'] = [_count_days_for_codes(row['code_days'], _baja_codes_xl) for _, row in df_c.iterrows()]
        
        codigos_no_vuelo = sorted([c for c in all_codes if not (c.isdigit() and len(c) == 4) and c not in ['--', '-', '']])
        
        for code in codigos_no_vuelo:
            col_name = code
            dias_por_trip = []
            for _, row in df_c.iterrows():
                n_dias = len(row['code_days'].get(code, []))
                dias_por_trip.append(n_dias)
            df_export[col_name] = dias_por_trip
        
        # Calcular totales incluyendo Diferencia
        total_inicial = sum(parse_hours(h) for h in df_c["Block_Inicial"])
        total_final = sum(parse_hours(h) for h in df_c["Block"])
        total_diferencia = total_final - total_inicial
        diff_text_total = format_diff_hm(total_diferencia)
        
        totals = {'ID': 'TOTAL', 'Nombre': f'{total_plantilla} tripulantes', 'Rol': '--',
                  'Block Inicial': f'{total_inicial:.0f}h',
                  'Block Final': f'{total_final:.0f}h',
                  'Diferencia': diff_text_total,
                  'Duty': f'{sum(parse_hours(h) for h in df_c["Duty"]):.0f}h',
                  'Cambios': df_export['Cambios'].sum() if 'Cambios' in df_export.columns else 0,
                  'Estabilidad': round(df_c['Estabilidad'].mean(), 1)}
        totals['FTG_Dias'] = df_export['FTG_Dias'].sum()
        totals['UNFIT_Dias'] = df_export['UNFIT_Dias'].sum()
        totals['Bajas_Dias'] = df_export['Bajas_Dias'].sum()
        for code in codigos_no_vuelo:
            totals[code] = df_export[code].sum()
        
        # Limpiar columnas temporales antes de exportar
        df_export_clean = df_export.drop(columns=['Horas_Inicial_Num', 'Horas_Final_Num'], errors='ignore')
        df_export_with_totals = pd.concat([df_export_clean, pd.DataFrame([totals])], ignore_index=True)
        
        cols_preview = ['ID', 'Nombre', 'Rol', 'Block Inicial', 'Block Final', 'Diferencia', 'Duty', 'Cambios', 'Estabilidad'] + codigos_no_vuelo[:10]
        st.dataframe(
            df_export_clean[cols_preview].head(50),
            use_container_width=True,
            column_config={
                "Nombre": st.column_config.TextColumn("Nombre", width="large"),
            }
        )
        
        if len(codigos_no_vuelo) > 10:
            st.info(f"📊 Se muestran las primeras 10 columnas de códigos. El Excel descargable incluye los {len(codigos_no_vuelo)} códigos detectados.")
        
        # Añadir columnas numéricas Horas BLOCK al export
        df_export_clean['Horas_BLOCK_Inicial'] = df_c['Block_Inicial'].apply(parse_hours)
        df_export_clean['Horas_BLOCK_Final'] = df_c['Block'].apply(parse_hours)
        
        st.markdown("#### 📥 Exportar datos (3 opciones):")
        col_exp1, col_exp2, col_exp3 = st.columns(3)
        
        with col_exp1:
            csv = df_export_with_totals.to_csv(index=False).encode('utf-8')
            st.download_button(
                "📄 Descargar CSV",
                csv,
                f"programaciones_{base_procesada}.csv",
                "text/csv",
                use_container_width=True
            )
        
        with col_exp2:
            try:
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df_export_with_totals.to_excel(writer, index=False, sheet_name='Programaciones')
                    
                    # --- Formatear hoja Programaciones ---
                    ws_prog = writer.sheets['Programaciones']
                    header_fill_basic = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
                    header_font_basic = Font(color="FFFFFF", bold=True, size=11)
                    totals_fill_basic = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
                    totals_font_basic = Font(bold=True, size=11)
                    
                    # Formatear fila de encabezados (fila 1)
                    for cell in ws_prog[1]:
                        cell.fill = header_fill_basic
                        cell.font = header_font_basic
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Formatear fila de totales (última fila)
                    last_row = ws_prog.max_row
                    for cell in ws_prog[last_row]:
                        cell.fill = totals_fill_basic
                        cell.font = totals_font_basic
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Ajustar ancho de columnas
                    for col_cells in ws_prog.columns:
                        max_len = max((len(str(cell.value or '')) for cell in col_cells), default=8)
                        ws_prog.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 25)
                    
                    if not df_r.empty:
                        df_codigos_resumen = df_r[['Codigo', 'Descripcion', 'Afectados', 'Total_Dias', 'Categoria']].copy()
                        df_codigos_resumen = df_codigos_resumen.rename(columns={
                            'Codigo': 'Código',
                            'Descripcion': 'Descripción',
                            'Afectados': 'Tripulantes Afectados',
                            'Total_Dias': 'Días Totales',
                            'Categoria': 'Categoría'
                        })
                        df_codigos_resumen.to_excel(writer, index=False, sheet_name='Resumen Códigos')
                        
                        # Formatear hoja Resumen Códigos
                        ws_cod = writer.sheets['Resumen Códigos']
                        for cell in ws_cod[1]:
                            cell.fill = header_fill_basic
                            cell.font = header_font_basic
                            cell.alignment = Alignment(horizontal='center')
                        for col_cells in ws_cod.columns:
                            max_len = max((len(str(cell.value or '')) for cell in col_cells), default=8)
                            ws_cod.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 25)
                    
                    # --- Hoja Block Horas Medias ---
                    block_summary_data = {
                        'Concepto': ['Block Medio JC (Inicial)', 'Block Medio JC (Final)',
                                     'Block Medio TCP (Inicial)', 'Block Medio TCP (Final)'],
                        'Horas (HH:MM)': [
                            k.get('block_medio_jc_ini_fmt', '0:00'),
                            k.get('block_medio_jc_fin_fmt', '0:00'),
                            k.get('block_medio_tc_ini_fmt', '0:00'),
                            k.get('block_medio_tc_fin_fmt', '0:00'),
                        ],
                        'Horas (decimal)': [
                            round(k.get('block_medio_jc_ini', 0), 2),
                            round(k.get('block_medio_jc_fin', 0), 2),
                            round(k.get('block_medio_tc_ini', 0), 2),
                            round(k.get('block_medio_tc_fin', 0), 2),
                        ]
                    }
                    df_block_summary = pd.DataFrame(block_summary_data)
                    df_block_summary.to_excel(writer, index=False, sheet_name='Block Horas Medias')
                    ws_bh = writer.sheets['Block Horas Medias']
                    for cell in ws_bh[1]:
                        cell.fill = header_fill_basic
                        cell.font = header_font_basic
                        cell.alignment = Alignment(horizontal='center')
                    for col_cells in ws_bh.columns:
                        max_len = max((len(str(cell.value or '')) for cell in col_cells), default=8)
                        ws_bh.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 30)
                
                excel_data = buffer.getvalue()
                st.download_button(
                    "📊 Excel Básico",
                    data=excel_data,
                    file_name=f"programaciones_{base_procesada}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            except Exception as e:
                st.error(f"Error generando Excel: {e}")
        
        with col_exp3:
            # TERCERA OPCIÓN: Excel con gráficos profesionales
            try:
                buffer_charts = io.BytesIO()
                wb = openpyxl.Workbook()
                
                # Hoja 1: Datos principales
                ws_data = wb.active
                ws_data.title = "Programaciones"
                
                # Estilos
                header_fill = PatternFill(start_color="0D5F5D", end_color="117F7C", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=11)
                border_style = Border(
                    left=Side(style='thin', color='B2DFDB'),
                    right=Side(style='thin', color='B2DFDB'),
                    top=Side(style='thin', color='B2DFDB'),
                    bottom=Side(style='thin', color='B2DFDB')
                )
                
                # Escribir cabeceras (incluyendo Diferencia)
                headers = ['ID', 'Nombre', 'Rol', 'Block Inicial', 'Block Final', 'Diferencia', 'Duty', 'Estabilidad']
                for col, header in enumerate(headers, 1):
                    cell = ws_data.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Escribir datos (incluyendo Diferencia)
                for row_idx, (_, row) in enumerate(df_c.iterrows(), 2):
                    horas_inicial = parse_hours(row['Block_Inicial'])
                    horas_final = parse_hours(row['Block'])
                    diferencia = horas_final - horas_inicial
                    diff_text = format_diff_hm(diferencia)
                    
                    ws_data.cell(row=row_idx, column=1, value=row['ID'])
                    ws_data.cell(row=row_idx, column=2, value=row['Nombre'])
                    ws_data.cell(row=row_idx, column=3, value=row['Cat'])
                    ws_data.cell(row=row_idx, column=4, value=row['Block_Inicial'])
                    ws_data.cell(row=row_idx, column=5, value=row['Block'])
                    ws_data.cell(row=row_idx, column=6, value=diff_text)
                    ws_data.cell(row=row_idx, column=7, value=row['Duty'])
                    ws_data.cell(row=row_idx, column=8, value=row['Estabilidad'])
                
                # Ajustar anchos de columna
                for col in range(1, 9):
                    ws_data.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                
                # Hoja 2: Resumen con gráficos
                ws_charts = wb.create_sheet("Gráficos Resumen")
                
                # Datos para gráfico de categorías
                ws_charts.cell(row=1, column=1, value="Categoría").font = header_font
                ws_charts.cell(row=1, column=1).fill = header_fill
                ws_charts.cell(row=1, column=2, value="Tripulantes").font = header_font
                ws_charts.cell(row=1, column=2).fill = header_fill
                ws_charts.cell(row=1, column=3, value="% Plantilla").font = header_font
                ws_charts.cell(row=1, column=3).fill = header_fill
                
                row_num = 2
                for cat_id, cat_info in CATEGORIAS_PRINCIPALES.items():
                    cat_data = categorias_resumen.get(cat_id, {'tripulantes': 0})
                    num_trip = cat_data.get('tripulantes', 0)
                    if num_trip > 0:
                        pct = round((num_trip / total_plantilla * 100), 1) if total_plantilla > 0 else 0
                        ws_charts.cell(row=row_num, column=1, value=cat_info['nombre'])
                        ws_charts.cell(row=row_num, column=2, value=num_trip)
                        ws_charts.cell(row=row_num, column=3, value=pct)
                        row_num += 1
                
                # Crear gráfico de barras
                if row_num > 2:
                    chart = BarChart()
                    chart.type = "col"
                    chart.style = 10
                    chart.title = "Distribución por Categoría"
                    chart.y_axis.title = "Tripulantes"
                    chart.x_axis.title = "Categoría"
                    
                    data = Reference(ws_charts, min_col=2, min_row=1, max_row=row_num-1, max_col=2)
                    cats = Reference(ws_charts, min_col=1, min_row=2, max_row=row_num-1)
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
                    chart.shape = 4
                    ws_charts.add_chart(chart, "E2")
                    
                    # Gráfico pie JC/TCP
                    ws_charts.cell(row=row_num + 2, column=1, value="Rol").font = header_font
                    ws_charts.cell(row=row_num + 2, column=1).fill = header_fill
                    ws_charts.cell(row=row_num + 2, column=2, value="Cantidad").font = header_font
                    ws_charts.cell(row=row_num + 2, column=2).fill = header_fill
                    ws_charts.cell(row=row_num + 3, column=1, value="JC")
                    ws_charts.cell(row=row_num + 3, column=2, value=k['jc'])
                    ws_charts.cell(row=row_num + 4, column=1, value="TCP")
                    ws_charts.cell(row=row_num + 4, column=2, value=k['tc'])
                    
                    pie = PieChart()
                    pie.title = "Distribución JC / TCP"
                    labels = Reference(ws_charts, min_col=1, min_row=row_num+3, max_row=row_num+4)
                    data_pie = Reference(ws_charts, min_col=2, min_row=row_num+2, max_row=row_num+4)
                    pie.add_data(data_pie, titles_from_data=True)
                    pie.set_categories(labels)
                    pie.dataLabels = DataLabelList()
                    pie.dataLabels.showPercent = True
                    pie.dataLabels.showVal = True
                    ws_charts.add_chart(pie, "E18")
                
                wb.save(buffer_charts)
                excel_charts_data = buffer_charts.getvalue()
                
                st.download_button(
                    "📈 Excel + Gráficos",
                    data=excel_charts_data,
                    file_name=f"programaciones_{base_procesada}_graficos.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            except Exception as e:
                st.error(f"Error generando Excel con gráficos: {e}")


    # El panel de admin ahora es una sección separada (admin_panel), no una pestaña

# Footer
st.markdown("---")
st.markdown('<p style="text-align:center;color:#94A3B8;font-size:0.8rem;font-weight:500;">✈️ Herramientas Sindicales © 2026</p>', unsafe_allow_html=True)